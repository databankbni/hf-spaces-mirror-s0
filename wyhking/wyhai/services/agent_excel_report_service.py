from __future__ import annotations

import csv
import math
import json
import os
import re
import shutil
import subprocess
import tempfile
from datetime import datetime
from functools import lru_cache
from io import BytesIO
from pathlib import Path
from typing import Any, Iterable
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


BLUE = "2563EB"
NAVY = "0F172A"
PALE_BLUE = "EFF6FF"
PALE_GREEN = "ECFDF5"
PALE_RED = "FEF2F2"
PALE_YELLOW = "FFFBEB"
WHITE = "FFFFFF"
MUTED = "64748B"
GRID = "D8E2F0"
SHANGHAI_TZ = ZoneInfo("Asia/Shanghai")
ROOT = Path(__file__).resolve().parents[1]
SELECTION_ARTIFACT_BUILDER = ROOT / "scripts" / "build_selection_report_artifact.mjs"
LEADERSHIP_40PCT_COHORT = ROOT / "data" / "selection" / "leadership_market_dsi_40pct_20260714.csv"
LEADERSHIP_40PCT_METADATA = ROOT / "data" / "selection" / "leadership_market_dsi_40pct_20260714.json"


def build_selection_report(payload: dict[str, Any]) -> tuple[BytesIO, str]:
    result = _dict(payload.get("result"))
    metrics = _dict(result.get("metrics"))
    scope = _dict(metrics.get("scope"))
    node = os.environ.get("NODE_EXECUTABLE") or shutil.which("node")
    builder_payload = dict(payload)
    builder_payload["strategy_evidence"] = _selection_strategy_evidence()
    is_avoid_export = _is_avoid_selection_export(result, scope)
    use_leadership_snapshot = not is_avoid_export and LEADERSHIP_40PCT_COHORT.is_file() and LEADERSHIP_40PCT_METADATA.is_file()
    output: BytesIO | None = None
    if node and SELECTION_ARTIFACT_BUILDER.is_file() and not use_leadership_snapshot:
        temp_root = ROOT / "tmp" / "selection_reports"
        temp_root.mkdir(parents=True, exist_ok=True)
        with tempfile.TemporaryDirectory(prefix="selection_", dir=temp_root) as temp_dir:
            temp_path = Path(temp_dir)
            input_path = temp_path / "payload.json"
            output_path = temp_path / "selection_report.xlsx"
            input_path.write_text(json.dumps(builder_payload, ensure_ascii=False, default=str), encoding="utf-8")
            completed = subprocess.run(
                [node, str(SELECTION_ARTIFACT_BUILDER), str(input_path), str(output_path)],
                cwd=ROOT,
                check=False,
                capture_output=True,
                text=True,
                timeout=90,
            )
            if completed.returncode == 0 and output_path.is_file():
                output = BytesIO(output_path.read_bytes())
    if output is None:
        output = _build_selection_report_portable(builder_payload)
    city = _clean_filename(scope.get("city") or "全国")
    stamp = datetime.now(SHANGHAI_TZ).strftime("%Y%m%d_%H%M%S")
    prefix = "AI懂车价_选品40%底池" if use_leadership_snapshot else "AI懂车价_选品报告"
    return output, f"{prefix}_{city}_{stamp}.xlsx"


def _build_selection_report_portable(payload: dict[str, Any]) -> BytesIO:
    """Server-safe equivalent used when the hosted container has no Node runtime."""
    result = _dict(payload.get("result"))
    metrics = _dict(result.get("metrics"))
    scope = _dict(metrics.get("scope"))
    is_avoid_export = _is_avoid_selection_export(result, scope)
    recommended = _strict_selection_rows(
        metrics.get("export_recommendations") or metrics.get("strict_recommendations") or [],
        kind="recommend",
    )
    avoid = _strict_selection_rows(
        metrics.get("export_avoid_items") or metrics.get("strict_avoid_items") or [],
        kind="avoid",
    )
    leadership_rows = list(_leadership_40pct_rows()) if not is_avoid_export else []
    leadership_meta = _leadership_40pct_metadata() if leadership_rows else {}
    qualification_projection = [
        _dict(item)
        for item in (metrics.get("export_qualification_items") or [])
        if isinstance(item, dict)
    ]
    export_rows = avoid if is_avoid_export else leadership_rows or recommended
    evidence = _dict(payload.get("strategy_evidence")) or _selection_strategy_evidence()

    workbook = Workbook()
    summary = workbook.active
    if leadership_meta and not is_avoid_export:
        summary.title = "40%底池摘要"
        grain_name = "全国车系×年款" if str(scope.get("city") or "全国") in {"全国", "全网", ""} else "城市×车系"
        projection_sheet = workbook.create_sheet(grain_name)
        detail_sheet = workbook.create_sheet("40%底池车辆")
        evidence_sheet = workbook.create_sheet("40%口径审计")
        _build_leadership_summary_portable(summary, leadership_meta, scope, len(qualification_projection))
        _build_qualification_projection_portable(projection_sheet, qualification_projection, scope)
        _build_leadership_detail_portable(detail_sheet, export_rows)
        _build_leadership_evidence_portable(evidence_sheet, leadership_meta)
    else:
        summary.title = "选品摘要"
        detail_sheet = workbook.create_sheet("避免收明细" if is_avoid_export else "推荐收明细")
        _build_selection_summary_portable(summary, export_rows, scope, evidence, is_avoid_export=is_avoid_export)
        _build_selection_detail_portable(
            detail_sheet,
            export_rows,
            "AvoidDetailTable" if is_avoid_export else "RecommendationDetailTable",
            avoid_sheet=is_avoid_export,
        )
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def _is_avoid_selection_export(result: dict[str, Any], scope: dict[str, Any]) -> bool:
    target = str(scope.get("selection_target") or "")
    return target in {"risk_series", "risk"} or "避免收" in str(result.get("title") or "")


@lru_cache(maxsize=1)
def _leadership_40pct_metadata() -> dict[str, Any]:
    if not LEADERSHIP_40PCT_METADATA.is_file():
        return {}
    try:
        loaded = json.loads(LEADERSHIP_40PCT_METADATA.read_text(encoding="utf-8"))
        return loaded if isinstance(loaded, dict) else {}
    except Exception:
        return {}


@lru_cache(maxsize=1)
def _leadership_40pct_rows() -> tuple[dict[str, Any], ...]:
    if not LEADERSHIP_40PCT_COHORT.is_file():
        return ()
    try:
        with LEADERSHIP_40PCT_COHORT.open("r", encoding="utf-8-sig", newline="") as handle:
            rows = tuple(dict(row) for row in csv.DictReader(handle))
        expected = int(_leadership_40pct_metadata().get("qualified_unique_vehicle_count") or 0)
        if expected and len(rows) != expected:
            return ()
        return rows
    except Exception:
        return ()


def _build_leadership_summary_portable(
    ws,
    metadata: dict[str, Any],
    scope: dict[str, Any],
    projection_count: int,
) -> None:
    _base_sheet(ws)
    ws.merge_cells("A1:H1")
    ws["A1"] = "AI懂车价 · 统一40%选品底池"
    _title(ws["A1"])
    ws.merge_cells("A3:H4")
    city = str(scope.get("city") or "全国")
    grain = "全国按车系×年款" if city in {"全国", "全网", ""} else f"{city}按城市×车系"
    baseline_count = int(metadata.get("baseline_unique_vehicle_count") or 0)
    qualified_count = int(metadata.get("qualified_unique_vehicle_count") or 0)
    qualification_rate = float(metadata.get("qualification_rate") or 0)
    ws["A3"] = (
        "本报告严格使用已确认的‘四类行情 AND DSI供不应求/供需平衡’准入底池："
        f"{qualified_count:,} / {baseline_count:,} 辆，覆盖率{qualification_rate:.2%}。{grain}展示；"
        "页面机会分只用于底池内排序和跟进优先级，不改变入选资格。"
    )
    ws["A3"].alignment = Alignment(wrap_text=True, vertical="center")
    ws["A3"].fill = PatternFill("solid", fgColor=PALE_BLUE)
    conversion = _dict(metadata.get("true_buyer_conversion"))
    sales = _dict(metadata.get("sales_conversion_45d"))
    cards = [
        ("统一底池范围", "全国"),
        ("全量唯一车辆", metadata.get("baseline_unique_vehicle_count")),
        ("准入唯一车辆", metadata.get("qualified_unique_vehicle_count")),
        ("底池覆盖率", metadata.get("qualification_rate")),
        ("当前榜单粒度", grain),
        ("当前投影行数", projection_count),
        ("准入池收车转化率", conversion.get("qualified")),
        ("准入池成熟售车转化率", sales.get("qualified")),
    ]
    for index, (label, value) in enumerate(cards):
        row = 6 + (index // 4) * 2
        col = 1 + (index % 4) * 2
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)
        ws.merge_cells(start_row=row + 1, start_column=col, end_row=row + 1, end_column=col + 1)
        ws.cell(row, col, label)
        ws.cell(row + 1, col, value)
        ws.cell(row, col).font = Font(size=9, bold=True, color=MUTED)
        ws.cell(row + 1, col).font = Font(size=14, bold=True, color=NAVY)
        ws.cell(row, col).fill = ws.cell(row + 1, col).fill = PatternFill("solid", fgColor="F8FAFC")
        if index in {3, 6, 7}:
            ws.cell(row + 1, col).number_format = "0.00%"
    ws.merge_cells("A12:H12")
    ws["A12"] = "口径边界"
    _section(ws["A12"])
    guidance = [
        "1. 入选资格只看四类行情与DSI供需标签，统一底池覆盖率固定为40.16%。",
        "2. 内部经营结果只用于回测验证和底池内跟进排序，不反向删除40%准入车辆。",
        "3. 全国榜按车系×年款聚合；城市榜按城市×车系聚合，两者投影自同一底池。",
        "4. 用户下载保留完整准入车辆与行情/DSI判定字段，不输出内部过程计数字段。",
    ]
    for row, value in enumerate(guidance, start=13):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        ws.cell(row, 1, value).alignment = Alignment(wrap_text=True)
    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].width = 16
    ws.freeze_panes = "A2"


def _build_qualification_projection_portable(ws, rows: list[dict[str, Any]], scope: dict[str, Any]) -> None:
    _base_sheet(ws)
    headers = [
        "排名", "城市", "品牌", "车系", "年款", "能源", "车身", "行情状态",
        "DSI供需", "准入依据", "底池内排序分", "跟进说明",
    ]
    ws.append(headers)
    city = str(scope.get("city") or "全国")
    is_national = city in {"全国", "全网", ""}
    for index, raw in enumerate(rows, start=1):
        item = _dict(raw)
        dsi = _dict(item.get("dsi_signal"))
        market_state = _safe_text(item.get("market_category_label") or item.get("market_category"))
        dsi_label = _safe_text(dsi.get("label") or item.get("dsi_label"))
        ws.append([
            index,
            "全国" if is_national else _safe_text(item.get("city") or city),
            _safe_text(item.get("brand")),
            _safe_text(item.get("series")),
            f"{item.get('model_year')}款" if is_national and item.get("model_year") else "-",
            _safe_text(item.get("energy_type")),
            _safe_text(item.get("body_type")),
            market_state,
            dsi_label,
            f"{market_state} + DSI{dsi_label}",
            _number(item.get("final_opportunity_score") or item.get("opportunity_score")),
            "仅决定跟进顺序，不改变40%底池入选资格",
        ])
    if ws.max_row == 1:
        ws.append(["-", city, "", "当前筛选下暂无可投影车系"] + [""] * 8)
    _style_table(ws, headers, "QualificationProjectionTable", ws.max_row)
    widths = [8, 12, 14, 24, 12, 12, 12, 16, 14, 30, 14, 36]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=cell.column in {10, 12})
    ws.freeze_panes = "A2"


def _build_leadership_detail_portable(ws, rows: list[dict[str, Any]]) -> None:
    _base_sheet(ws)
    headers = [
        "序号", "唯一货品ID", "城市", "品牌", "车系", "能源类型", "车身类型", "价格带",
        "行情状态", "状态匹配层级", "DSI标签", "DSI车款数", "DSI一致度", "判定依据",
    ]
    ws.append(headers)
    for index, item in enumerate(rows, start=1):
        ws.append([
            index,
            _safe_text(item.get("unique_product_id")),
            _safe_text(item.get("city")),
            _safe_text(item.get("brand")),
            _safe_text(item.get("series")),
            _safe_text(item.get("energy_type")),
            _safe_text(item.get("body_type")),
            _safe_text(item.get("price_band")),
            _safe_text(item.get("market_state")),
            _safe_text(item.get("market_match_scope")),
            _safe_text(item.get("dsi_label")),
            _number(item.get("dsi_model_count")),
            _ratio(item.get("dsi_consistency")),
            _safe_text(item.get("qualification_reason")),
        ])
    _style_table(ws, headers, "LeadershipCohortTable", ws.max_row)
    for row in range(2, ws.max_row + 1):
        ws.cell(row, 13).number_format = "0.00%"
    widths = [8, 16, 12, 14, 24, 12, 12, 14, 16, 28, 14, 14, 14, 32]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=cell.column in {10, 14})
    ws.freeze_panes = "A2"


def _build_leadership_evidence_portable(ws, metadata: dict[str, Any]) -> None:
    _base_sheet(ws)
    ws.merge_cells("A1:D1")
    ws["A1"] = "统一40%底池 · 口径审计"
    _title(ws["A1"])
    ws.append(["口径", "全量基线", "40%准入底池", "变化 / 说明"])
    conversion = _dict(metadata.get("true_buyer_conversion"))
    sales = _dict(metadata.get("sales_conversion_45d"))
    profit = _dict(metadata.get("avg_profit"))
    turnover = _dict(metadata.get("avg_turnover_days"))
    loss = _dict(metadata.get("loss_rate"))
    conversion_lift = _relative_change(conversion.get("relative_lift"))
    sales_lift = _relative_change(sales.get("relative_lift"))
    profit_lift = _relative_change(profit.get("relative_lift"))
    turnover_change = _relative_change(turnover.get("relative_ratio"))
    baseline_loss = _number(loss.get("baseline"))
    qualified_loss = _number(loss.get("qualified"))
    loss_change = (
        f"下降{(baseline_loss - qualified_loss) * 100:.2f}个百分点"
        if baseline_loss is not None and qualified_loss is not None
        else "仅回测验证"
    )
    rows = [
        ["唯一车辆", metadata.get("baseline_unique_vehicle_count"), metadata.get("qualified_unique_vehicle_count"), metadata.get("qualification_rate")],
        ["买手报价→B2C收车转化率", conversion.get("baseline"), conversion.get("qualified"), conversion_lift],
        ["成熟上架售出率", sales.get("baseline"), sales.get("qualified"), sales_lift],
        ["单车平均毛利", profit.get("baseline"), profit.get("qualified"), profit_lift],
        ["平均周转天数", turnover.get("baseline"), turnover.get("qualified"), turnover_change],
        ["亏损率", loss.get("baseline"), loss.get("qualified"), loss_change],
        ["准入规则", "全部合法车辆", metadata.get("qualification_rule"), "内部经营结果不反向筛选"],
        ["榜单粒度", "-", "城市×车系 / 全国车系×年款", "同一40%底池的不同投影"],
    ]
    for row in rows:
        ws.append(row)
    _format_header(ws, 2, 1, 4)
    ws["B3"].number_format = ws["C3"].number_format = "#,##0"
    ws["D3"].number_format = "0.00%"
    for row in (4, 5):
        for col in range(2, 5):
            ws.cell(row, col).number_format = "0.00%"
    ws["B6"].number_format = ws["C6"].number_format = '#,##0.00"元"'
    ws["D6"].number_format = "0.00%"
    ws["B7"].number_format = ws["C7"].number_format = '0.00"天"'
    ws["D7"].number_format = "0.00%"
    ws["B8"].number_format = ws["C8"].number_format = "0.00%"
    for row in range(3, ws.max_row + 1):
        for cell in ws[row]:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    widths = [28, 24, 34, 28]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width
    ws.freeze_panes = "A3"


def _strict_selection_rows(source_rows: Iterable[Any], *, kind: str) -> list[dict[str, Any]]:
    seen: set[tuple[str, str, str, str]] = set()
    rows: list[dict[str, Any]] = []
    for raw in source_rows:
        item = _dict(raw)
        if kind == "recommend" and not item.get("business_recommend"):
            continue
        if kind == "avoid" and not item.get("business_avoid"):
            continue
        dsi = _safe_text(_dict(item.get("dsi_signal")).get("label") or item.get("dsi_label"))
        if kind == "recommend" and dsi not in {"供不应求", "供需平衡"}:
            continue
        key = (
            _safe_text(item.get("city")),
            _safe_text(item.get("brand")),
            _safe_text(item.get("series")),
            _safe_text(item.get("model_year")),
        )
        if key in seen:
            continue
        seen.add(key)
        rows.append(item)
    return rows


def _pct_text(value: Any) -> str:
    parsed = _number(value)
    return "暂无" if parsed is None else f"{parsed * 100:.1f}%"


def _days_text(value: Any) -> str:
    parsed = _number(value)
    return "暂无" if parsed is None else f"{parsed:.1f}天"


def _money_text(value: Any) -> str:
    parsed = _number(value)
    return "暂无" if parsed is None else f"{parsed:,.0f}元"


def _build_selection_summary_portable(ws, export_rows, scope, evidence, *, is_avoid_export: bool) -> None:
    _base_sheet(ws)
    ws.merge_cells("A1:H1")
    ws["A1"] = "AI懂车价 · 选品决策报告"
    _title(ws["A1"])
    ws.merge_cells("A3:H4")
    action_label = "避免收" if is_avoid_export else "推荐收"
    ws["A3"] = (
        f"本报告导出{scope.get('city') or '全国'}{action_label}完整清单，共{len(export_rows)}条。"
        "全国按‘车系×年款’展示，城市按‘城市×车系’展示；页面保持简洁，Excel保留完整严格业务池。"
    )
    ws["A3"].alignment = Alignment(wrap_text=True, vertical="center")
    ws["A3"].fill = PatternFill("solid", fgColor=PALE_BLUE)
    cards = [
        ("报告范围", f"{scope.get('city') or '全国'} · {scope.get('energy_filter') or '全部能源'} · {scope.get('body_filter') or '全部车身'}"),
        ("导出清单", action_label),
        ("导出车辆数", len(export_rows)),
        ("数据窗口", scope.get("time_window") or "90天"),
        ("全量收车转化率", evidence.get("baseline_acquisition_conversion_rate")),
        ("推荐池收车转化率", evidence.get("priority_acquisition_conversion_rate")),
        ("推荐池售车转化率", evidence.get("priority_sales_conversion_rate")),
        ("推荐池单车平均毛利", evidence.get("priority_avg_profit")),
    ]
    for index, (label, value) in enumerate(cards):
        row = 6 + (index // 4) * 2
        col = 1 + (index % 4) * 2
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)
        ws.merge_cells(start_row=row + 1, start_column=col, end_row=row + 1, end_column=col + 1)
        ws.cell(row, col, label)
        ws.cell(row + 1, col, value)
        ws.cell(row, col).font = Font(size=9, bold=True, color=MUTED)
        ws.cell(row + 1, col).font = Font(size=14, bold=True, color=NAVY)
        ws.cell(row, col).fill = ws.cell(row + 1, col).fill = PatternFill("solid", fgColor="F8FAFC")
        if index == 3:
            ws.cell(row + 1, col).alignment = Alignment(horizontal="right", vertical="center")
        if index in {4, 5, 6}:
            ws.cell(row + 1, col).number_format = "0.0%"
        if index == 7:
            ws.cell(row + 1, col).number_format = '#,##0"元"'
    ws.merge_cells("A12:H12")
    ws["A12"] = "指标对比与使用说明"
    _section(ws["A12"])
    guidance = [
        f"1. 同一份全量90天数据下，全量收车转化率为{_pct_text(evidence.get('baseline_acquisition_conversion_rate'))}，推荐池为{_pct_text(evidence.get('priority_acquisition_conversion_rate'))}。",
        f"2. 推荐池成熟上架售出率为{_pct_text(evidence.get('priority_sales_conversion_rate'))}，平均卖出用时{_days_text(evidence.get('priority_avg_days_to_sell'))}，单车平均毛利{_money_text(evidence.get('priority_avg_profit'))}。",
        "3. 推荐收可主动找车，但具体车辆仍须进入单车定价确定最高收车边界；避免收不主动补库，已有库存优先去化。",
        "4. 推荐资格只接受供不应求或供需平衡，并同时通过行情、经营和风险门槛；完整明细不混入观察项。",
    ]
    for row, value in enumerate(guidance, start=13):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        ws.cell(row, 1, value).alignment = Alignment(wrap_text=True)
    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].width = 16
    ws.freeze_panes = "A2"


def _build_selection_detail_portable(ws, rows, table_name: str, *, avoid_sheet: bool) -> None:
    _base_sheet(ws)
    headers = ["排名", "品牌", "车系", "年款", "建议动作", "供需状态", "平均周转(天)", "单车典型毛利(元)", "亏损车占比", "90天总毛利贡献", "成交价格带", "样本说明", "主要依据/风险", "下一步动作"]
    ws.append(headers)
    for index, raw in enumerate(rows[:500], start=1):
        item = _dict(raw)
        dsi = _dict(item.get("dsi_signal"))
        low, high = _number(item.get("deal_price_low_90d")), _number(item.get("deal_price_high_90d"))
        price_band = f"{low / 10000:.1f}万-{high / 10000:.1f}万" if low is not None and high is not None else "暂无"
        ws.append([
            index, _safe_text(item.get("brand")), _safe_text(item.get("series")), f"{item.get('model_year')}款" if item.get("model_year") else "-", "避免收" if avoid_sheet else "推荐收",
            _safe_text(dsi.get("label") or item.get("dsi_label") or "未知"), _number(item.get("avg_deal_cycle")), _number(item.get("median_gross_profit")),
            _ratio(item.get("loss_rate")), _ratio(item.get("total_profit_contribution")), price_band, _safe_text(item.get("sample_note")),
            _safe_text("；".join(str(value) for value in ((item.get("risks") if avoid_sheet else item.get("reasons")) or [])[:3])), _safe_text(item.get("action")),
        ])
    if ws.max_row == 1:
        ws.append(["-", "", "暂无符合严格标准的车辆"] + [""] * 11)
    _style_table(ws, headers, table_name, ws.max_row)
    for col in (9, 10):
        for row in range(2, ws.max_row + 1):
            ws.cell(row, col).number_format = "0.0%"
    widths = [8, 13, 22, 11, 13, 14, 14, 18, 14, 18, 18, 28, 34, 34]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=cell.column >= 12)
    ws.freeze_panes = "A2"


def _build_selection_evidence_portable(ws, evidence) -> None:
    _base_sheet(ws)
    ws.merge_cells("A1:H1")
    ws["A1"] = "同一份全量数据的策略对比"
    _title(ws["A1"])
    headers = ["策略", "覆盖全部候选", "每100条最终收下", "成熟上架售出率", "平均周转天数", "单车平均毛利", "亏损车占比", "历史总毛利保留"]
    ws.append([])
    ws.append(headers)
    comparison = [
        ["全量车源（不筛选）", 1, evidence.get("baseline_acquisition_conversion_rate"), evidence.get("baseline_sales_conversion_rate"), evidence.get("baseline_avg_days_to_sell"), evidence.get("baseline_avg_profit"), evidence.get("baseline_loss_rate"), 1],
        ["4类行情 × DSI", evidence.get("leader_selection_rate"), evidence.get("leader_acquisition_conversion_rate"), evidence.get("leader_sales_conversion_rate"), evidence.get("leader_avg_days_to_sell"), evidence.get("leader_avg_profit"), evidence.get("leader_loss_rate"), evidence.get("leader_profit_retention_rate")],
        ["当前策略", evidence.get("priority_selection_rate"), evidence.get("priority_acquisition_conversion_rate"), evidence.get("priority_sales_conversion_rate"), evidence.get("priority_avg_days_to_sell"), evidence.get("priority_avg_profit"), evidence.get("priority_loss_rate"), evidence.get("priority_profit_retention_rate")],
    ]
    for row in comparison:
        ws.append(row)
    _format_header(ws, 3, 1, 8)
    for row in range(4, 7):
        for col in (2, 3, 4, 7, 8):
            ws.cell(row, col).number_format = "0.0%"
        ws.cell(row, 5).number_format = "0.0"
        ws.cell(row, 6).number_format = "#,##0"
    ws.merge_cells("A8:H9")
    ws["A8"] = "结论：推荐池在收车转化、售车转化、周转、单车毛利和亏损率上优于全量；覆盖范围反映筛选严格程度，不等同于准确率。"
    ws["A8"].alignment = Alignment(wrap_text=True, vertical="center")
    ws["A8"].fill = PatternFill("solid", fgColor=PALE_BLUE)
    definitions = [
        ("数据窗口", "过去90天唯一合法车源；按车源商品ID优先去重"),
        ("真实收车转化", "toc且2号岗定价人员为买手的唯一车源中，最终B2C收车成功数 / 买手出价唯一车源数"),
        ("推荐行情范围", "流动行情、结构性行情、上涨行情、常规行情"),
        ("推荐供需范围", "供不应求、供需平衡"),
        ("利润资格", "总毛利为正、亏损车占比不高于42.5%、每候选毛利不少于250元，并满足最小有效样本要求"),
        ("推荐指标", "周转不高于基线0.9倍；单车毛利、真实收车转化、售车转化均不低于基线1.1倍"),
        ("避免指标", "周转不低于基线1.1倍；单车毛利、真实收车转化、售车转化均不高于基线0.9倍"),
        ("全量每100条最终收下", evidence.get("baseline_acquisition_conversion_rate")),
        ("推荐池每100条最终收下", evidence.get("priority_acquisition_conversion_rate")),
        ("推荐池成熟上架售出率", evidence.get("priority_sales_conversion_rate")),
        ("推荐池平均卖出用时", evidence.get("priority_avg_days_to_sell")),
        ("推荐池单车平均毛利", evidence.get("priority_avg_profit")),
        ("推荐池占全部候选", evidence.get("priority_selection_rate")),
        ("推荐池保留历史总毛利", evidence.get("priority_profit_retention_rate")),
        ("时间稳定性", evidence.get("time_stability_summary")),
        ("留一品牌验证", evidence.get("brand_stress_summary")),
        ("说明", "行情与DSI缺少逐日历史快照；时间稳定性是固定车系组合的事件切片回放，不等价于未来因果收益承诺。"),
    ]
    ws.append([])
    ws.append(["口径", "说明 / 结果"])
    start = ws.max_row
    for label, value in definitions:
        ws.append([label, value])
    _format_header(ws, start, 1, 2)
    for row in range(start + 8, start + 11):
        ws.cell(row, 2).number_format = "0.0%"
    ws.cell(start + 11, 2).number_format = "0.0\"天\""
    ws.cell(start + 12, 2).number_format = "#,##0\"元\""
    for row in (start + 13, start + 14):
        ws.cell(row, 2).number_format = "0.0%"
    ws.column_dimensions["A"].width = 27
    ws.column_dimensions["B"].width = 38
    for col in range(3, 9):
        ws.column_dimensions[get_column_letter(col)].width = 18
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A4"


def _selection_strategy_evidence() -> dict[str, Any]:
    profit_path = ROOT / "results" / "evals" / "selection_profit_frontier_champion_20260713.json"
    priority_path = ROOT / "results" / "evals" / "selection_market_dsi_global_champion_20260713.json"
    leader_true_path = ROOT / "results" / "evals" / "selection_leadership_true_buyer_conversion_20260714.json"
    leader_old_path = ROOT / "outputs" / "selection_market_state_dsi_20260710" / "selection_export_data.json"
    profit = json.loads(profit_path.read_text(encoding="utf-8")) if profit_path.is_file() else {}
    priority = json.loads(priority_path.read_text(encoding="utf-8")) if priority_path.is_file() else {}
    leader_true = json.loads(leader_true_path.read_text(encoding="utf-8")) if leader_true_path.is_file() else {}
    leader_old = json.loads(leader_old_path.read_text(encoding="utf-8")) if leader_old_path.is_file() else {}
    baseline = profit.get("baseline") or {}
    champion = priority.get("weight_champion") or {}
    stability = priority.get("time_stability") or {}
    stress = profit.get("stress_test") or {}
    leader_selected = leader_old.get("selected") or {}
    leader_metrics = (leader_true.get("metrics") or {}).get("market_four_states_dsi_not_oversupplied") or {}
    return {
        "baseline_acquisition_conversion_rate": baseline.get("baseline_acquisition_conversion_rate"),
        "priority_acquisition_conversion_rate": champion.get("acquisition_conversion_rate"),
        "priority_sales_conversion_rate": champion.get("sales_conversion_rate"),
        "priority_avg_days_to_sell": champion.get("avg_days_to_sell"),
        "priority_avg_profit": champion.get("avg_profit"),
        "priority_selection_rate": champion.get("selection_rate"),
        "priority_profit_retention_rate": champion.get("profit_retention_rate"),
        "baseline_sales_conversion_rate": baseline.get("baseline_sales_conversion_rate"),
        "baseline_avg_days_to_sell": baseline.get("baseline_avg_days_to_sell"),
        "baseline_avg_profit": baseline.get("baseline_avg_profit"),
        "baseline_loss_rate": baseline.get("baseline_loss_rate"),
        "leader_selection_rate": (leader_true.get("leadership_rule") or {}).get("portfolio_coverage_rate") or leader_selected.get("selection_rate"),
        "leader_acquisition_conversion_rate": leader_metrics.get("rate"),
        "leader_sales_conversion_rate": leader_selected.get("sales_conversion_rate"),
        "leader_avg_days_to_sell": leader_selected.get("avg_days_to_sell"),
        "leader_avg_profit": leader_selected.get("avg_profit"),
        "leader_loss_rate": leader_selected.get("loss_rate"),
        "leader_profit_retention_rate": leader_selected.get("profit_retention_rate"),
        "priority_loss_rate": champion.get("loss_rate"),
        "time_stability_summary": (
            f"毛利{stability.get('avg_profit_pass_count', 0)}/{stability.get('realized_period_count', 0)}、"
            f"周转{stability.get('turnover_pass_count', 0)}/{stability.get('realized_period_count', 0)}、"
            f"收车转化{stability.get('acquisition_conversion_pass_count', 0)}/{stability.get('realized_period_count', 0)}、"
            f"售车转化{stability.get('sales_conversion_pass_count', 0)}/{stability.get('mature_sales_period_count', 0)}个切片达标"
        ),
        "brand_stress_summary": (
            f"{stress.get('leader_all_pass_count', 0)}/{stress.get('brand_count', 0)}个留一品牌折通过全部领导指标"
        ),
    }


def build_pricing_report(payload: dict[str, Any]) -> tuple[BytesIO, str]:
    result = _dict(payload.get("result"))
    metrics = _dict(result.get("metrics"))
    report = _dict(payload.get("report"))
    if not report:
        report = _dict(metrics.get("pricing_final_report"))
    if not report:
        report = _dict(_dict(metrics.get("pricing_agent")).get("final_report"))
    slots = _dict(payload.get("slots")) or _dict(metrics.get("six_elements")) or _dict(metrics.get("slots"))
    calculator = _dict(payload.get("calculator"))

    workbook = Workbook()
    summary = workbook.active
    summary.title = "估价报告"
    comparable = workbook.create_sheet("可比车证据")
    audit = workbook.create_sheet("定价依据")

    _build_pricing_summary(summary, result, report, slots, calculator)
    _build_comparable_sheet(comparable, metrics, report)
    _build_pricing_audit(audit, report)

    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    vehicle = _clean_filename(report.get("vehicle_title") or slots.get("standard_vehicle") or "当前车辆")
    stamp = datetime.now(SHANGHAI_TZ).strftime("%Y%m%d_%H%M%S")
    return output, f"AI懂车价_估价报告_{vehicle}_{stamp}.xlsx"


def _build_selection_summary(ws, result, metrics, items, strict_items, risk_items, scope) -> None:
    _base_sheet(ws)
    ws.merge_cells("A1:H1")
    ws["A1"] = "AI懂车价 · 选品决策报告"
    _title(ws["A1"])
    ws.merge_cells("A2:H3")
    ws["A2"] = _selection_summary_text(result.get("summary"))
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="center")
    ws["A2"].fill = PatternFill("solid", fgColor=PALE_BLUE)

    filters = result.get("filters") or []
    filter_text = " / ".join(str(value) for value in filters) if isinstance(filters, list) else str(filters)
    labels = [
        ("报告范围", f"{scope.get('city') or '全国'} · {scope.get('energy_filter') or '全部能源'} · {scope.get('body_filter') or '全部车身'}"),
        ("筛选条件", filter_text or "全部条件"),
        ("候选车系", len(items)),
        ("严格推荐", len(strict_items)),
        ("风险车系", len(risk_items)),
        ("数据窗口", scope.get("time_window") or "90天"),
        ("生成时间", datetime.now(SHANGHAI_TZ).strftime("%Y-%m-%d %H:%M:%S")),
        ("决策口径", "行情四状态 + DSI准入；经营实证校验、排序与降级"),
    ]
    for index, (label, value) in enumerate(labels):
        row = 5 + (index // 4) * 3
        col = 1 + (index % 4) * 2
        ws.cell(row, col, label)
        ws.cell(row + 1, col, _safe_text(value))
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)
        ws.merge_cells(start_row=row + 1, start_column=col, end_row=row + 1, end_column=col + 1)
        ws.cell(row, col).font = Font(size=10, bold=True, color=MUTED)
        ws.cell(row + 1, col).font = Font(size=14, bold=True, color=NAVY)
        for target_row in (row, row + 1):
            ws.cell(target_row, col).fill = PatternFill("solid", fgColor="F8FAFC")

    ws.merge_cells("A12:H12")
    ws["A12"] = "一线业务怎么用"
    _section(ws["A12"])
    guidance = [
        "1. 先看推荐动作与证据等级：证据不足只作观察，不能主动高价收。",
        "2. 再看周转、平均毛利、亏损率、收车转化和售车转化，确认业务表现。",
        "3. 行情状态与DSI决定市场方向；经营指标决定证据强弱与同池排序。",
        "4. 日报/政策只用于风险解释，不进入选品分数。",
    ]
    for row, text in enumerate(guidance, start=13):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        ws.cell(row, 1, text)
        ws.cell(row, 1).alignment = Alignment(wrap_text=True)

    ws.column_dimensions["A"].width = 17
    for col in range(2, 9):
        ws.column_dimensions[get_column_letter(col)].width = 15
    ws.row_dimensions[2].height = 28
    ws.freeze_panes = "A5"


SELECTION_HEADERS = [
    "排名", "品牌", "车系", "建议动作", "行情状态", "DSI供需", "最终机会分", "业务分", "置信度",
    "成交周期(天)", "平均毛利(元)", "总毛利(元)", "每10候选预计毛利(元)", "亏损率",
    "收车转化率", "售车转化率", "入库后售出率", "毛利贡献", "能源", "车身", "样本说明", "主要风险", "业务动作",
]


def _build_selection_detail(ws, rows: list[dict[str, Any]], table_name: str, risk_sheet: bool = False) -> None:
    _base_sheet(ws)
    ws.append(SELECTION_HEADERS)
    for rank, item in enumerate(rows[:500], start=1):
        dsi = _dict(item.get("dsi_signal"))
        ws.append([
            item.get("rank") or rank,
            _safe_text(item.get("brand")),
            _safe_text(item.get("series")),
            _safe_text(item.get("recommendation_label") or item.get("recommendation_level")),
            _safe_text(item.get("market_category_label") or item.get("market_category")),
            _safe_text(dsi.get("label") or item.get("dsi_label") or "未知"),
            _number(item.get("final_opportunity_score") or item.get("opportunity_score")),
            _number(item.get("business_score")),
            _ratio(item.get("confidence_score")),
            _number(item.get("avg_deal_cycle")),
            _number(item.get("median_gross_profit")),
            _number(item.get("total_gross_profit")),
            _number(item.get("expected_profit_per_10_candidates")),
            _ratio(item.get("loss_rate")),
            _ratio(item.get("acquisition_conversion_rate")),
            _ratio(item.get("sale_conversion_rate")),
            _ratio(item.get("sold_from_acquired_rate")),
            _ratio(item.get("total_profit_contribution")),
            _safe_text(item.get("energy_type")),
            _safe_text(item.get("body_type")),
            _safe_text(item.get("sample_note")),
            _safe_text("；".join(str(x) for x in (item.get("risks") or [])[:3])),
            _safe_text(item.get("action")),
        ])
    _style_table(ws, SELECTION_HEADERS, table_name, len(rows[:500]) + 1)
    for column in (9, 14, 15, 16, 17, 18):
        for cell in ws.iter_cols(min_col=column, max_col=column, min_row=2, max_row=max(ws.max_row, 2)):
            cell[0].number_format = "0.0%"
    for column in (13, 14, 15):
        for cell in ws.iter_cols(min_col=column, max_col=column, min_row=2, max_row=max(ws.max_row, 2)):
            cell[0].number_format = "#,##0"
    if ws.max_row >= 2:
        score_range = f"G2:G{ws.max_row}"
        ws.conditional_formatting.add(score_range, CellIsRule(operator="greaterThanOrEqual", formula=["70"], fill=PatternFill("solid", fgColor=PALE_GREEN)))
        ws.conditional_formatting.add(score_range, CellIsRule(operator="lessThan", formula=["40"], fill=PatternFill("solid", fgColor=PALE_RED)))
        if not risk_sheet:
            chart = BarChart()
            chart.type = "bar"
            chart.title = "候选车系总毛利 Top 10"
            chart.height = 7
            chart.width = 13
            end = min(ws.max_row, 11)
            chart.add_data(Reference(ws, min_col=14, min_row=1, max_row=end), titles_from_data=True)
            chart.set_categories(Reference(ws, min_col=3, min_row=2, max_row=end))
            ws.add_chart(chart, "Z2")


def _build_selection_evidence(ws, metrics: dict[str, Any], scope: dict[str, Any]) -> None:
    _base_sheet(ws)
    ws.append(["类别", "字段", "内容"])
    rows: list[tuple[str, str, Any]] = []
    for key, value in scope.items():
        rows.append(("任务范围", key, value))
    for group in ("selection_audit", "score_policy", "data_source"):
        for key, value in _dict(metrics.get(group)).items():
            rows.append((group, key, value))
    rows.extend([
        ("方法", "主信号", "行情四状态 + DSI供需"),
        ("方法", "经营验证", "真实收车转化、售车转化、周转、平均/总毛利、亏损率"),
        ("方法", "扩量情景", "每10个候选预计毛利=10×实际收车转化率×成熟收车售出率×中位毛利；仅用于统一规模对比，不等同未来承诺，不参与当前排名"),
        ("方法", "日报/政策", "仅用于风险解释，不参与数值排名"),
        ("样本门控", "观察", "有效经营证据不足：保留观察，不进入可执行推荐"),
        ("样本门控", "利润资格", "通过最小有效样本、利润和亏损约束"),
        ("城市标签", "可关注", "本地经营证据达到基础门槛，并继续通过经营指标与风险门控"),
        ("城市标签", "重点关注", "本地经营证据充分，并继续通过经营指标与风险门控"),
    ])
    for category, field, value in rows:
        ws.append([_safe_text(category), _safe_text(field), _safe_text(_compact(value))])
    _style_table(ws, ["类别", "字段", "内容"], "EvidenceTable", len(rows) + 1)
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 90
    for row in range(2, ws.max_row + 1):
        ws.cell(row, 3).alignment = Alignment(wrap_text=True, vertical="top")


def _build_pricing_summary(ws, result, report, slots, calculator) -> None:
    _base_sheet(ws)
    ws.merge_cells("A1:H1")
    ws["A1"] = "AI懂车价 · 综合估价报告"
    _title(ws["A1"])
    ws.merge_cells("A2:H2")
    ws["A2"] = _safe_text(report.get("vehicle_title") or slots.get("standard_vehicle") or result.get("title") or "当前车辆")
    ws["A2"].font = Font(size=14, bold=True, color=NAVY)

    elements = [
        ("标准车型", slots.get("standard_vehicle") or report.get("vehicle_title")),
        ("上牌时间", slots.get("first_license_date") or slots.get("registration_date")),
        ("里程", slots.get("mileage_wan_km") or slots.get("mileage_km")),
        ("城市", slots.get("city")),
        ("过户次数", slots.get("transfer_count")),
        ("颜色", slots.get("color")),
        ("车况等级", slots.get("condition_grade") or slots.get("condition")),
    ]
    ws.append([])
    ws.append(["七要素", "值"])
    for label, value in elements:
        ws.append([label, _safe_text(value)])
    _format_header(ws, 4, 1, 2)

    listing = _first_number(report, "listing_price_yuan", "listing_price")
    listing_low = _first_number(report, "listing_price_low_yuan", "listing_price_low")
    listing_high = _first_number(report, "listing_price_high_yuan", "listing_price_high")
    sale = _first_number(report, "sale_price_yuan", "sale_price")
    sale_low = _first_number(report, "sale_price_low_yuan", "sale_price_low")
    sale_high = _first_number(report, "sale_price_high_yuan", "sale_price_high")
    purchase = _number(calculator.get("currentPurchaseYuan")) or _first_number(report, "point_price_yuan", "purchase_price_yuan", "purchase_price")
    purchase_low = _first_number(report, "purchase_price_low_yuan", "lower_yuan", "purchase_price_low")
    purchase_high = _first_number(report, "purchase_price_high_yuan", "upper_yuan", "purchase_price_high")
    max_purchase = _first_number(report, "max_c2b_price_yuan", "max_c2b_price", "upper_yuan", "purchase_price_high_yuan")
    current_sale = _number(calculator.get("currentSaleYuan")) or sale

    start = 4
    ws.cell(start, 4, "价格决策梯度")
    _format_header(ws, start, 4, 5)
    price_rows = [
        ("建议挂牌价", listing, listing_low, listing_high, "对外挂牌与议价起点"),
        ("预计实际售车价", current_sale, sale_low, sale_high, "预计最终成交口径"),
        ("C2B最高收车价", max_purchase, None, max_purchase, "含成本、风险与最低利润约束"),
        ("预计实际收车价", purchase, purchase_low, purchase_high, "预计最终签约口径"),
    ]
    ws.append([])
    ws.cell(start + 1, 4, "价格层级")
    ws.cell(start + 1, 5, "中心价(元)")
    ws.cell(start + 1, 6, "区间下限(元)")
    ws.cell(start + 1, 7, "区间上限(元)")
    ws.cell(start + 1, 8, "用途")
    _format_header(ws, start + 1, 4, 8)
    for row_index, row in enumerate(price_rows, start=start + 2):
        for col_index, value in enumerate(row, start=4):
            ws.cell(row_index, col_index, _safe_text(value) if col_index in (4, 8) else value)
        for col_index in (5, 6, 7):
            ws.cell(row_index, col_index).number_format = "¥#,##0"

    calc_row = 13
    ws.merge_cells(start_row=calc_row, start_column=1, end_row=calc_row, end_column=2)
    ws.cell(calc_row, 1, "利润测算")
    _section(ws.cell(calc_row, 1))
    recon = _number(calculator.get("reconYuan")) or _first_number(report, "estimated_recon_cost_yuan", "estimated_recon_cost")
    platform = _number(calculator.get("platformYuan")) or _first_number(report, "platform_service_cost_yuan", "platform_service_cost")
    buffer = _number(calculator.get("bufferYuan")) or _first_number(report, "risk_buffer_yuan", "risk_buffer")
    calc_values = [
        ("试算收车价", purchase), ("试算售车价", current_sale), ("整备成本", recon),
        ("平台/服务成本", platform), ("风险缓冲", buffer),
    ]
    for index, (label, value) in enumerate(calc_values):
        row = calc_row + 1 + index
        ws.cell(row, 1, label)
        ws.cell(row, 2, value)
        ws.cell(row, 2).number_format = "¥#,##0"
    gross_row = calc_row + 6
    ws.cell(gross_row, 1, "预计毛利")
    ws.cell(gross_row, 2, f"=B{calc_row + 2}-SUM(B{calc_row + 1},B{calc_row + 3}:B{calc_row + 5})")
    ws.cell(gross_row, 2).number_format = "¥#,##0"
    ws.cell(gross_row + 1, 1, "预计毛利率")
    ws.cell(gross_row + 1, 2, f"=IFERROR(B{gross_row}/B{calc_row + 2},0)")
    ws.cell(gross_row + 1, 2).number_format = "0.0%"

    ws.merge_cells("D13:H13")
    ws["D13"] = "为什么是这个价"
    _section(ws["D13"])
    ws.merge_cells("D14:H18")
    ws["D14"] = _safe_text(report.get("why_this_price") or report.get("summary_why") or "基于同款证据、六/七要素修正、市场边界与利润约束形成。")
    ws["D14"].alignment = Alignment(wrap_text=True, vertical="top")

    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].width = 18 if col != 1 else 22
    ws.column_dimensions["D"].width = 24
    ws.column_dimensions["H"].width = 34
    ws.freeze_panes = "A4"


def _build_comparable_sheet(ws, metrics: dict[str, Any], report: dict[str, Any]) -> None:
    _base_sheet(ws)
    candidates = _list_of_dicts(metrics.get("comparable_evidence"))
    if not candidates:
        reasoning = _dict(report.get("price_reasoning"))
        candidates = _list_of_dicts(reasoning.get("comparable_evidence") or reasoning.get("candidates"))
    headers = ["来源", "车型", "年份", "上牌时间", "里程", "城市", "颜色", "过户", "价格(元)", "证据关系", "时效", "链接"]
    ws.append(headers)
    for item in candidates[:300]:
        ws.append([
            _safe_text(item.get("source") or item.get("platform")),
            _safe_text(item.get("vehicle") or item.get("title") or item.get("model")),
            _safe_text(item.get("model_year") or item.get("year")),
            _safe_text(item.get("registration_date") or item.get("first_license_date")),
            _safe_text(item.get("mileage") or item.get("mileage_km")),
            _safe_text(item.get("city")),
            _safe_text(item.get("color")),
            _safe_text(item.get("transfer_count")),
            _number(item.get("price") or item.get("price_yuan") or item.get("listing_price_yuan")),
            _safe_text(item.get("relation") or item.get("relation_label")),
            _safe_text(item.get("recency") or item.get("observed_at") or item.get("event_time")),
            _safe_text(item.get("url")),
        ])
    _style_table(ws, headers, "ComparableEvidence", len(candidates[:300]) + 1)
    for cell in ws["I"][1:]:
        cell.number_format = "¥#,##0"


def _build_pricing_audit(ws, report: dict[str, Any]) -> None:
    _base_sheet(ws)
    ws.append(["模块", "证据/解释"])
    rows: list[tuple[str, Any]] = []
    for key in ("internal_basis", "main_risks", "action_guide", "technical_audit", "price_boundary"):
        value = report.get(key)
        if isinstance(value, list):
            for item in value:
                rows.append((key, _compact(item)))
        elif value:
            rows.append((key, _compact(value)))
    rows.extend([
        ("口径", "挂牌价来自三方同款挂牌证据并做六/七要素调整，不直接把挂牌中位数当成交价。"),
        ("口径", "预计实际售车价与预计实际收车价来自定价模型、内部交易证据和市场约束；售车价不得低于收车价。"),
        ("口径", "旧交易记录必须做时效处理；same_series_year不能作为主价格锚。"),
    ])
    for label, value in rows:
        ws.append([_safe_text(label), _safe_text(value)])
    _style_table(ws, ["模块", "证据/解释"], "PricingAudit", len(rows) + 1)
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 110
    for cell in ws["B"][1:]:
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def _base_sheet(ws) -> None:
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"


def _title(cell) -> None:
    cell.fill = PatternFill("solid", fgColor=BLUE)
    cell.font = Font(size=18, bold=True, color=WHITE)
    cell.alignment = Alignment(vertical="center")
    cell.parent.row_dimensions[cell.row].height = 32


def _section(cell) -> None:
    cell.fill = PatternFill("solid", fgColor=PALE_BLUE)
    cell.font = Font(size=12, bold=True, color=NAVY)
    cell.alignment = Alignment(vertical="center")


def _format_header(ws, row: int, start_col: int, end_col: int) -> None:
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row, col)
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.font = Font(bold=True, color=WHITE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _style_table(ws, headers: list[str], name: str, row_count: int) -> None:
    _format_header(ws, 1, 1, len(headers))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(row_count, 1)}"
    if row_count >= 2:
        table = Table(displayName=name, ref=f"A1:{get_column_letter(len(headers))}{row_count}")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showFirstColumn=False, showLastColumn=False)
        ws.add_table(table)
    widths = [8, 14, 24, 18, 16, 14, 14, 12, 12, 12, 12, 14, 16, 16, 20, 12, 14, 14, 14, 12, 12, 12, 34, 44, 44]
    for index, width in enumerate(widths[: len(headers)], start=1):
        ws.column_dimensions[get_column_letter(index)].width = width
    thin = Side(style="thin", color=GRID)
    for row in ws.iter_rows(min_row=1, max_row=max(ws.max_row, 1), min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = Border(bottom=thin)
            cell.alignment = Alignment(vertical="top", wrap_text=cell.column >= 22)


def _dict(value: Any) -> dict[str, Any]:
    return value if isinstance(value, dict) else {}


def _selection_summary_text(value: Any) -> str:
    if isinstance(value, str) and value.strip():
        return value.strip()
    if isinstance(value, dict):
        for key in ("headline", "direct_answer", "summary", "answer", "text"):
            candidate = value.get(key)
            if isinstance(candidate, str) and candidate.strip():
                return candidate.strip()
        findings = value.get("key_findings")
        if isinstance(findings, list):
            compact = [str(item).strip() for item in findings if str(item).strip()]
            if compact:
                return " ".join(compact[:3])
    return "基于90天内部经营表现、行情四状态与DSI供需信号生成。"


def _list_of_dicts(value: Any) -> list[dict[str, Any]]:
    return [item for item in value if isinstance(item, dict)] if isinstance(value, list) else []


def _safe_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value)
    if text[:1] in {"=", "+", "-", "@"}:
        return "'" + text
    return text


def _number(value: Any, *, integer: bool = False) -> float | int | None:
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    if not math.isfinite(number):
        return None
    return int(round(number)) if integer else round(number, 4)


def _ratio(value: Any) -> float | None:
    number = _number(value)
    if number is None:
        return None
    return number / 100 if abs(number) > 1 else number


def _relative_change(value: Any) -> float | None:
    """Convert a stored relative ratio (for example 1.3559x) into +35.59%."""
    number = _number(value)
    return None if number is None else number - 1


def _first_number(values: dict[str, Any], *keys: str) -> float | None:
    for key in keys:
        value = _number(values.get(key))
        if value is not None and value != 0:
            return value * 10000 if value < 1000 and not key.endswith("_yuan") else value
    return None


def _compact(value: Any) -> str:
    if isinstance(value, dict):
        return "；".join(f"{key}={_compact(item)}" for key, item in value.items())
    if isinstance(value, list):
        return "；".join(_compact(item) for item in value)
    return "" if value is None else str(value)


def _clean_filename(value: Any) -> str:
    text = re.sub(r"[\\/:*?\"<>|\s]+", "_", str(value or "报告")).strip("_")
    return text[:80] or "报告"
