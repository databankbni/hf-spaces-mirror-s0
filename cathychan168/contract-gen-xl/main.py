"""
XL 合同生成微服务 v5.7（贤凌科技版）
POST /generate 接收合同 JSON，返回 Excel + PDF 二进制（base64）

公司：深圳市贤凌科技有限公司
- 正式合同前缀：XL
- 仅订单前缀：DD
- 公章：seal_xianling.png（4cm 透明 PNG，浮于文字上方）
"""
import base64
import io
import json
import os
import re
import subprocess
import tempfile
from copy import copy
from datetime import datetime
from typing import List, Optional

from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, Side
from pydantic import BaseModel

app = FastAPI(title="贤凌合同生成服务", version="5.7-XL")
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])

TEMPLATE_PATH = os.getenv("TEMPLATE_PATH", "/app/合同模板_占位符版.xlsx")
LIBREOFFICE_BIN = os.getenv("LIBREOFFICE_BIN", "libreoffice")


# ============== v5.7 单据类型配置 ==============
DOC_TYPE_CONFIG = {
    "正式合同": {
        "no_prefix": "XL",
        "title_replacements": [],  # 不改
    },
    "仅订单": {
        "no_prefix": "DD",
        "title_replacements": [
            ("销售合同", "销售订单"),
            ("购销合同", "销售订单"),
            ("产品销售合同", "销售订单"),
        ],
    },
}


# ============== 数据模型 ==============
class Product(BaseModel):
    name: str
    unit: str = "kgs"
    quantity: float
    unit_price: float

    @property
    def amount(self) -> float:
        return round(self.quantity * self.unit_price, 2)


class ContractRequest(BaseModel):
    contract_no: Optional[str] = None
    party_a: str = "深圳市贤凌科技有限公司"
    party_b: str
    sign_place: str = "深圳"
    sign_date: Optional[str] = None  # YYYY-MM-DD
    products: List[Product]
    delivery_method: str = "送货上门"
    transport_method: str = "甲方代办运输,运输费用甲方负担"
    payment_method: str = "款到发货"
    party_b_phone: str = ""
    party_b_fax: str = ""
    party_b_address: str = ""
    party_b_bank: str = ""
    party_b_account: str = ""
    output_format: str = "both"  # excel | pdf | both
    doc_type: str = "正式合同"   # v5.7：正式合同 / 仅订单


# ============== 工具函数 ==============
def num2cn(num: float) -> str:
    """把数字转成中文大写金额（符合财务规范）"""
    digits = "零壹贰叁肆伍陆柒捌玖"
    units = ["", "拾", "佰", "仟"]
    big_units = ["", "万", "亿", "兆"]

    if num == 0:
        return "零元整"

    integer, _, decimal = f"{num:.2f}".partition(".")
    int_part = int(integer)

    if int_part == 0:
        int_str = ""
    else:
        groups = []
        g_idx = 0
        tmp = int_part
        while tmp > 0:
            g = tmp % 10000
            groups.append((g, g_idx))
            tmp //= 10000
            g_idx += 1

        parts = []
        for g, idx in reversed(groups):
            if g == 0:
                parts.append("")
                continue
            s = ""
            has_prev_zero = False
            for i in range(3, -1, -1):
                d = (g // (10 ** i)) % 10
                if d == 0:
                    has_prev_zero = True
                else:
                    if has_prev_zero and s:
                        s += "零"
                    s += digits[d] + units[i]
                    has_prev_zero = False
            parts.append(s + big_units[idx])

        int_str = ""
        for i, p in enumerate(parts):
            if not p:
                if int_str and not int_str.endswith("零"):
                    int_str += "零"
            else:
                int_str += p
        int_str = int_str.rstrip("零")

    jiao = int(decimal[0])
    fen = int(decimal[1])

    if int_str:
        result = int_str + "元"
        if jiao == 0 and fen == 0:
            result += "整"
        else:
            if jiao > 0:
                result += digits[jiao] + "角"
            elif fen > 0:
                result += "零"
            if fen > 0:
                result += digits[fen] + "分"
    else:
        result = ""
        if jiao > 0:
            result += digits[jiao] + "角"
        if fen > 0:
            result += digits[fen] + "分"
        if not result:
            result = "零元整"
    return result


COUNTER_FILE = os.getenv("COUNTER_FILE", "/app/counter.txt")
COUNTER_START = int(os.getenv("COUNTER_START", "0"))


def _counter_path_for(doc_type: str) -> str:
    """v5.7: 两种单据独立计数器，避免序号冲突"""
    cfg = DOC_TYPE_CONFIG.get(doc_type, DOC_TYPE_CONFIG["正式合同"])
    prefix = cfg["no_prefix"]
    if prefix == "XL":
        return COUNTER_FILE  # 主单据用默认 counter.txt
    base, ext = os.path.splitext(COUNTER_FILE)
    return f"{base}_{prefix.lower()}{ext}"


def next_seq(doc_type: str = "正式合同") -> int:
    """按年累计计数器，跨年自动重置回 0001。各 doc_type 独立。"""
    counter_file = _counter_path_for(doc_type)
    today_year = datetime.now().year
    data = {"year": today_year, "counter": COUNTER_START}

    try:
        with open(counter_file, "r") as f:
            content = f.read().strip()
        if content.startswith("{"):
            data = json.loads(content)
        elif content:
            data = {"year": today_year, "counter": int(content)}
    except (FileNotFoundError, ValueError, json.JSONDecodeError):
        pass

    if data.get("year") != today_year:
        data = {"year": today_year, "counter": 0}

    data["counter"] = data.get("counter", 0) + 1

    try:
        with open(counter_file, "w") as f:
            json.dump(data, f)
    except OSError:
        pass

    return data["counter"]


def gen_contract_no(doc_type: str = "正式合同") -> str:
    """生成单据编号 = 前缀 + YYYYMMDD + 4位序号（年度累计，跨年重置）"""
    cfg = DOC_TYPE_CONFIG.get(doc_type, DOC_TYPE_CONFIG["正式合同"])
    prefix = cfg["no_prefix"]
    return f"{prefix}{datetime.now().strftime('%Y%m%d')}{next_seq(doc_type):04d}"


# ============== 模板填充 ==============
def fill_template(req: ContractRequest) -> bytes:
    wb = load_workbook(TEMPLATE_PATH)
    ws = wb.active

    # === v5.7 单据类型差异化处理（标题替换）===
    cfg = DOC_TYPE_CONFIG.get(req.doc_type, DOC_TYPE_CONFIG["正式合同"])
    title_replacements = cfg["title_replacements"]
    if title_replacements:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    new_val = cell.value
                    for old, new in title_replacements:
                        if old in new_val:
                            new_val = new_val.replace(old, new)
                    if new_val != cell.value:
                        cell.value = new_val

    # 主动插入 logo（按原图比例等比缩放，限定最大高 65px）
    LOGO_PATH = os.path.join(os.path.dirname(TEMPLATE_PATH), "xl_logo.png")
    logo_to_add = None
    if os.path.exists(LOGO_PATH):
        from openpyxl.drawing.image import Image as XLImage
        from PIL import Image as PILImage
        logo_to_add = XLImage(LOGO_PATH)
        try:
            with PILImage.open(LOGO_PATH) as _pil:
                _ow, _oh = _pil.size
            _max_h = 65
            if _oh > 0:
                logo_to_add.height = _max_h
                logo_to_add.width = int(_max_h * _ow / _oh)
        except Exception:
            logo_to_add.width = 78
            logo_to_add.height = 60

    # 收紧页边距 + 列宽压缩 + 强制单页等比缩放（LibreOffice 渲染时自动 fit 到 A4 宽度）
    try:
        ws.page_margins.left = 0.3
        ws.page_margins.right = 0.3
        ws.page_margins.top = 0.4
        ws.page_margins.bottom = 0.4
        from openpyxl.utils import get_column_letter
        total_w = sum((ws.column_dimensions[get_column_letter(i)].width or 10) for i in range(1, 7))
        if total_w > 78:
            factor = 78 / total_w
            for i in range(1, 7):
                col = get_column_letter(i)
                w = ws.column_dimensions[col].width
                if w:
                    ws.column_dimensions[col].width = w * factor
        # fit-to-page：让 LibreOffice 转 PDF 时按 A4 宽度等比缩放，自动撑满
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 1
        ws.sheet_properties.pageSetUpPr.fitToPage = True
    except Exception:
        pass

    product_rows = []
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and "{{PRODUCT_ROW}}" in cell.value:
                product_rows.append(cell.row)
                break
    product_rows.sort()

    if not product_rows:
        raise RuntimeError("模板缺失 {{PRODUCT_ROW}} 占位符")

    products = req.products
    n_products = len(products)
    max_slots = len(product_rows)

    if n_products > 45:
        raise RuntimeError(f"产品数量 {n_products} 超过单张合同上限 45，请拆分订单")

    # v5.13: 超过模板预留行数时动态插行（合同自动变多页：第1页货物表，第2页条款+签字）
    if n_products > max_slots:
        import copy as _copy
        extra = n_products - max_slots
        last = product_rows[-1]
        ins_at = last + 1
        moved = []
        for mr in list(ws.merged_cells.ranges):
            if mr.min_row >= ins_at:
                moved.append((mr.min_row, mr.min_col, mr.max_row, mr.max_col))
                ws.unmerge_cells(mr.coord)
        heights = {r: d.height for r, d in ws.row_dimensions.items() if r >= ins_at and d.height is not None}
        ws.insert_rows(ins_at, extra)
        for r in sorted(heights, reverse=True):
            ws.row_dimensions[r + extra].height = heights[r]
        for r1, c1, r2, c2 in moved:
            ws.merge_cells(start_row=r1 + extra, start_column=c1, end_row=r2 + extra, end_column=c2)
        for k in range(extra):
            nr = ins_at + k
            for col in range(1, 10):
                src = ws.cell(row=last, column=col)
                dst = ws.cell(row=nr, column=col)
                dst.font = _copy.copy(src.font)
                dst.border = _copy.copy(src.border)
                dst.fill = _copy.copy(src.fill)
                dst.alignment = _copy.copy(src.alignment)
                dst.number_format = src.number_format
            ws.row_dimensions[nr].height = ws.row_dimensions[last].height or 24.0
            ws.merge_cells(start_row=nr, start_column=1, end_row=nr, end_column=2)
            product_rows.append(nr)
        max_slots = len(product_rows)

    def delete_row_keeping_merges(sheet, r):
        to_remerge = []
        for mr in list(sheet.merged_cells.ranges):
            if mr.min_row > r:
                to_remerge.append((mr.min_row - 1, mr.min_col, mr.max_row - 1, mr.max_col))
                sheet.unmerge_cells(mr.coord)
            elif mr.min_row == r and mr.max_row == r:
                sheet.unmerge_cells(mr.coord)
        heights_below = {}
        for row_idx, dim in list(sheet.row_dimensions.items()):
            if row_idx > r and dim.height is not None:
                heights_below[row_idx - 1] = dim.height
        sheet.delete_rows(r, 1)
        for row_idx, h in heights_below.items():
            sheet.row_dimensions[row_idx].height = h
        for minr, minc, maxr, maxc in to_remerge:
            sheet.merge_cells(start_row=minr, start_column=minc, end_row=maxr, end_column=maxc)

    from openpyxl.styles import Font as _Font
    product_font = _Font(name="宋体", size=10.5, color="000000")
    for i, p in enumerate(products):
        r = product_rows[i]
        ws.cell(row=r, column=1, value=p.name)
        ws.cell(row=r, column=1).font = product_font
        ws.cell(row=r, column=3, value=p.unit)
        ws.cell(row=r, column=3).font = product_font
        ws.cell(row=r, column=4, value=p.quantity)
        ws.cell(row=r, column=4).font = product_font
        ws.cell(row=r, column=5, value=p.unit_price)
        ws.cell(row=r, column=5).font = product_font
        ws.cell(row=r, column=6, value=p.amount)
        ws.cell(row=r, column=6).font = product_font

    unused_rows = product_rows[n_products:]
    for r in sorted(unused_rows, reverse=True):
        delete_row_keeping_merges(ws, r)

    # v5.13: >6 行时记录货物表结束行，后面用双打印区域分页（manual row break 在旧版 LibreOffice 会丢内容）
    _split_after_row = None
    if n_products > 6:
        _fit_one_page = bool(getattr(ws.sheet_properties.pageSetUpPr, 'fitToPage', False))
        if not _fit_one_page:
            for _row in ws.iter_rows():
                for _cell in _row:
                    if isinstance(_cell.value, str) and '合计人民币金额' in _cell.value:
                        _split_after_row = _cell.row
                        break
                if _split_after_row:
                    break

    long_clause_keywords = ("如对货物数量和规格",)
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if isinstance(v, str):
                if any(kw in v for kw in long_clause_keywords):
                    ws.row_dimensions[cell.row].height = 45
                    break
                if "八、争议解决方式" in v:
                    ws.row_dimensions[cell.row].height = 32
                    break

    total = sum(p.amount for p in products)
    total_str = str(int(total)) if total == int(total) else f"{total:.2f}".rstrip('0').rstrip('.')

    contract_no_final = req.contract_no or gen_contract_no(req.doc_type)

    repl = {
        "{{contract_no}}": contract_no_final,
        "{{party_a}}": req.party_a,
        "{{party_b}}": req.party_b,
        "{{sign_place}}": req.sign_place,
        "{{sign_date}}": req.sign_date or datetime.now().strftime("%Y-%m-%d"),
        "{{total_amount}}": total_str,
        "{{total_amount_cn}}": num2cn(total),
        "{{delivery_method}}": req.delivery_method,
        "{{transport_method}}": req.transport_method,
        "{{payment_method}}": req.payment_method,
        "{{party_b_phone}}": req.party_b_phone,
        "{{party_b_fax}}": req.party_b_fax,
        "{{party_b_address}}": req.party_b_address,
        "{{party_b_bank}}": req.party_b_bank,
        "{{party_b_account}}": req.party_b_account,
    }

    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            val = cell.value
            if isinstance(val, str):
                for k, v in repl.items():
                    if k in val:
                        val = val.replace(k, str(v))
                cell.value = val

    # 乙方联系信息行（地址/开户行/账号/联系电话/传真）按显示宽度自动加宽行高
    # 中文/全角字符按 1.0 算，ASCII（数字/字母/半角符号）按 0.5 算，模拟实际占位
    def _display_width(s: str) -> float:
        w = 0.0
        for ch in s:
            w += 0.5 if ord(ch) < 0x80 else 1.0
        return w

    contact_prefixes = ("地址：", "开户行：", "账号：", "联系电话：", "传真：")
    chars_per_line = 20.0  # 列 E 一行可容纳的"显示宽度"
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if isinstance(v, str) and cell.column >= 4:
                stripped = v.lstrip()
                if any(stripped.startswith(p) for p in contact_prefixes):
                    dw = _display_width(stripped)
                    if dw > chars_per_line:  # 单行能放下就不动
                        import math
                        n_lines = math.ceil(dw / chars_per_line)
                        ws.row_dimensions[cell.row].height = max(22.0, n_lines * 22.0)
                    break

    try:
        last_real_row = 0
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if v is not None and not (isinstance(v, str) and v.strip() == ""):
                    if cell.row > last_real_row:
                        last_real_row = cell.row
        if last_real_row > 0 and ws.max_row > last_real_row:
            ws.delete_rows(last_real_row + 1, ws.max_row - last_real_row)
        if last_real_row > 0:
            if _split_after_row and _split_after_row < last_real_row:
                # 双打印区域：第1页=抬头+货物表+合计，第2页=条款+签字（每个区域独立成页，跨 LO 版本可靠）
                ws.print_area = [f"A1:F{_split_after_row}", f"A{_split_after_row + 1}:F{last_real_row}"]
            else:
                ws.print_area = f"A1:F{last_real_row}"
    except Exception:
        pass

    if logo_to_add is not None:
        if hasattr(ws, "_images"):
            ws._images.clear()
        # logo 锚在 A1，向右偏 30px → 从最左稍右移，但不撞 A2 居中抬头
        from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor as _OCA, AnchorMarker as _AM
        from openpyxl.drawing.xdr import XDRPositiveSize2D as _XDR
        from openpyxl.utils.units import pixels_to_EMU as _p2e
        _logo_w_emu = _p2e(logo_to_add.width or 73)
        _logo_h_emu = _p2e(logo_to_add.height or 65)
        logo_to_add.anchor = _OCA(
            _from=_AM(col=0, colOff=_p2e(30), row=0, rowOff=_p2e(18)),
            ext=_XDR(cx=_logo_w_emu, cy=_logo_h_emu),
        )
        ws.add_image(logo_to_add)

    # 公章（合同/订单都盖贤凌章；浮于文字上方 = 绘图层后插入 = 渲染时在最上层）
    SEAL_PATH = os.path.join(os.path.dirname(TEMPLATE_PATH), "seal_xianling.png")
    if os.path.exists(SEAL_PATH):
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
        from openpyxl.drawing.xdr import XDRPositiveSize2D
        from openpyxl.utils.units import pixels_to_EMU, cm_to_EMU

        seal_cell = None
        for row in ws.iter_rows():
            for cell in row:
                if cell.value == "盖章：" and cell.column == 1:
                    seal_cell = cell
                    break
            if seal_cell:
                break

        if seal_cell:
            seal_img = XLImage(SEAL_PATH)
            size_emu = cm_to_EMU(4.0)
            marker = AnchorMarker(
                col=seal_cell.column - 1,
                colOff=pixels_to_EMU(35),
                row=seal_cell.row - 1,
                rowOff=pixels_to_EMU(-5),
            )
            seal_img.anchor = OneCellAnchor(
                _from=marker,
                ext=XDRPositiveSize2D(cx=size_emu, cy=size_emu),
            )
            ws.add_image(seal_img)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def excel_to_pdf(xlsx_bytes: bytes) -> bytes:
    """用 LibreOffice 把 xlsx 转 pdf，并删掉文字极少的空白页"""
    with tempfile.TemporaryDirectory() as tmp:
        xlsx_path = os.path.join(tmp, "contract.xlsx")
        with open(xlsx_path, "wb") as f:
            f.write(xlsx_bytes)

        r = subprocess.run(
            [LIBREOFFICE_BIN, "--headless", "--convert-to", "pdf",
             "--outdir", tmp, xlsx_path],
            capture_output=True, text=True, timeout=120,
        )
        if r.returncode != 0:
            raise RuntimeError(f"LibreOffice failed: {r.stderr}")
        pdf_path = os.path.join(tmp, "contract.pdf")
        if not os.path.exists(pdf_path):
            raise RuntimeError(f"PDF not produced; libreoffice output: {r.stdout}")
        with open(pdf_path, "rb") as f:
            pdf_bytes = f.read()

        try:
            from pypdf import PdfReader, PdfWriter
            reader = PdfReader(io.BytesIO(pdf_bytes))
            if len(reader.pages) > 1:
                # v5.13.2: 只删几乎无文字的真空白页（老逻辑"只保留文字最多的一页"会把多页合同的条款签字页删掉）
                keep = [p for p in reader.pages if len((p.extract_text() or "").strip()) >= 30]
                if keep and len(keep) < len(reader.pages):
                    writer = PdfWriter()
                    for p in keep:
                        writer.add_page(p)
                    out = io.BytesIO()
                    writer.write(out)
                    pdf_bytes = out.getvalue()
        except Exception:
            pass
        return pdf_bytes


# ============== API ==============
@app.get("/")
def health():
    return {
        "ok": True,
        "service": "contract-gen",
        "version": "5.11-XL",
        "stamp": True,
        "template_exists": os.path.exists(TEMPLATE_PATH),
        "supported_doc_types": list(DOC_TYPE_CONFIG.keys()),
    }


@app.post("/generate")
def generate(req: ContractRequest):
    try:
        contract_no = req.contract_no or gen_contract_no(req.doc_type)
        req.contract_no = contract_no

        xlsx = fill_template(req)
        out = {
            "contract_no": contract_no,
            "doc_type": req.doc_type,
            "total_amount": sum(p.amount for p in req.products),
            "total_amount_cn": num2cn(sum(p.amount for p in req.products)),
            "filename_xlsx": f"{contract_no}_{req.party_b}.xlsx",
            "filename_pdf": f"{contract_no}_{req.party_b}.pdf",
        }

        if req.output_format in ("excel", "both"):
            out["excel_base64"] = base64.b64encode(xlsx).decode()
        if req.output_format in ("pdf", "both"):
            pdf = excel_to_pdf(xlsx)
            out["pdf_base64"] = base64.b64encode(pdf).decode()

        return out
    except Exception as e:
        raise HTTPException(500, str(e))




# ============== v5.8 对方合同盖章 /stamp ==============
STAMP_SEAL_PATH = os.path.join(os.path.dirname(TEMPLATE_PATH), "seal_xianling.png")
STAMP_KEYWORDS = [
    "深圳市贤凌科技有限公司",
    "贤凌科技",
    "（盖章）", "(盖章)", "盖章", "签章", "公章",
]
SEAL_SIZE_PT = 4.0 / 2.54 * 72  # 4cm 国标章 ≈ 113.4pt
STAMP_MARGIN_PT = 42  # 兜底位置离页边的距离


class StampRequest(BaseModel):
    file_base64: str
    file_ext: str = "pdf"                 # pdf / doc / docx / png / jpg / jpeg / bmp / webp
    position_hint: Optional[str] = None   # 如 "末页右下" / "第2页左下" / "骑缝" / "不骑缝"
    filename: Optional[str] = None
    riding_seal: Optional[bool] = None    # 骑缝章：None=多页自动加，True/False 强制


def _office_to_pdf(data: bytes, ext: str) -> bytes:
    """Word 等办公文档 → PDF（复用 LibreOffice）"""
    with tempfile.TemporaryDirectory() as tmp:
        src = os.path.join(tmp, f"incoming.{ext}")
        with open(src, "wb") as f:
            f.write(data)
        r = subprocess.run(
            [LIBREOFFICE_BIN, "--headless", "--convert-to", "pdf", "--outdir", tmp, src],
            capture_output=True, text=True, timeout=120,
        )
        pdf_path = os.path.join(tmp, "incoming.pdf")
        if r.returncode != 0 or not os.path.exists(pdf_path):
            raise RuntimeError(f"LibreOffice 转换失败: {r.stderr or r.stdout}")
        with open(pdf_path, "rb") as f:
            return f.read()


_OFFSET_RE = re.compile(
    r"(?:往|向)?(左上|左下|右上|右下|左|右|上|下)?(?:边|方|面)?"
    r"(?:挪|平移|偏移|移动|移|偏)\s*([\d.]+)\s*(厘米|公分|[cC][mM]|毫米|[mM][mM])"
)


def _parse_offsets(hint: str):
    """v5.11 解析 '上挪2厘米'/'往左移1.5cm' 等偏移 → (dx_pt, dy_pt, 剥掉偏移后的提示, 提示语)

    方向可组合出现多次（'左挪1厘米 上挪2厘米'）；'右下挪3厘米' = 右、下各挪3厘米。
    没写方向的偏移（'挪10厘米'）无法执行，忽略并在卡片提示正确写法。"""
    dx = dy = 0.0
    note = ""

    def _apply(m):
        nonlocal dx, dy, note
        direc, num, unit = m.group(1), float(m.group(2)), m.group(3)
        pt = num * (2.8346 if unit in ("毫米",) or unit.lower() == "mm" else 28.346)
        if not direc:
            note = "⚠️偏移没写方向已忽略，请写如: 上挪2厘米 / 左挪1.5厘米"
            return ""
        if "左" in direc:
            dx -= pt
        if "右" in direc:
            dx += pt
        if "上" in direc:
            dy -= pt
        if "下" in direc:
            dy += pt
        return ""

    rest = _OFFSET_RE.sub(_apply, hint).strip(" \t+、,，;；/·")
    return dx, dy, rest, note


def _parse_position_hint(hint: str, total_pages: int):
    """解析 '末页右下' / '第2页左下' 这类提示 → (page_index, corner)"""
    page_idx = total_pages - 1
    m = re.search(r"第\s*(\d+)\s*页", hint)
    if m:
        page_idx = min(max(int(m.group(1)) - 1, 0), total_pages - 1)
    elif "首页" in hint:
        page_idx = 0
    corner = "右下"
    for c in ("右下", "左下", "中下", "右上", "左上", "居中"):
        if c in hint:
            corner = c
            break
    return page_idx, corner


def _corner_center(page_rect, corner: str):
    w, h = page_rect.width, page_rect.height
    half = SEAL_SIZE_PT / 2
    m = STAMP_MARGIN_PT
    xs = {"右": w - m - half, "左": m + half, "中": w / 2, "居": w / 2}
    ys = {"下": h - m - half, "上": m + half, "中": h / 2, "居": h / 2}
    return xs.get(corner[0], w - m - half), ys.get(corner[-1], h - m - half)


STAMP_PARTY_WORDS = ["需方", "买方", "购货方", "购方", "供方", "卖方", "供货方", "销售方", "甲方", "乙方"]


def _our_party_words(doc):
    """判断我方在这份合同里的角色词：找 '角色词…我方品牌名' 同行邻近出现"""
    brand = STAMP_KEYWORDS[1][:2]  # 祥润 / 贤凌
    try:
        text_all = "\n".join(p.get_text() for p in doc)
    except Exception:
        return []
    return [pw for pw in STAMP_PARTY_WORDS
            if re.search(pw + r"[^\n]{0,40}?" + brand, text_all)]


SIGN_CTX_WORDS = ("盖章", "签章", "公章", "签字", "签名", "法定代表人", "授权代表", "委托代理人", "日期")


def _parse_riding(hint):
    """从提示词解析骑缝章开关 → (riding: True/False/None, 去掉骑缝词后的提示)"""
    h = (hint or "").strip()
    riding = None
    for w in ("不骑缝", "无骑缝", "不要骑缝", "取消骑缝", "去骑缝"):
        if w in h:
            riding = False
            h = h.replace(w, "")
    if riding is None and "骑缝" in h:
        riding = True
    h = h.replace("骑缝章", "").replace("骑缝", "").strip(" \t+、,，;；/·")
    return riding, h


def _detect_table_box(page):
    """v5.10 渲染页面找末尾签字表格的边框 → fitz.Rect（pt 坐标），找不到返回 None。

    给扫描件（无文字层）提供版面感知：二值化 → 纵向膨胀±2px（把扫描倾斜
    摊薄的横线接回连续带）→ 横向腐蚀32px（只留长横线，文字笔画全部消失）→
    行覆盖率≥15% 判为表格横线 → 按间隙聚类，取最靠底部、≥3条横线的簇。
    纯 PIL 实现，不依赖 numpy。"""
    import fitz
    from PIL import Image, ImageChops
    dpi = 100
    scale = dpi / 72.0
    pix = page.get_pixmap(dpi=dpi, colorspace=fitz.csGRAY)
    img = Image.frombuffer("L", (pix.width, pix.height), pix.samples,
                           "raw", "L", pix.stride, 1)
    bw = img.point(lambda v: 255 if v < 200 else 0)  # 深色→255
    W, H = bw.size
    dil = bw
    for dy in (1, 2, -1, -2):
        dil = ImageChops.lighter(dil, ImageChops.offset(bw, 0, dy))
    er = dil
    for off in (1, 2, 4, 8, 16):
        er = ImageChops.darker(er, ImageChops.offset(er, off, 0))
    er = er.crop((32, 0, W, H))  # 去掉 offset 横向环绕污染
    eW = er.size[0]
    rows = list(er.resize((1, H), Image.BOX).getdata())
    line_rows = [y for y, v in enumerate(rows) if v >= 255 * 0.15]
    if not line_rows:
        return None
    gap = int(80 * scale)  # 行距超过80pt视为两张表
    clusters, cur = [], [line_rows[0]]
    for y in line_rows[1:]:
        if y - cur[-1] <= gap:
            cur.append(y)
        else:
            clusters.append(cur)
            cur = [y]
    clusters.append(cur)

    def _distinct(c):
        n, last = 1, c[0]
        for y in c:
            if y - last > 4:
                n += 1
            last = y
        return n

    clusters = [c for c in clusters if _distinct(c) >= 3]
    if not clusters:
        return None
    tbl = clusters[-1]  # 最靠底部的表格＝签字栏
    y0, y1 = tbl[0], tbl[-1]
    x0, x1 = None, None
    for y in tbl:  # 表格水平范围＝各横线的最左/最右延伸
        bb = er.crop((0, y, eW, y + 1)).getbbox()
        if not bb:
            continue
        x0 = bb[0] if x0 is None else min(x0, bb[0])
        x1 = bb[2] if x1 is None else max(x1, bb[2])
    if x0 is None:
        return None
    # 补回腐蚀截掉的 32px 起点偏移和横线两端各~16px 的腐蚀损耗
    x0 = max(0, x0 + 32 - 16)
    x1 = min(W, x1 + 32 + 16)
    if (x1 - x0) < W * 0.3:  # 太窄不像签字表格
        return None
    if y1 / scale < page.rect.height * 0.65:  # v5.11 签字表格必然靠下；条款下划线在页中部误检时弃用
        return None
    return fitz.Rect(x0 / scale, y0 / scale, x1 / scale, y1 / scale)


def _corner_center_in_box(box, corner):
    """v5.10 九宫格方位 → 表格框内坐标；左/右取左右半栏中心（贴合双栏签字表）"""
    half = SEAL_SIZE_PT / 2
    inset = 10
    w = box.width
    lx = box.x0 + min(max(w * 0.25, half + inset), w / 2)
    rx = box.x1 - min(max(w * 0.25, half + inset), w / 2)
    xs = {"右": rx, "左": lx, "中": (box.x0 + box.x1) / 2, "居": (box.x0 + box.x1) / 2}
    ys = {"下": box.y1 - half - inset, "上": box.y0 + half + inset,
          "中": (box.y0 + box.y1) / 2, "居": (box.y0 + box.y1) / 2}
    cy = ys.get(corner[-1], box.y1 - half - inset)
    if box.height < SEAL_SIZE_PT + 2 * inset:  # 表格太矮就垂直居中
        cy = (box.y0 + box.y1) / 2
    return xs.get(corner[0], rx), cy


def _auto_locate(doc):
    """返回 (page_idx, cx, cy, method)。

    v5.9 全局打分制（解决两页文档定位不准）：
    不再"从末页往前找到即停"，而是把全文档所有候选点收集后打分取最优：
      基础分  公司全称100 > 简称90 > 我方角色词62 > 盖章/签章/公章标签55
      加分    越靠页面下方越像签字栏(+0~30)；靠后页微弱加分(+0~8)；
              名称/角色词竖向70pt内有盖章·签字·日期等签字栏上下文(+35)；
              章标签40pt内横向邻近我方名称——区分甲乙两栏(+30)；章标签偏右微加(我方常在右列)
      扣分    位于页面顶部22%（抬头/页眉里的公司名）-45
    找不到任何关键词才兜底末页右下。"""
    total = doc.page_count
    our_words = _our_party_words(doc)
    cands = []
    for pi in range(total):
        page = doc[pi]
        ph, pw_ = page.rect.height, page.rect.width
        groups = [(STAMP_KEYWORDS[0], 100, "全称"), (STAMP_KEYWORDS[1], 90, "简称")]
        groups += [(w, 62, "我方角色词") for w in our_words]
        groups += [(w, 55, "章标签") for w in ("盖章", "签章", "公章")]

        ctx_rects = []
        for w in SIGN_CTX_WORDS:
            ctx_rects += page.search_for(w)
        name_rects = page.search_for(STAMP_KEYWORDS[0]) + page.search_for(STAMP_KEYWORDS[1])

        for kw, base, tag in groups:
            for r in page.search_for(kw):
                ry = (r.y0 + r.y1) / 2
                rx = (r.x0 + r.x1) / 2
                y_ratio = r.y1 / ph
                score = base + 30 * y_ratio + 8 * (pi + 1) / total
                if y_ratio < 0.22:
                    score -= 45  # 页眉/合同抬头里的名字不是签字栏
                if tag in ("全称", "简称", "我方角色词"):
                    if any(abs((c.y0 + c.y1) / 2 - ry) < 70 and c != r for c in ctx_rects):
                        score += 35  # 附近有盖章/签字/日期 → 签字栏上下文
                else:  # 章标签
                    if any(abs((n.y0 + n.y1) / 2 - ry) < 40 and abs((n.x0 + n.x1) / 2 - rx) < 180 for n in name_rects):
                        score += 30  # 章标签紧挨我方名称（区分甲乙两栏）
                    score += 5 * rx / pw_  # 对方起草的合同我方通常在右列
                cx = rx + (30 if tag == "我方角色词" else 0)
                cands.append((score, pi, cx, ry + 6, f"{tag}:{kw}(第{pi + 1}页,{int(score)}分)"))
    if cands:
        score, pi, cx, cy, label = max(cands, key=lambda c: c[0])
        return pi, cx, cy, "打分定位:" + label
    page = doc[total - 1]
    try:
        box = _detect_table_box(page)  # v5.10 扫描件按签字表格边框兜底
    except Exception:
        box = None
    if box:
        cx, cy = _corner_center_in_box(box, "右下")
        return total - 1, cx, cy, "兜底:末页签字表格框内右下（扫描件按表格边框定位，请核对我方栏位）"
    cx, cy = _corner_center(page.rect, "右下")
    return total - 1, cx, cy, "兜底:末页右下（扫描件或未找到关键词，请仔细核对位置）"


def _add_riding_seal(doc, fitz):
    """骑缝章：公章竖切 N 条，第 i 条贴第 i 页右缘垂直居中，叠页可拼回完整章"""
    from PIL import Image
    import io
    n = doc.page_count
    if n < 2:
        return False
    img = Image.open(STAMP_SEAL_PATH).convert("RGBA")
    W, H = img.size
    for i in range(n):
        left = int(round(i * W / n))
        right = int(round((i + 1) * W / n))
        if right <= left:
            continue
        buf = io.BytesIO()
        img.crop((left, 0, right, H)).save(buf, format="PNG")
        page = doc[i]
        pr = page.rect
        pt_w = SEAL_SIZE_PT * (right - left) / W
        cy = pr.height * 0.5
        rect = fitz.Rect(pr.width - pt_w, cy - SEAL_SIZE_PT / 2, pr.width, cy + SEAL_SIZE_PT / 2)
        if page.rotation:  # v5.11 旋转页：显示坐标 → 物理坐标
            rect = rect * page.derotation_matrix
        page.insert_image(rect, stream=buf.getvalue(), overlay=True,
                          rotate=page.rotation)
    return True


@app.post("/stamp")
def stamp(req: StampRequest):
    """对方合同盖章：PDF/Word/图片 → 定位签字栏盖透明公章 → 返回盖章版 PDF

    定位优先级：手动 position_hint > 关键词（从末页往前：公司全称/简称/盖章/签章/公章）> 兜底末页右下
    """
    try:
        import fitz  # PyMuPDF，延迟导入，安装异常不影响 /generate
    except ImportError:
        raise HTTPException(500, "PyMuPDF 未安装：requirements.txt 需含 PyMuPDF 并重新 build")

    if not os.path.exists(STAMP_SEAL_PATH):
        raise HTTPException(500, f"公章文件缺失: {STAMP_SEAL_PATH}")

    try:
        data = base64.b64decode(req.file_base64)
        ext = (req.file_ext or "pdf").lower().lstrip(".")

        # 1. 统一转 PDF
        if ext in ("png", "jpg", "jpeg", "bmp", "webp"):
            imgdoc = fitz.open(stream=data, filetype=ext)
            data = imgdoc.convert_to_pdf()
            imgdoc.close()
        elif ext in ("doc", "docx"):
            data = _office_to_pdf(data, ext)
        elif ext != "pdf":
            raise HTTPException(400, f"不支持的文件类型: {ext}")

        doc = fitz.open(stream=data, filetype="pdf")
        total = doc.page_count

        # 2. 解析骑缝章开关（提示词里的"骑缝/不骑缝"），并把骑缝词从定位提示里剥掉
        riding_flag, hint_clean = _parse_riding(req.position_hint)
        if req.riding_seal is not None:
            riding_flag = req.riding_seal
        do_riding = riding_flag if riding_flag is not None else (total >= 2)  # 多页默认加骑缝章

        # 3. 定位盖章中心点（剥掉骑缝词后提示为空 → 走自动定位）
        if hint_clean:
            dx, dy, hint_pos, off_note = _parse_offsets(hint_clean)  # v5.11 先剥偏移词
            page_idx, corner = _parse_position_hint(hint_pos or hint_clean, total)
            page = doc[page_idx]
            try:
                box = _detect_table_box(page)  # v5.10 方位以签字表格边框为基准
            except Exception:
                box = None
            if box:
                cx, cy = _corner_center_in_box(box, corner)
                method = f"手动指定:{hint_clean}(按表格边框定位)"
            else:
                cx, cy = _corner_center(page.rect, corner)
                method = f"手动指定:{hint_clean}"
            if dx or dy:
                cx += dx
                cy += dy
                method += f"(已按偏移调整)"
            if off_note:
                method += f" {off_note}"
        else:
            page_idx, cx, cy, method = _auto_locate(doc)  # v5.9 全局打分定位
            page = doc[page_idx]

        # 4. 盖主章（章不出页边）
        half = SEAL_SIZE_PT / 2
        pr = page.rect
        cx = min(max(cx, half + 4), pr.width - half - 4)
        cy = min(max(cy, half + 4), pr.height - half - 4)
        rect = fitz.Rect(cx - half, cy - half, cx + half, cy + half)
        # v5.11 旋转页修正：定位算的是显示坐标，insert_image 要物理坐标（横放扫描件 /Rotate≠0 时不转会盖出页外）
        if page.rotation:
            rect = rect * page.derotation_matrix
        page.insert_image(rect, filename=STAMP_SEAL_PATH, overlay=True,
                          rotate=page.rotation)

        # 5. 骑缝章（≥2页默认；提示"不骑缝"关闭；单页自动跳过）
        riding_done = _add_riding_seal(doc, fitz) if do_riding else False
        if riding_done:
            method += f" +骑缝章({total}页)"

        out = doc.tobytes(deflate=True, garbage=3)
        doc.close()
        return {
            "ok": True,
            "pdf_base64": base64.b64encode(out).decode(),
            "method": method,
            "page": page_idx + 1,
            "total_pages": total,
            "riding_seal": riding_done,
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f"盖章失败: {e}")


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=int(os.getenv("PORT", 8000)))
