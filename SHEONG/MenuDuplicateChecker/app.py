# -*- coding: utf-8 -*-
"""
주간 식단표 중복/충돌 검사기 — Hugging Face Spaces (Gradio) 버전
====================================================================
- 매주 바뀌는 주간 식단표 엑셀(.xlsx)을 올리면, 여러 규칙 위반을 걸러 준다.
- 식단 데이터는 코드에 고정돼 있지 않다. 업로드한 엑셀 내용만 검사한다.
- 날짜 수에 맞춰 7일씩 자동으로 '주(week)'를 나눠 주간 규칙(오리/순대/잡채 등)을 적용한다.

엑셀 표 형태:
--------------------------------------------------------------------
· 2행: 날짜가 가로로 나열 (07/06, 07/07, …)
· A열: 조 식 / 중 식 / 석 식 라벨
· 각 끼니 칸: 쌀밥/잡곡밥 → 국 → 반찬 여러 개 → 포기김치, 중식엔 죽 한 줄
· 쌀밥/잡곡밥·포기김치는 규칙 대상에서 자동 제외
--------------------------------------------------------------------
(과거 텍스트 붙여넣기 방식 parse_menu_text 도 함수로 남겨 뒀다 — 지금 UI는 엑셀 업로드)
"""

import re
import traceback
from collections import defaultdict
from datetime import date, datetime

import gradio as gr
import openpyxl

# --- Gradio 버그 우회 -------------------------------------------------------
# 일부 Gradio/gradio_client 버전은 API 스키마 생성 시 스키마 값이 불리언이면
# "argument of type 'bool' is not iterable" 오류로 앱이 시작조차 못 한다.
try:
    import gradio_client.utils as _gcu

    _orig_js2pt = _gcu._json_schema_to_python_type

    def _safe_js2pt(schema, *args, **kwargs):
        if isinstance(schema, bool):
            return "Any"
        return _orig_js2pt(schema, *args, **kwargs)

    _gcu._json_schema_to_python_type = _safe_js2pt

    if hasattr(_gcu, "get_type"):
        _orig_get_type = _gcu.get_type

        def _safe_get_type(schema):
            if not isinstance(schema, dict):
                return "Any"
            return _orig_get_type(schema)

        _gcu.get_type = _safe_get_type
except Exception:
    pass
# ---------------------------------------------------------------------------

MEAL_ORDER = ["조식", "중식", "석식"]

# ------------------------------------------------------------------
# 재료/유형 분류 사전 (필요하면 자유롭게 추가하세요)
# 매주 새 메뉴명이 나오면, 아래 사전에 없는 항목은 '분류 실패'로 결과 하단에 뜬다.
# ------------------------------------------------------------------

FISH_KW = ["고등어", "삼치", "갈치", "꽁치", "가자미", "코다리", "우럭", "조기",
           "동태", "황태", "임연수", "전어", "박대", "도미", "방어", "생선"]
# 멸치는 육수/잔멸치 반찬으로 거의 매일 쓰이는 재료라 '메인 생선요리'로 보지 않음(제외).

PORK_KW = ["돈", "제육", "돼지", "삼겹", "목살", "돈육", "탕수육", "동그랑땡"]
BEEF_KW = ["소고기", "쇠고기", "우육", "우채", "우불고기", "언양불고기", "스테이크"]
# ※ '돈채'는 PORK_KW의 '돈'으로, '우채'는 위 '우채'로 잡힌다 (채 썬 고기).
CHICKEN_KW = ["닭", "치킨"]
DUCK_KW = ["오리"]
SUNDAE_KW = ["순대"]
JAPCHAE_KW = ["잡채"]   # 주 1회만
CURRY_KW = ["카레"]     # 주 1회만
JJAJANG_KW = ["짜장"]   # 주 1회만
MANDU_KW = ["만두", "교자"]        # 물만두/교자만두/만두찜 등 -> 주 1회만
KKASU_KW = ["까스", "튀김"]        # 기성 튀김류(까스·튀김) -> 주 3회 이상 금지(=최대 2회)
ALTANG_KW = ["알탕", "매운탕"]     # 알탕·매운탕 같은 취급 -> 합쳐서 주 1회만
JWIEOCHAE_KW = ["쥐어채", "명엽채"]  # 쥐어채·명엽채 비슷 -> 합쳐서 주 1~2회
DORAJI_KW = ["도라지"]            # 주 최대 2회
# 면류 -> 주 1회만
NOODLE_KW = ["국수", "냉면", "스파게티", "파스타", "우동", "라면", "쫄면", "소바", "짬뽕"]
# 누룽지탕 = 숭늉 (같은 메뉴, 보통 조식). 주 최대 2회 + 최소 3일 간격
NURUNGJI_KW = ["누룽지", "숭늉"]

# '주 1회' 같은 규칙은 달력상의 주로 자르면 일요일↔다음주 월요일을 놓친다.
# 그래서 '연속 21끼니(7일)' 이동 창으로 센다.
WEEK_MEALS = 21
# 기성 떡갈비류: 같은 날 금지 + 하루 걸러서만 (연이틀 금지). 산적도 같은 취급.
TTEOKGALBI_KW = ["너비아니", "떡갈비", "갈비경단", "경단", "산적"]
MUSHROOM_KW = ["버섯", "느타리", "새송이", "양송이", "표고", "팽이", "목이"]
JANGJORIM_KW = ["장조림"]
NAMUL_SUFFIX = ["나물", "무침", "무생채"]
NAMUL_BASE_HINTS = ["청경채", "콩나물", "숙주", "고사리", "미나리", "시금치", "열무",
                    "노각", "가지", "비름", "도라지", "무말랭이", "무짠지", "참나물",
                    "오복채", "낙지젓알", "들깨무", "우채", "명엽채"]

# 영양상 거의 매 끼니 들어가는 기본 채소들 -> 겹쳐도 정상이라 문제로 잡지 않고,
# '분류 실패' 목록에서도 제외한다. (필요하면 추가하세요)
# ※ 감자·도토리묵·청포묵은 '메인 재료라 겹치면 안 됨' 요청에 따라
#    필러에서 빼고 SIDE_GAP_RULES로 옮김.
FILLER_KW = ["당근", "양파", "야채", "피망", "파프리카", "양배추", "대파",
             "브로콜리", "애호박", "호박", "고추", "부추", "콩"]

# 나물 외에 '너무 자주 겹치면 안 되는' 반찬 주재료.
# (이름, 키워드들, within, count_soup)
#   within  = 이 끼니 수 안에 다시 나오면 위반 (끼니 인덱스 차이 <= within 이면 걸림)
#   count_soup = 국에 나와도 카운트할지 (묵류는 냉국도 메인 사용으로 봄)
SIDE_GAP_RULES = [
    ("계란",    ["계란", "달걀", "메란", "메추리알"], 6, False),  # 메란=메추리알=계란류
    ("두부",    ["두부"], 6, False),                # 순두부 포함
    ("멸치",    ["멸치"], 6, False),                # 국물용 멸치(국)는 제외
    ("마늘쫑",  ["마늘쫑", "마늘종"], 6, False),
    ("어묵",    ["어묵"], 6, False),                # 어묵탕(국)은 제외
    ("감자",    ["감자"], 6, False),
    ("가지",    ["가지"], 6, False),
    ("고구마",  ["고구마"], 6, False),
    ("옥수수",  ["옥수수", "콘"], 6, False),   # 옥수수=콘 같은 재료

    # ※ 장조림은 계란장조림/돈육장조림처럼 종류가 다르면 다른 메뉴라
    #    아래 classify_jangjorim_type 으로 종류별('장조림:계란' 등)로 따로 센다.
    ("기성품전", ["해물완자전", "동그랑땡", "오색전"], 6, False),  # 기성 냉동 전류
    ("미트볼",  ["미트볼"], 6, False),
    ("연근",    ["연근"], 5, False),
    ("굴소스",  ["굴소스"], 5, False),
    ("소시지",  ["소시지", "비엔나", "후랑크", "프랑크"], 5, False),
    # ※ 묵류는 종류가 다양해서(도토리묵/청포묵/동부묵…) 아래 check_muk_rules 로 따로 처리
]
WITHIN_MAP = {name: within for name, kws, within, soup in SIDE_GAP_RULES}
NAMUL_WITHIN = 6  # 나물류는 6끼니 안 반복 금지

# '까스/튀김'류는 기성 가공식품을 그대로 쓰는 경우가 많아 신선육 메인요리와
# 다른 카테고리로 취급 -> 그날의 '메인재료'로 세지 않음.
KKASU_PATTERN = re.compile(r"까스|튀김")

# 국 종류가 같은지 비교할 때 쓰는 키워드
SOUP_TYPE_KW = ["된장", "김치", "육개장", "미역", "뭇국", "콩나물", "계란", "만두",
                "어묵", "순두부", "백탕", "들깨", "도토리묵", "아욱", "수제비",
                "근대", "부대찌개", "시금치", "건새우", "우거지", "냉국", "탕국",
                "곰탕", "시락국", "무채", "무국"]

# 분류 실패 후보를 찾을 때 '메인요리처럼 보이는' 접미어
MAIN_DISH_SUFFIX = ["조림", "볶음", "구이", "찜", "전", "까스", "튀김", "탕",
                    "스테이크", "장조림"]


def classify_protein(name):
    """돼지/소/닭/오리/생선/순대 등 메인 재료 카테고리 판별. set 반환."""
    cats = set()
    if KKASU_PATTERN.search(name):
        return cats  # 기성품 튀김류 -> 메인재료로 세지 않음
    if any(k in name for k in PORK_KW):
        cats.add("돼지")
    if any(k in name for k in FISH_KW):
        cats.add("생선")
    if any(k in name for k in CHICKEN_KW):
        cats.add("닭")
    if any(k in name for k in DUCK_KW):
        cats.add("오리")
    if any(k in name for k in SUNDAE_KW):
        cats.add("순대")
    if any(k in name for k in BEEF_KW) and "돼지" not in cats:
        cats.add("소")
    return cats


def classify_mushroom(name):
    return {k for k in MUSHROOM_KW if k in name}


def classify_tofu(name):
    if "순두부" in name:
        return "순두부"
    if "두부" in name:
        return "두부"
    return None


def classify_jangjorim(name):
    return any(k in name for k in JANGJORIM_KW)


# 장조림 종류: 계란장조림과 돈육장조림은 서로 다른 메뉴로 본다.
# 메란장조림·메알장조림·계란장조림은 모두 같은 '계란장조림'류.
JANGJORIM_TYPE_KW = [
    ("계란", ["계란", "달걀", "메란", "메알", "메추리알"]),
    ("돈육", ["돈육", "돼지", "돈"]),
    ("닭",   ["닭", "치킨"]),
    ("우육", ["우육", "소고기", "쇠고기", "우채"]),
]


def classify_jangjorim_type(name):
    """장조림이면 종류('계란'/'돈육'/'닭'/'우육'/'기타')를 돌려준다. 아니면 None."""
    if "장조림" not in name:
        return None
    for key, kws in JANGJORIM_TYPE_KW:
        if any(k in name for k in kws):
            return key
    return "기타"


# 주 N회 등으로 따로 관리해서 '나물 간격' 검사에서는 빼는 재료
NAMUL_GAP_EXEMPT = {"도라지"}   # 도라지는 규칙21(주 2회)로 관리


def classify_namul_base(name):
    if not any(s in name for s in NAMUL_SUFFIX):
        return None
    base = None
    for hint in NAMUL_BASE_HINTS:
        if hint in name:
            base = hint
            break
    if base is None:
        base = re.split("|".join(NAMUL_SUFFIX), name)[0]
        base = base if base else name
    if base in NAMUL_GAP_EXEMPT:
        return None
    return base


def extract_tracked_side_keys(name):
    """반찬 하나에서 '너무 자주 겹치면 안 되는' 주재료 키를 뽑는다.
    - 계란/두부/멸치/소시지/묵/장조림 등은 SIDE_GAP_RULES 이름으로
    - 위에 안 걸리면 나물류는 '나물:<기본재료>' 형태로 (묵무침이 나물로 중복되지 않게)
    반환: set (없으면 빈 set)
    """
    keys = set()
    for cat, kws, within, soup in SIDE_GAP_RULES:
        if any(k in name for k in kws):
            keys.add(cat)
    jt = classify_jangjorim_type(name)   # 계란장조림/돈육장조림은 서로 다른 메뉴
    if jt:
        keys.add("장조림:" + jt)
    if not keys and not extract_muk_type(name):  # 묵류는 check_muk_rules 담당
        base = classify_namul_base(name)
        if base:
            keys.add("나물:" + base)
    return keys


def soup_tracked_keys(name):
    """국에 나와도 카운트하는 재료 키(현재 없음, 확장용)."""
    return {cat for cat, kws, within, soup in SIDE_GAP_RULES
            if soup and any(k in name for k in kws)}


# '묵'으로 끝나는 재료명(도토리묵/청포묵/동부묵/메밀묵…)만 뽑는다.
# '묵은지'처럼 묵으로 시작하는 단어는 앞에 글자가 없어 매칭되지 않는다.
MUK_RE = re.compile(r"[가-힣]+묵")


def extract_muk_type(name):
    """반찬/국 이름에서 묵(젤리) 종류를 뽑는다. 예: '도토리묵냉국' -> '도토리묵'.
    '어묵/붕어묵/봉어묵'은 생선살이라 묵류가 아니므로 제외. 없으면 None."""
    for m in MUK_RE.finditer(name):
        t = m.group(0)
        if "어묵" in t:      # 어묵류(생선살)는 묵(젤리) 아님
            continue
        return t
    return None


def is_filler_only(name):
    """기본 채소로만 이뤄진 반찬인지(=겹쳐도 되는지) 대략 판단."""
    if not any(k in name for k in FILLER_KW):
        return False
    # 필러 채소가 들어있고, 고기/생선/추적재료가 아니면 필러로 본다
    if classify_protein(name):
        return False
    if extract_tracked_side_keys(name):
        return False
    return True


# ------------------------------------------------------------------
# 입력 텍스트 -> MENU(dict) 파서
# ------------------------------------------------------------------

DAY_RE = re.compile(r"^\d{1,2}\s*/\s*\d{1,2}")  # 07/06, 7/6 ...


def parse_menu_text(text):
    """입력 텍스트를 MENU(순서 있는 dict)로 변환.
    반환: (menu, warnings)
    menu = {day: {meal: {"국": str|None, "찬": [..], "죽": str|None}}}
    """
    menu = {}
    warns = []
    cur_day = None

    for lineno, raw in enumerate(text.splitlines(), 1):
        line = raw.strip()
        if not line:
            continue
        # 주석(#)이나 구분선(---)은 건너뜀
        if line.startswith("#") or set(line) <= set("-=·—"):
            continue

        first_tok = line.split("|", 1)[0].strip()

        # 끼니 줄인가?
        meal_name = None
        for m in MEAL_ORDER:
            if first_tok.startswith(m):
                meal_name = m
                break

        if meal_name is None:
            # 날짜(요일) 줄로 취급
            if DAY_RE.match(line) or "(" in line:
                cur_day = line
                menu.setdefault(cur_day, {})
            else:
                warns.append(f"[{lineno}행] 날짜도 끼니도 아니라 무시함: {line}")
            continue

        if cur_day is None:
            warns.append(f"[{lineno}행] 날짜 줄이 먼저 나와야 합니다. 무시함: {line}")
            continue

        soup = None
        porridge = None
        dishes = []
        parts = [p.strip() for p in line.split("|")]
        for p in parts[1:]:  # parts[0]은 끼니명
            if not p:
                continue
            if p.startswith("국:") or p.startswith("국："):
                v = p.split(":", 1)[-1].split("：", 1)[-1].strip()
                soup = None if v in ("", "없음", "-", "X", "x") else v
            elif p.startswith("죽:") or p.startswith("죽："):
                v = p.split(":", 1)[-1].split("：", 1)[-1].strip()
                porridge = v or None
            else:
                if p.startswith("찬:") or p.startswith("찬："):
                    p = p.split(":", 1)[-1].split("：", 1)[-1].strip()
                for d in re.split(r"[,，]", p):
                    d = d.strip()
                    if d:
                        dishes.append(d)

        menu[cur_day][meal_name] = {"국": soup, "찬": dishes, "죽": porridge}

    return menu, warns


# ------------------------------------------------------------------
# 엑셀(.xlsx) -> MENU 파서
#   표 형태: 2행에 날짜(열별 7일), A열에 조식/중식/석식 라벨,
#   각 끼니 블록은 [쌀밥 → 국 → 반찬 여러 개 → 포기김치], 중식엔 죽 한 줄.
#   쌀밥/잡곡밥, 포기김치는 규칙 대상에서 제외한다.
# ------------------------------------------------------------------

STAPLE = {"쌀밥/잡곡밥", "쌀밥", "잡곡밥"}   # 주식 -> 제외
EXCLUDE = {"포기김치"}                        # 규칙 제외 대상
WEEKDAY_KR = ["월", "화", "수", "목", "금", "토", "일"]


def _clean_cell(v):
    if v is None:
        return ""
    return str(v).replace("\n", " ").strip()


def parse_one_sheet(ws, result, warns):
    """워크시트 하나를 읽어 result(dict: date -> {meal: {...}})에 채운다."""
    # 1) 날짜 행 찾기 (날짜형 셀이 3개 이상인 첫 행)
    date_cols, date_row = {}, None
    for r in range(1, min(ws.max_row, 12) + 1):
        found = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if isinstance(v, datetime):
                found[c] = v.date()
            elif isinstance(v, date):
                found[c] = v
        if len(found) >= 3:
            date_row, date_cols = r, found
            break
    if not date_cols:
        warns.append(f"[{ws.title}] 날짜 행(가로로 나열된 날짜)을 못 찾았습니다.")
        return
    first_date_col = min(date_cols)

    # 2) 끼니 라벨 행 찾기 (날짜 열 왼쪽 칸에서 '조식/중식/석식')
    meal_rows = []
    for r in range(date_row + 1, ws.max_row + 1):
        for c in range(1, first_date_col):
            s = _clean_cell(ws.cell(r, c).value).replace(" ", "")
            if s in MEAL_ORDER:
                meal_rows.append((r, s))
                break
    if not meal_rows:
        warns.append(f"[{ws.title}] 조식/중식/석식 라벨을 못 찾았습니다.")
        return

    # 3) 끼니 블록별로 국/찬/죽 추출
    for i, (start, meal) in enumerate(meal_rows):
        end = meal_rows[i + 1][0] - 1 if i + 1 < len(meal_rows) else ws.max_row
        for c, dt in date_cols.items():
            cells = [(r, _clean_cell(ws.cell(r, c).value)) for r in range(start, end + 1)]
            # 주식(쌀밥) 바로 아랫줄을 '국'으로 본다.
            staple_idx = next((k for k, (r, v) in enumerate(cells) if v in STAPLE), None)
            soup, dishes, porridge = None, [], None
            for k, (r, v) in enumerate(cells):
                if not v or v in STAPLE or v in EXCLUDE:
                    continue
                if staple_idx is not None and k == staple_idx + 1:
                    soup = v
                    continue
                if v.endswith("죽"):
                    porridge = v
                    continue
                dishes.append(v)
            result.setdefault(dt, {})[meal] = {"국": soup, "찬": dishes, "죽": porridge}


def parse_menu_xlsx(paths):
    """엑셀 파일 여러 개를 읽어 날짜순으로 정렬된 MENU(dict)로 만든다.
    반환: (menu, warnings)"""
    result = {}   # date -> {meal: {...}}
    warns = []
    for p in paths:
        try:
            wb = openpyxl.load_workbook(p, data_only=True)
        except Exception as e:
            warns.append(f"[열기 실패] {p} ({e})")
            continue
        for ws in wb.worksheets:
            parse_one_sheet(ws, result, warns)

    menu = {}
    for dt in sorted(result):
        label = f"{dt.month:02d}/{dt.day:02d}({WEEKDAY_KR[dt.weekday()]})"
        menu[label] = result[dt]
    return menu, warns


def _to_paths(files):
    """Gradio 업로드 값에서 실제 파일 경로 리스트를 뽑는다.
    (버전에 따라 str / dict / 파일객체로 올 수 있어 모두 대응)"""
    paths = []
    for f in files or []:
        if isinstance(f, str):
            paths.append(f)
        elif isinstance(f, dict):
            paths.append(f.get("path") or f.get("name"))
        else:
            p = getattr(f, "name", None)
            if p:
                paths.append(p)
    return [p for p in paths if p]


# ------------------------------------------------------------------
# MENU -> 검사 컨텍스트(타임라인/주 구분)
# ------------------------------------------------------------------

def build_context(menu):
    days = list(menu.keys())
    timeline = []
    idx = 0
    for day in days:
        meals = menu[day]
        for meal in MEAL_ORDER:
            info = meals.get(meal)
            if not info:
                continue
            timeline.append({
                "idx": idx, "day": day, "meal": meal,
                "soup": info.get("국"),
                "dishes": list(info.get("찬", [])),
                "porridge": info.get("죽"),
            })
            idx += 1

    # 날짜를 7일씩 묶어 주(week)로 나눔 (14일 고정 아님)
    weeks = []
    for i in range(0, len(days), 7):
        chunk = days[i:i + 7]
        if len(days) <= 7:
            wname = "이번 주"
        else:
            wname = f"{i // 7 + 1}주차"
        weeks.append((chunk, wname))

    return {"menu": menu, "days": days, "timeline": timeline, "weeks": weeks}


# ------------------------------------------------------------------
# 규칙별 검사 (모두 ctx를 받음)
# ------------------------------------------------------------------

def check_rule1_soup_type_same_day(ctx):
    """1. 같은 날 조/중/석식 국의 '종류'가 같으면 안 됨."""
    issues = []
    for day, meals in ctx["menu"].items():
        soups = {m: meals[m].get("국") for m in MEAL_ORDER
                 if meals.get(m) and meals[m].get("국")}
        keys = list(soups.items())
        for i in range(len(keys)):
            for j in range(i + 1, len(keys)):
                m1, s1 = keys[i]
                m2, s2 = keys[j]
                common = {kw for kw in SOUP_TYPE_KW if kw in s1 and kw in s2}
                if common:
                    issues.append(f"[규칙1] {day} {m1}({s1}) vs {m2}({s2}) - 같은 국 유형 겹침: {common}")
    return issues


def check_rule2_3_consecutive_protein(ctx):
    """2,3. 연속된 끼니에 같은 메인 재료 반복 금지 (전날 석식->다음날 조식 포함)."""
    issues = []
    tl = ctx["timeline"]
    protein_by_idx = []
    for t in tl:
        cats = set()
        for d in [t["soup"]] + t["dishes"]:
            if d:
                cats |= classify_protein(d)
        protein_by_idx.append(cats)

    for i in range(len(tl) - 1):
        overlap = protein_by_idx[i] & protein_by_idx[i + 1]
        if overlap:
            t1, t2 = tl[i], tl[i + 1]
            issues.append(f"[규칙2/3] {t1['day']} {t1['meal']} -> {t2['day']} {t2['meal']} : 연속 끼니 동일 재료 반복 ({overlap})")
    return issues


def check_rule4_fish_gap(ctx):
    """4. 생선은 최소 3끼니는 건너뛰어야 함 (인덱스 차이 >= 4)."""
    issues = []
    fish = []
    for t in ctx["timeline"]:
        cats = set()
        for d in [t["soup"]] + t["dishes"]:
            if d:
                cats |= classify_protein(d)
        if "생선" in cats:
            fish.append(t)
    for i in range(len(fish) - 1):
        gap = fish[i + 1]["idx"] - fish[i]["idx"]
        if gap < 4:
            issues.append(f"[규칙4] {fish[i]['day']} {fish[i]['meal']} -> "
                          f"{fish[i+1]['day']} {fish[i+1]['meal']} : 생선 간격 부족 ({gap}끼니 차이, 최소 4 필요)")
    return issues


def check_window_limit(ctx, keyword_set, label, limit=1, window=WEEK_MEALS):
    """'주 N회' 규칙. 달력상의 주로 자르지 않고 연속 {window}끼니(기본 7일) 이동 창으로 센다.
    -> 일요일 조식 ↔ 다음주 월요일 석식처럼 주 경계를 넘는 중복도 잡힌다."""
    occ = []
    for t in ctx["timeline"]:
        for d in [t["soup"]] + t["dishes"]:
            if d and any(k in d for k in keyword_set):
                occ.append((t["idx"], t["day"], t["meal"], d))
    occ.sort()

    issues = []
    for i in range(len(occ) - limit):
        first, last = occ[i], occ[i + limit]
        if last[0] - first[0] < window:
            grp = occ[i:i + limit + 1]
            detail = ", ".join(f"{g[1]}{g[2]}({g[3]})" for g in grp)
            issues.append(f"[{label}] {window // 3}일 안에 {limit}회 초과 ({limit + 1}회): {detail}")
    return issues


def check_rule7_mushroom_same_day(ctx):
    """7. 버섯류는 하루에 1끼니 넘게 겹치면 안 됨."""
    issues = []
    for day, meals in ctx["menu"].items():
        found = defaultdict(list)
        for m in MEAL_ORDER:
            info = meals.get(m)
            if not info:
                continue
            for d in [info.get("국")] + info.get("찬", []):
                if d:
                    for kw in classify_mushroom(d):
                        found[kw].append((m, d))
        for kw, occ in found.items():
            if len(occ) > 1:
                issues.append(f"[규칙7] {day}: 버섯류 '{kw}' 하루 중복 -> {occ}")
    return issues


def check_rule8_soondubu_tofu(ctx):
    """8. 순두부국이 있으면 같은 끼니에 두부 반찬 금지."""
    issues = []
    for day, meals in ctx["menu"].items():
        for m in MEAL_ORDER:
            info = meals.get(m)
            if not info:
                continue
            soup = info.get("국") or ""
            if classify_tofu(soup) == "순두부":
                for d in info.get("찬", []):
                    if classify_tofu(d) == "두부":
                        issues.append(f"[규칙8] {day} {m}: 국({soup}) + 반찬({d}) 두부 중복")
    return issues


def check_rule9_soup_dish_overlap(ctx):
    """9. 국과 반찬의 메인 재료가 겹치면 안 됨."""
    issues = []
    for day, meals in ctx["menu"].items():
        for m in MEAL_ORDER:
            info = meals.get(m)
            if not info:
                continue
            soup = info.get("국")
            if not soup:
                continue
            soup_cats = classify_protein(soup)
            if not soup_cats:
                continue
            for d in info.get("찬", []):
                overlap = soup_cats & classify_protein(d)
                if overlap:
                    issues.append(f"[규칙9] {day} {m}: 국({soup}) vs 반찬({d}) 재료 중복 ({overlap})")
    return issues


def check_rule10_side_gap(ctx):
    """10. 같은 반찬 주재료가 정해진 끼니 수 안에 다시 나오면 안 됨.
       재료별 기준(within)은 SIDE_GAP_RULES 참고. 기본 채소(당근/양파/야채)는 제외.
       묵류는 국(냉국)에 나와도 카운트하고, 여러 주 파일을 올리면 주 경계도 이어서 본다."""
    # 같은 두 메뉴가 여러 키로 중복 보고되지 않게(예: '계란' + '계란장조림')
    # (앞끼니, 뒤끼니, 앞메뉴, 뒤메뉴)별로 가장 구체적인 메시지 하나만 남긴다.
    best = {}
    seen = defaultdict(list)
    idx_lookup = {(t["day"], t["meal"]): t["idx"] for t in ctx["timeline"]}
    for day, meals in ctx["menu"].items():
        for m in MEAL_ORDER:
            info = meals.get(m)
            if not info:
                continue
            for d in info.get("찬", []):
                for key in extract_tracked_side_keys(d):
                    seen[key].append((idx_lookup[(day, m)], day, m, d))
            soup = info.get("국")
            if soup:
                for key in soup_tracked_keys(soup):  # 묵류만 국도 카운트
                    seen[key].append((idx_lookup[(day, m)], day, m, soup))
    for key, occ in seen.items():
        occ.sort()
        if key.startswith("나물:"):
            within, rank = NAMUL_WITHIN, 1
            label = f"나물 '{key.split(':', 1)[1]}'"
        elif key.startswith("장조림:"):
            within, rank = 6, 2          # 계란장조림 등 더 구체적인 쪽을 우선
            label = f"'{key.split(':', 1)[1]}장조림'"
        else:
            within, rank = WITHIN_MAP[key], 0
            label = f"'{key}'"
        for i in range(len(occ) - 1):
            gap = occ[i + 1][0] - occ[i][0]
            if gap <= within:
                pair = (occ[i][0], occ[i + 1][0], occ[i][3], occ[i + 1][3])
                msg = (f"[규칙10] {label} 반복: {occ[i][1]}{occ[i][2]}({occ[i][3]}) -> "
                       f"{occ[i+1][1]}{occ[i+1][2]}({occ[i+1][3]}) "
                       f"(간격 {gap}끼니 — {within}끼니 안에 다시 나옴)")
                if pair not in best or rank > best[pair][0]:
                    best[pair] = (rank, msg)
    return [msg for _, msg in best.values()]


def check_rule11_jangjorim_weekly(ctx):
    """11. 장조림류는 주 2~3회만."""
    issues = []
    for week_days, wname in ctx["weeks"]:
        occ = []
        for t in ctx["timeline"]:
            if t["day"] not in week_days:
                continue
            for d in [t["soup"]] + t["dishes"]:
                if d and classify_jangjorim(d):
                    occ.append((t["day"], t["meal"], d))
        # 부분 주(7일 미만)는 횟수 규칙을 강하게 적용하지 않음
        full_week = len(week_days) == 7
        if full_week and not (2 <= len(occ) <= 3):
            issues.append(f"[규칙11] {wname}: 장조림류 {len(occ)}회 (권장 2~3회) -> {occ}")
    return issues


MUK_WITHIN = 5  # 묵류(종류 무관)는 5끼니 안에 다시 나오면 안 됨


def check_muk_rules(ctx):
    """묵 규칙:
    (가) 묵류(종류 무관)는 5끼니 안에 다시 나오면 안 됨. 냉국 등 국도 카운트.
    (나) 같은 종류의 묵은 한 주에 한 번만.
    여러 주 파일을 올리면 (가)는 주 경계도 이어서 본다."""
    issues = []
    idx_lookup = {(t["day"], t["meal"]): t["idx"] for t in ctx["timeline"]}
    occ = []  # (idx, day, meal, dish, muktype)
    for day, meals in ctx["menu"].items():
        for m in MEAL_ORDER:
            info = meals.get(m)
            if not info:
                continue
            for d in info.get("찬", []) + ([info["국"]] if info.get("국") else []):
                mt = extract_muk_type(d)
                if mt:
                    occ.append((idx_lookup[(day, m)], day, m, d, mt))
    occ.sort()

    # (가) 종류 무관 간격
    for i in range(len(occ) - 1):
        gap = occ[i + 1][0] - occ[i][0]
        if gap <= MUK_WITHIN:
            a, b = occ[i], occ[i + 1]
            issues.append(f"[규칙-묵간격] 묵류 반복: {a[1]}{a[2]}({a[3]}) -> "
                          f"{b[1]}{b[2]}({b[3]}) (간격 {gap}끼니 — {MUK_WITHIN}끼니 안에 다시 나옴)")

    # (나) 같은 종류 묵은 7일(21끼니) 안에 한 번만
    by_type = defaultdict(list)
    for idx, day, m, d, mt in occ:
        by_type[mt].append((idx, day, m, d))
    for mt, lst in by_type.items():
        lst.sort()
        for i in range(len(lst) - 1):
            if lst[i + 1][0] - lst[i][0] < WEEK_MEALS:
                a, b = lst[i], lst[i + 1]
                issues.append(f"[규칙-묵주간] 같은 묵 '{mt}' 7일 안에 2회: "
                              f"{a[1]}{a[2]}({a[3]}) -> {b[1]}{b[2]}({b[3]})")
    return issues


def menu_identity(name):
    """'같은 메뉴'로 볼 식별자를 만든다.
    - '생선이름 + 조림'은 부재료가 달라도 같은 메뉴 (삼치조림 = 삼치무조림, 고등어조림 = 고등어무조림)
    - 그 밖에는 메뉴명 그대로 (이름이 완전히 같으면 같은 메뉴, 예: 다시마부각)"""
    if "조림" in name:
        for f in FISH_KW:
            if f == "생선":
                continue
            if f in name:
                return f + "조림"
    return name


def check_same_menu(ctx):
    """같은 메뉴는 7일(21끼니) 안에 두 번 나오면 안 됨. (죽 제외)
    나물류(나물/무침/무생채)는 규칙10에서 따로 간격을 보므로 여기서는 제외한다."""
    seen = defaultdict(list)
    for t in ctx["timeline"]:
        for d in [t["soup"]] + t["dishes"]:
            if not d:
                continue
            if any(s in d for s in NAMUL_SUFFIX):   # 나물류는 규칙10 담당
                continue
            seen[menu_identity(d)].append((t["idx"], t["day"], t["meal"], d))
    issues = []
    for ident, lst in seen.items():
        lst.sort()
        for i in range(len(lst) - 1):
            if lst[i + 1][0] - lst[i][0] < WEEK_MEALS:
                a, b = lst[i], lst[i + 1]
                same = "같은 메뉴" if a[3] == b[3] else f"같은 메뉴({ident})"
                issues.append(f"[규칙-같은메뉴] {same} 7일 안에 2회: "
                              f"{a[1]}{a[2]}({a[3]}) -> {b[1]}{b[2]}({b[3]})")
    return issues


def check_day_gap(ctx, keyword_set, label, min_days):
    """같은 메뉴군은 최소 {min_days}일은 떨어져 있어야 한다(끼니가 아니라 '일' 단위).
    같은 날(0일) 또는 {min_days}일 미만으로 붙어 있으면 위반."""
    day_pos = {d: i for i, d in enumerate(ctx["days"])}
    occ = []  # (day_index, day, meal, dish)
    for day, meals in ctx["menu"].items():
        for m in MEAL_ORDER:
            info = meals.get(m)
            if not info:
                continue
            for d in [info.get("국")] + info.get("찬", []):
                if d and any(k in d for k in keyword_set):
                    occ.append((day_pos[day], day, m, d))
    occ.sort()
    issues = []
    for i in range(len(occ) - 1):
        dgap = occ[i + 1][0] - occ[i][0]
        if dgap < min_days:
            a, b = occ[i], occ[i + 1]
            same_day = "같은 날" if dgap == 0 else f"{dgap}일 간격"
            issues.append(f"[{label}] 최소 {min_days}일은 띄어야 함 (현재 {same_day}): "
                          f"{a[1]}{a[2]}({a[3]}) -> {b[1]}{b[2]}({b[3]})")
    return issues


# 주 N회 등으로 이미 관리되는 키워드 (분류 실패 목록에서 제외)
WEEKLY_HANDLED_KW = (MANDU_KW + KKASU_KW + JAPCHAE_KW + CURRY_KW
                     + JJAJANG_KW + DUCK_KW + SUNDAE_KW + ALTANG_KW + TTEOKGALBI_KW
                     + JWIEOCHAE_KW + DORAJI_KW + NOODLE_KW + NURUNGJI_KW)


def find_unclassified_protein_candidates(ctx):
    """분류 사전에 안 걸리는데 '고기/생선 메인요리일 수도 있는' 항목만 추린다.
    기본 채소·나물·버섯·계란·두부·묵·주N회 관리 항목 등 이미 처리되는 건 뺀다."""
    flagged = []
    for t in ctx["timeline"]:
        for d in [t["soup"]] + t["dishes"]:
            if not d:
                continue
            if not any(s in d for s in MAIN_DISH_SUFFIX):
                continue
            if classify_protein(d):           # 이미 고기/생선으로 분류됨
                continue
            if classify_mushroom(d):          # 새송이 등 버섯류(규칙7에서 관리)
                continue
            if is_filler_only(d):             # 당근/양파/야채 등 기본 채소 반찬
                continue
            if extract_tracked_side_keys(d):  # 나물/계란/두부/소시지/장조림 등(규칙10)
                continue
            if extract_muk_type(d):           # 도토리묵/청포묵/동부묵 등(묵 규칙)
                continue
            if any(k in d for k in WEEKLY_HANDLED_KW):  # 만두/까스/잡채/카레/짜장/오리/순대/알탕/떡갈비
                continue
            flagged.append((t["day"], t["meal"], d))
    return flagged


def run_all_checks(ctx):
    out = []
    out += check_rule1_soup_type_same_day(ctx)
    out += check_rule2_3_consecutive_protein(ctx)
    out += check_rule4_fish_gap(ctx)
    out += check_window_limit(ctx, DUCK_KW, "규칙5-오리")
    out += check_window_limit(ctx, SUNDAE_KW, "규칙6-순대")
    out += check_window_limit(ctx, JAPCHAE_KW, "규칙12-잡채")
    out += check_window_limit(ctx, CURRY_KW, "규칙13-카레")
    out += check_window_limit(ctx, JJAJANG_KW, "규칙14-짜장")
    out += check_window_limit(ctx, MANDU_KW, "규칙15-만두")
    out += check_window_limit(ctx, KKASU_KW, "규칙16-까스/튀김", limit=2)
    out += check_window_limit(ctx, ALTANG_KW, "규칙17-알탕/매운탕")
    out += check_window_limit(ctx, JWIEOCHAE_KW, "규칙20-쥐어채/명엽채", limit=2)
    out += check_window_limit(ctx, DORAJI_KW, "규칙21-도라지", limit=2)
    out += check_window_limit(ctx, NOODLE_KW, "규칙22-면류")
    out += check_rule7_mushroom_same_day(ctx)
    out += check_rule8_soondubu_tofu(ctx)
    out += check_rule9_soup_dish_overlap(ctx)
    out += check_rule10_side_gap(ctx)
    out += check_rule11_jangjorim_weekly(ctx)
    out += check_window_limit(ctx, NURUNGJI_KW, "규칙24-누룽지/숭늉", limit=2)
    out += check_muk_rules(ctx)
    # 일(day) 단위 간격 규칙
    out += check_day_gap(ctx, TTEOKGALBI_KW, "규칙19-떡갈비류", 2)   # 하루 걸러서만
    out += check_day_gap(ctx, NURUNGJI_KW, "규칙24-누룽지/숭늉", 3)  # 3일은 띄우기
    out += check_same_menu(ctx)
    return out


def sort_issues_by_date(issues, ctx):
    """규칙 순서가 아니라 '날짜(그리고 끼니) 순'으로 정렬한다.
    각 메시지에 들어있는 날짜 라벨(예: '07/08(수)')과 끼니(조/중/석)로 순위를 매긴다.
    날짜가 여러 개면 가장 이른 날짜를 기준으로 한다."""
    day_rank = {d: i for i, d in enumerate(ctx["days"])}

    def key(msg):
        ranks = [r for d, r in day_rank.items() if d in msg]
        dr = min(ranks) if ranks else 10 ** 6
        # 끼니: 메시지에서 가장 먼저 등장하는 조/중/석식 위치로
        positions = [msg.find(m) for m in MEAL_ORDER if m in msg]
        if positions:
            first = min(positions)
            mr = next(i for i, m in enumerate(MEAL_ORDER) if msg.find(m) == first)
        else:
            mr = 0
        return (dr, mr)

    return sorted(issues, key=key)


# ------------------------------------------------------------------
# Gradio 콜백
# ------------------------------------------------------------------

def run(files):
    try:
        paths = _to_paths(files)
        if not paths:
            return "식단표 엑셀(.xlsx) 파일을 먼저 올려 주세요."

        menu, parse_warns = parse_menu_xlsx(paths)
        if not menu:
            return ("엑셀에서 식단을 못 읽었습니다.\n"
                    "· 2행에 날짜가 가로로, A열에 조식/중식/석식이 있는 표인지 확인하세요.\n"
                    + ("\n".join(parse_warns) if parse_warns else ""))

        ctx = build_context(menu)

        lines = []
        # 인식 요약
        n_days = len(ctx["days"])
        n_meals = len(ctx["timeline"])
        lines.append(f"[인식] {n_days}일 / {n_meals}끼니")
        lines.append("[날짜] " + ", ".join(ctx["days"]))
        lines.append("")

        if parse_warns:
            lines.append("── 입력 파싱 경고 ──")
            lines.extend(parse_warns)
            lines.append("")

        issues = sort_issues_by_date(run_all_checks(ctx), ctx)
        lines.append(f"■ 규칙 위반 총 {len(issues)}건 (날짜순)")
        if issues:
            for i, msg in enumerate(issues, 1):
                lines.append(f"{i}. {msg}")
        else:
            lines.append("  (위반 없음 👍)")

        unclassified = find_unclassified_protein_candidates(ctx)
        lines.append("")
        lines.append(f"■ 분류 실패(사전 보강 확인 필요) {len(unclassified)}건")
        if unclassified:
            lines.append("  아래 항목은 재료 사전에 없어 재료 규칙 검사에서 빠졌을 수 있어요.")
            for day, meal, dish in unclassified:
                lines.append(f"  · {day} {meal}: {dish}")
        else:
            lines.append("  (없음)")

        return "\n".join(lines)
    except Exception:
        return "[처리 중 오류]\n" + traceback.format_exc()


HELP = """### 엑셀 형식
아래 형태의 **주간 식단표 엑셀(.xlsx)** 을 그대로 올리면 됩니다.
- **2행**: 날짜가 가로로 나열 (07/06, 07/07, …)
- **A열**: `조 식` / `중 식` / `석 식` 라벨
- 각 끼니 칸: `쌀밥/잡곡밥` → 국 → 반찬 여러 개 → `포기김치`, 중식엔 죽 한 줄
- **쌀밥/잡곡밥·포기김치는 자동으로 검사에서 제외**됩니다.

여러 주(週)를 한꺼번에 검사하려면, 주간 식단표 파일을 **여러 개** 올리세요.
날짜순으로 이어 붙여 연속 끼니·재료 간격까지 함께 봅니다.
"""

with gr.Blocks(title="주간 식단표 중복 검사기") as demo:
    gr.Markdown(
        "# 주간 식단표 중복·충돌 검사기\n"
        "주간 식단표 엑셀(.xlsx)을 올리고 **검사하기**를 누르면 "
        "식단 중복·충돌 규칙 위반을 걸러 줍니다. (사람이 눈으로 잡기 어려운 반복/충돌 1차 필터)"
    )
    with gr.Accordion("엑셀 형식 도움말", open=False):
        gr.Markdown(HELP)

    inp = gr.File(
        label="식단표 엑셀(.xlsx) — 여러 개 가능",
        file_count="multiple",
        file_types=[".xlsx"],
    )
    btn = gr.Button("검사하기", variant="primary")
    out = gr.Textbox(label="검사 결과", lines=25, show_copy_button=True)
    btn.click(run, inputs=inp, outputs=out)


if __name__ == "__main__":
    # HF Spaces에서는 SSR(실험적) 모드가 프록시 뒤 localhost 점검에 실패해 오류가 난다.
    # ssr_mode=False 로 끄고, 0.0.0.0:7860 에 바인딩한다.
    demo.launch(
        show_error=True,
        ssr_mode=False,
        server_name="0.0.0.0",
        server_port=7860,
    )
