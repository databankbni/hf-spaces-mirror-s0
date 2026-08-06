import re
import sys
import json
from datetime import datetime, timezone
from pathlib import Path

import fitz  # PyMuPDF
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from processors.base import BaseProcessor

ROW_GAP_THRESHOLD = 5.0
HEADER_FILL = "1F4E78"


def cluster_words_into_rows(words, min_y=None, max_y=None):
    filtered = []
    for w in words:
        x0, y0, x1, y1, text = w[:5]
        yc = (y0 + y1) / 2
        if min_y is not None and yc < min_y:
            continue
        if max_y is not None and yc > max_y:
            continue
        filtered.append((x0, y0, x1, y1, text, yc))

    filtered.sort(key=lambda t: (t[5], t[0]))

    rows = []
    current = []
    current_y = None

    for item in filtered:
        if current_y is None or abs(item[5] - current_y) <= ROW_GAP_THRESHOLD:
            current.append(item)
            current_y = item[5] if current_y is None else (current_y + item[5]) / 2
        else:
            rows.append(sorted(current, key=lambda t: t[0]))
            current = [item]
            current_y = item[5]

    if current:
        rows.append(sorted(current, key=lambda t: t[0]))

    return rows


def merge_row_words_to_phrases(row, max_gap=8):
    phrases = []
    current = None

    for x0, y0, x1, y1, text, yc in row:
        if current is None:
            current = [x0, y0, x1, y1, [text]]
            continue

        gap = x0 - current[2]
        if gap <= max_gap:
            current[2] = x1
            current[3] = max(current[3], y1)
            current[4].append(text)
        else:
            phrases.append((current[0], current[1], current[2], current[3], " ".join(current[4])))
            current = [x0, y0, x1, y1, [text]]

    if current is not None:
        phrases.append((current[0], current[1], current[2], current[3], " ".join(current[4])))

    return phrases


def clean_text_value(text):
    return text.strip()


def normalize_target_value(text):
    text = text.strip()
    if not text:
        return None
    if not text.startswith("$"):
        text = f"${text}"
    return text


def parse_notes_section(text):
    match = re.search(r"Notes:\s*(.+?)\s*Rating:", text, flags=re.I | re.S)
    return match.group(1).strip() if match else ""


def parse_adjusted_note_numbers(notes_text):
    """
    Dynamically find note numbers for EBITDA values that should be excluded.
    Current report wording:
      '(5) shown on adj. basis and excludes certain items'
    """
    pattern = r"\((\d+)\)\s+shown on adj\. basis and excludes certain items"
    matches = re.findall(pattern, notes_text, flags=re.I)
    return sorted({int(m) for m in matches})


def build_fundamental_group_specs(page):
    """
    Detect FY columns dynamically from Fundamental Data page.
    """

    words = page.get_text("words")

    fy_headers = []

    for w in words:
        x0, y0, x1, y1, text = w[:5]

        text = text.strip()

        if re.fullmatch(r"FY\d{2}[AE]", text):
            fy_headers.append({
                "label": text,
                "xc": (x0 + x1) / 2,
                "x0": x0,
                "x1": x1
            })

    fy_headers.sort(key=lambda x: x["xc"])

    if len(fy_headers) < 12:
        raise ValueError(
            f"Expected at least 12 FY columns but found {len(fy_headers)}"
        )

    return {
        "EPS": fy_headers[0:2],
        "EBITDA": fy_headers[4:6],
        "Revenue": fy_headers[8:10],
    }


def extract_valuation_table(pdf):
    """
    Extract from Valuation Data:
    - Ticker
    - Rating
    - Price Target
    """
    page = pdf[1]
    rows = cluster_words_into_rows(page.get_text("words"), min_y=160, max_y=520)

    rating_xc = 338.7
    target_xc = 367.5

    records = []

    for row in rows:
        tokens = [w[4] for w in row]
        if not tokens:
            continue

        first_token = tokens[0]
        if not re.fullmatch(r"[A-Z]{1,5}", first_token):
            continue

        ticker = first_token
        if ticker in {"SPX", "Average"}:
            continue

        rating = None
        price_target = None

        for x0, y0, x1, y1, text_value, yc in row:
            xc = (x0 + x1) / 2
            value = clean_text_value(text_value)

            if abs(xc - rating_xc) <= 18 and value in {"OP", "SP", "UP", "R"}:
                rating = value

            if abs(xc - target_xc) <= 18 and re.match(r"^\$?\d", value):
                price_target = normalize_target_value(value)

        records.append({
            "Ticker": ticker,
            "Rating": rating,
            "Price Target": price_target
        })

    df = pd.DataFrame(records, columns=["Ticker", "Rating", "Price Target"])
    return df, {"source_page": 2}


def extract_fundamental_table(pdf):
    """
    Extract from Fundamental data:
    - Ticker
    - EPS for each FY column
    - EBITDA for each FY column, excluding adjusted-note tickers
    - Revenue for each FY column

    EBITDA rule:
    If the ticker has a note number whose text says:
      'shown on adj. basis and excludes certain items'
    then EBITDA values are excluded (left blank initially, later converted to '-').
    """
    page = pdf[2]
    text = page.get_text("text")

    notes_text = parse_notes_section(text)
    adjusted_note_numbers = parse_adjusted_note_numbers(notes_text)
    group_specs = build_fundamental_group_specs(page)

    years = {metric: [s["label"] for s in specs] for metric, specs in group_specs.items()}

    rows = cluster_words_into_rows(page.get_text("words"), min_y=80, max_y=520)

    records = []
    raw_records = []
    excluded_ebitda_tickers = []

    for row in rows:
        tokens = [w[4] for w in row]
        if not tokens:
            continue

        first_token = tokens[0]
        if not re.fullmatch(r"[A-Z]{1,5}", first_token):
            continue

        ticker = first_token
        if ticker in {"SPX", "Average"}:
            continue

        notes = []
        note_tokens = [t for t in tokens if re.fullmatch(r"\([\d,]+\)", t)]
        for nt in note_tokens:
            notes.extend(int(x) for x in re.findall(r"\d+", nt))

        data_words = [w for w in row if w[0] > 200]

        row_map = {}
        for x0, y0, x1, y1, text_value, yc in data_words:
            xc = (x0 + x1) / 2
            value = clean_text_value(text_value)

            matched = False
            for metric_name, specs in group_specs.items():
                for spec in specs:
                    if abs(xc - spec["xc"]) <= 16:
                        row_map[f"{metric_name}_{spec['label']}"] = value
                        matched = True
                        break
                if matched:
                    break

        record = {"Ticker": ticker}
        raw_record = {"Ticker": ticker}

        for period in years["EPS"]:
            val = row_map.get(f"EPS_{period}")
            record[f"EPS_{period}"] = val
            raw_record[f"EPS_{period}"] = val

        if any(note_num in adjusted_note_numbers for note_num in notes):
            for period in years["EBITDA"]:
                record[f"EBITDA_{period}"] = None
                raw_record[f"EBITDA_{period}"] = row_map.get(f"EBITDA_{period}")
            excluded_ebitda_tickers.append(ticker)
        else:
            for period in years["EBITDA"]:
                val = row_map.get(f"EBITDA_{period}")
                record[f"EBITDA_{period}"] = val
                raw_record[f"EBITDA_{period}"] = val

        for period in years["Revenue"]:
            val = row_map.get(f"Revenue_{period}")
            record[f"Revenue_{period}"] = val
            raw_record[f"Revenue_{period}"] = val

        records.append(record)
        raw_records.append(raw_record)

    ordered_columns = (
        ["Ticker"]
        + [f"EPS_{y}" for y in years["EPS"]]
        + [f"EBITDA_{y}" for y in years["EBITDA"]]
        + [f"Revenue_{y}" for y in years["Revenue"]]
    )

    df = pd.DataFrame(records, columns=ordered_columns)
    raw_df = pd.DataFrame(raw_records, columns=ordered_columns)

    return df, {
        "source_page": 3,
        "notes_text": notes_text,
        "adjusted_note_numbers": adjusted_note_numbers,
        "excluded_ebitda_tickers": excluded_ebitda_tickers,
        "years": years,
        "raw_df": raw_df,
    }


def fill_blank_cells_with_dash(df):
    df = df.copy()
    df = df.fillna("-")
    df = df.replace(r"^\s*$", "-", regex=True)
    return df


def autosize_worksheet(ws):
    for col_cells in ws.columns:
        max_len = 0
        col_idx = col_cells[0].column
        for cell in col_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 12), 28)


def style_header(ws, fill_color=HEADER_FILL):
    header_fill = PatternFill("solid", fgColor=fill_color)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill


def write_dataframe_to_sheet(ws, df):
    ws.append(list(df.columns))
    for row in df.itertuples(index=False):
        ws.append(list(row))

    style_header(ws)
    ws.freeze_panes = "A2"
    autosize_worksheet(ws)


def write_metadata_sheet(ws, metadata_rows, raw_fundamental_df=None):
    ws.append(["Field", "Value"])
    for row in metadata_rows:
        ws.append(row)

    style_header(ws)
    ws.freeze_panes = "A2"
    
    if raw_fundamental_df is not None:
        ws.append([])
        ws.append(["Raw Fundamental Data"])
        
        sub_header_cell = ws.cell(row=ws.max_row, column=1)
        sub_header_cell.font = Font(bold=True)
        
        header_row_idx = ws.max_row + 1
        ws.append(list(raw_fundamental_df.columns))
        
        header_fill = PatternFill("solid", fgColor=HEADER_FILL)
        for cell in ws[header_row_idx]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            
        for row in raw_fundamental_df.itertuples(index=False):
            ws.append(list(row))

    autosize_worksheet(ws)


class RBCPerlinsPonderingsProcessor(BaseProcessor):
    PROCESSOR_NAME = "RBC Perlin's Ponderings"
    SUPPORTED_EXTENSIONS = [".pdf"]

    def process(self, filepath: str, job) -> str:
        job.message = "Initializing RBC Perlin's Ponderings extraction..."
        job.progress = 5

        pdf_path = Path(filepath)
        if not pdf_path.exists():
            raise FileNotFoundError(f"PDF not found: {pdf_path}")

        job.message = "Opening PDF..."
        job.progress = 10
        pdf = fitz.open(str(pdf_path))

        job.message = "Extracting Valuation Table..."
        job.progress = 30
        valuation_df, valuation_meta = extract_valuation_table(pdf)

        job.message = "Extracting Fundamental Table..."
        job.progress = 60
        fundamental_df, fundamental_meta = extract_fundamental_table(pdf)
        
        raw_fundamental_df = fundamental_meta["raw_df"]
        raw_fundamental_df = fill_blank_cells_with_dash(raw_fundamental_df)

        valuation_df = fill_blank_cells_with_dash(valuation_df)
        fundamental_df = fill_blank_cells_with_dash(fundamental_df)

        job.companies_found = len(fundamental_df)
        job.message = "Writing Excel report..."
        job.progress = 80

        output_name = f"RBC_Perlins_Ponderings_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        output_path = self.output_folder / output_name

        wb = Workbook()

        ws1 = wb.active
        ws1.title = "Valuation Data"
        write_dataframe_to_sheet(ws1, valuation_df)

        ws2 = wb.create_sheet("Fundamental Data")
        write_dataframe_to_sheet(ws2, fundamental_df)

        ws3 = wb.create_sheet("Metadata")
        metadata_rows = [
            ["source_file", pdf_path.name],
            ["source_file_path", str(pdf_path.resolve())],
            ["extracted_at_utc", datetime.now(timezone.utc).isoformat()],
            ["valuation_source_page", valuation_meta["source_page"]],
            ["fundamental_source_page", fundamental_meta["source_page"]],
            ["adjusted_note_numbers", json.dumps(fundamental_meta["adjusted_note_numbers"])],
            ["excluded_ebitda_tickers", json.dumps(fundamental_meta["excluded_ebitda_tickers"])],
            ["metric_years", json.dumps(fundamental_meta["years"])],
            ["notes_text", fundamental_meta["notes_text"]],
            ["valuation_row_count", len(valuation_df)],
            ["fundamental_row_count", len(fundamental_df)],
            ["output_note", "Metadata sheet contains the raw fundamental data at the bottom."]
        ]
        write_metadata_sheet(ws3, metadata_rows, raw_fundamental_df)

        wb.save(str(output_path))

        job.output_file = output_name
        job.progress = 100
        job.message = "RBC Perlin's Ponderings extraction complete."
        
        return output_name
