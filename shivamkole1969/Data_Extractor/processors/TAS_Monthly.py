"""
TAS Monthly Report Processor
Extracts data from TA Securities "Corporate Earnings Summary – Stocks Analysis" monthly reports.
"""

from __future__ import annotations

import logging
import re
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from processors.base import BaseProcessor

log = logging.getLogger("tas_monthly")

# ---- Config ----------------------------------------------------------------

REQUIRED_HEADER_TOKENS = ("Net Earnings", "EPS", "Gross Div", "PER", "Target")
YEAR_RE = re.compile(r"^(?:FY|CY)\d{2}$")
FY_YEAR_RE = re.compile(r"^FY\d{2}$")
MIN_YEAR_COLS = 3

METRIC_GROUPS = ["Net Earnings", "EPS", "EPS Growth", "PER", "Gross Div", "Div Yield"]
IDENTITY_HEADERS = {
    "Company": ("Company",), "FYE": ("FYE",),
    "Recom.": ("Recom.", "Recom"), "upside": ("upside",)
}

VALUE_RE = re.compile(r"^([<>]?\(?-?[\d,]+\.?\d*%?\)?|na|nm|n\.?a\.?|-)$", re.I)
MIN_DATA_VALUES = 5
STOP_MARKERS = ("DEFINITION OF RECOMMENDATION",)

COLUMN_SELECTION = {
    "identity": ["Sector", "Company", "FYE", "Recom.", "Share Price", "Target Price"],
    "Net Earnings": "forecast", "EPS": "forecast", "Gross Div": "all",
}

TABLE_X_MIN, TABLE_X_MAX = 45.0, 800.0
ROW_Y_TOL, HEADER_BAND, X_WORD_TOL = 3.0, 30.0, 1.5

DIVIDE_COLS = {
    "EPS FY25", "EPS FY26", "EPS FY27",
    "Gross Div FY26", "Gross Div FY27"
}


@dataclass
class Tok:
    text: str
    x0: float
    x1: float
    top: float

    @property
    def cx(self) -> float:
        return (self.x0 + self.x1) / 2


def page_tokens(page):
    return [Tok(w["text"], w["x0"], w["x1"], w["top"])
            for w in page.extract_words(x_tolerance=X_WORD_TOL, keep_blank_chars=False)
            if TABLE_X_MIN <= w["x0"] < TABLE_X_MAX]


def cluster_rows(toks, tol=ROW_Y_TOL):
    rows = []
    for t in sorted(toks, key=lambda z: (z.top, z.x0)):
        for row in rows:
            if abs(row[0].top - t.top) <= tol:
                row.append(t)
                break
        else:
            rows.append([t])
    for row in rows:
        row.sort(key=lambda z: z.cx)
    rows.sort(key=lambda r: r[0].top)
    return rows


@dataclass
class ColumnModel:
    anchors: list
    company_cx: float
    fye_cx: float
    years_by_group: dict
    fy_row_top: float

    def nearest(self, cx, exclude=("Company",)):
        best, bestd = None, 1e9
        for name, ax in self.anchors:
            if name in exclude:
                continue
            d = abs(cx - ax)
            if d < bestd:
                best, bestd = name, d
        return best


def _find_fy_row(rows):
    best, best_n = None, 0
    for row in rows:
        n = sum(1 for t in row if YEAR_RE.match(t.text))
        if n > best_n:
            best, best_n = row, n
    if not best or best_n < MIN_YEAR_COLS:
        raise ValueError("no FY/CY header row found")
    return best, best[0].top


def _header_band(rows, fy_top):
    return [t for row in rows
            if fy_top - HEADER_BAND <= row[0].top <= fy_top + ROW_Y_TOL for t in row]


def _group_centres(band, x_lo, x_hi):
    return sorted(t.cx for t in band
                  if re.match(r"^\(.+\)$", t.text) and x_lo <= t.cx <= x_hi)


def _locate_header_cx(band, names):
    return next((t.cx for t in band if t.text in names), None)


def _group_centres_by_name(band):
    by_text = {}
    for t in band:
        by_text.setdefault(t.text, []).append(t)

    def pair(a, b):
        for ta in by_text.get(a, []):
            for tb in by_text.get(b, []):
                if abs(ta.top - tb.top) <= ROW_Y_TOL and 0 < tb.x0 - ta.x1 < 12:
                    return (ta.x0 + tb.x1) / 2
        return None

    centre = {"Net Earnings": pair("Net", "Earnings"), "EPS Growth": pair("EPS", "Growth"),
              "Gross Div": pair("Gross", "Div"), "Div Yield": pair("Div", "Yield")}
    for t in by_text.get("PER", []):
        centre["PER"] = t.cx
    for t in by_text.get("EPS", []):
        if not any(abs(g.top - t.top) <= ROW_Y_TOL and 0 < g.x0 - t.x1 < 12
                   for g in by_text.get("Growth", [])):
            centre["EPS"] = t.cx
    return {k: v for k, v in centre.items() if v is not None}


def build_column_model(page):
    rows = cluster_rows(page_tokens(page))
    fy_row, fy_top = _find_fy_row(rows)
    band = _header_band(rows, fy_top)

    year_leaves = sorted(((t.text, t.cx) for t in fy_row if YEAR_RE.match(t.text)),
                         key=lambda z: z[1])
    x_lo = min(c for _, c in year_leaves)
    x_hi = max(c for _, c in year_leaves)

    gcentres = _group_centres(band, x_lo, x_hi)
    if len(gcentres) == len(METRIC_GROUPS):
        group_centre = dict(zip(METRIC_GROUPS, gcentres))
    else:
        log.warning("group-unit count %d != %d; using name fallback",
                    len(gcentres), len(METRIC_GROUPS))
        group_centre = _group_centres_by_name(band)

    anchors, years_by_group = [], {g: [] for g in METRIC_GROUPS}
    for label, cx in year_leaves:
        g = min(group_centre, key=lambda n: abs(cx - group_centre[n]))
        anchors.append((f"{g} {label}", cx))
        years_by_group[g].append(label)

    company_cx = _locate_header_cx(band, IDENTITY_HEADERS["Company"]) or (x_lo - 170)
    fye_cx = _locate_header_cx(band, IDENTITY_HEADERS["FYE"]) or (company_cx + 45)
    recom_cx = _locate_header_cx(band, IDENTITY_HEADERS["Recom."]) or (fye_cx + 20)
    anchors += [("Company", company_cx), ("FYE", fye_cx), ("Recom.", recom_cx)]

    rm_left = sorted(t.cx for t in fy_row if t.text == "(RM)" and t.cx < x_lo)
    if len(rm_left) >= 2:
        anchors += [("Share Price", rm_left[0]), ("Target Price", rm_left[1])]
    else:
        sp, tp = _locate_header_cx(band, ("Share",)), _locate_header_cx(band, ("Target",))
        if sp:
            anchors.append(("Share Price", sp))
        if tp:
            anchors.append(("Target Price", tp))

    upside_cx = _locate_header_cx(band, IDENTITY_HEADERS["upside"])
    if upside_cx:
        anchors.append(("% upside", upside_cx))

    right_units = sorted(t.cx for t in fy_row
                         if re.match(r"^\(.+\)$", t.text) and t.cx > x_hi)
    for name, cx in zip(["NTA/Shr", "Price/NTA", "Debt/Equity", "Mkt Cap"], right_units):
        anchors.append((name, cx))

    anchors.sort(key=lambda z: z[1])
    return ColumnModel(anchors, company_cx, fye_cx, years_by_group, fy_top)


@dataclass
class Record:
    sector: str = ""
    company: str = ""
    values: dict = field(default_factory=dict)


def parse_page(page, cm):
    rows = cluster_rows(page_tokens(page))
    boundary = (cm.company_cx + cm.fye_cx) / 2
    records, sector = [], ""

    for row in rows:
        if row[0].top <= cm.fy_row_top + ROW_Y_TOL:
            continue
        joined = " ".join(t.text for t in row).strip()
        if not joined:
            continue
        if any(m.lower() in joined.lower() for m in STOP_MARKERS):
            break

        data_toks = [t for t in row if t.cx >= boundary]
        if sum(1 for t in data_toks if VALUE_RE.match(t.text)) < MIN_DATA_VALUES:
            first = row[0].text
            if first.isalpha() and len(first) >= 2 and first.isupper():
                sector = joined
            continue

        company_toks = [t for t in row if t.cx < boundary]
        if not company_toks:
            continue

        rec = Record(sector, " ".join(t.text for t in company_toks).strip())
        bucket = {}
        for t in data_toks:
            bucket.setdefault(cm.nearest(t.cx), []).append(t)
        for col, ts in bucket.items():
            rec.values[col] = " ".join(z.text for z in sorted(ts, key=lambda z: z.cx))
        records.append(rec)
    return records


def page_is_summary(page):
    text = page.extract_text() or ""
    if not all(tok in text for tok in REQUIRED_HEADER_TOKENS):
        return False
    return any(sum(1 for w in line.split() if FY_YEAR_RE.match(w)) >= MIN_YEAR_COLS
               for line in text.splitlines())


def find_summary_pages(pdf):
    matched, skipped = [], []
    for i, page in enumerate(pdf.pages):
        txt = page.extract_text() or ""
        if "Corporate Earnings Summary" in txt or "Net Earnings" in txt:
            (matched if page_is_summary(page) else skipped).append(i)
    return matched, skipped


def _divide_value(val_str: str) -> str:
    s = val_str.strip()
    if not s:
        return val_str

    prefix = s[0] if s and s[0] in "<>" else ""
    if prefix: s = s[1:].strip()

    suffix = "%" if s and s[-1] == "%" else ""
    if suffix: s = s[:-1].strip()

    is_negative = False
    if s.startswith("(") and s.endswith(")"):
        is_negative, s = True, s[1:-1].strip()
    elif s.startswith("-") and len(s) > 1:
        is_negative, s = True, s[1:].strip()

    s = s.replace(",", "")

    try:
        val = float(s)
        if is_negative: val = -val
        return f"{prefix}{val / 100.0:g}{suffix}"
    except ValueError:
        return val_str


def resolve_year_columns(group, sel, years):
    if sel == "forecast":
        chosen = years[1:] if len(years) > 1 else years
    elif isinstance(sel, (list, tuple)):
        chosen = [y for y in years if y in sel]
    else:
        chosen = years
    return [f"{group} {y}" for y in chosen]


def build_output_columns(cm):
    cols = list(COLUMN_SELECTION["identity"])
    for g in METRIC_GROUPS:
        if g in COLUMN_SELECTION:
            cols += resolve_year_columns(g, COLUMN_SELECTION[g], cm.years_by_group.get(g, []))
    return cols


def record_to_row(rec, columns):
    out = {}
    for c in columns:
        if c == "Sector":
            out[c] = rec.sector
        elif c == "Company":
            out[c] = rec.company
        else:
            val = rec.values.get(c, "")
            out[c] = _divide_value(val) if c in DIVIDE_COLS else val
    return out


def write_xlsx(path, columns, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Stocks Analysis"
    ws.append(columns)

    fill = PatternFill("solid", fgColor="1F4E78")
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = fill

    for r in rows:
        ws.append([r.get(c, "") for c in columns])

    for i, col in enumerate(columns, 1):
        width = max(len(col), *(len(str(r.get(col, ""))) for r in rows)) + 2
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = min(width, 40)

    ws.freeze_panes = "A2"
    wb.save(path)


def extract(pdf_path, job=None, full=False):
    records, out_columns = [], None
    with pdfplumber.open(pdf_path) as pdf:
        matched, skipped = find_summary_pages(pdf)
        if not matched:
            raise ValueError("No stocks-summary pages found - layout may have "
                             "changed; check REQUIRED_HEADER_TOKENS.")
        log.info("Summary pages (0-indexed): %s; skipped: %s", matched, skipped)

        total_summary = len(matched)
        for idx, i in enumerate(matched):
            page = pdf.pages[i]
            cm = build_column_model(page)

            if out_columns is None:
                if full:
                    out_columns = ["Sector"] + [n for n, _ in cm.anchors]
                else:
                    out_columns = build_output_columns(cm)

            recs = parse_page(page, cm)
            log.info("page %d: %d rows, FY cols %s", i, len(recs),
                     {g: cm.years_by_group[g] for g in ("Net Earnings", "Gross Div")
                      if cm.years_by_group.get(g)})
            records.extend(recs)

            if job:
                job.progress = int(10 + (idx / max(total_summary, 1)) * 80)
                job.message = f"Parsed summary page {idx + 1} of {total_summary} ({len(records)} companies so far)..."

    return out_columns, [record_to_row(r, out_columns) for r in records]


class TASMonthlyProcessor(BaseProcessor):
    PROCESSOR_NAME = "TAS Monthly Report"
    SUPPORTED_EXTENSIONS = [".pdf"]

    def process(self, filepath: str, job) -> str:
        job.message = "Initializing TAS Monthly extraction..."
        job.progress = 5

        try:
            job.message = "Scanning PDF for summary pages..."
            job.progress = 10

            columns, rows = extract(str(filepath), job=job)

            job.companies_found = len(rows)
            job.message = f"Found {len(rows)} companies. Writing Excel report..."
            job.progress = 92

            output_name = f"TAS_Monthly_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            output_path = self.output_folder / output_name
            write_xlsx(str(output_path), columns, rows)

            job.progress = 100
            job.message = f"Successfully extracted {len(rows)} companies x {len(columns)} columns."
            job.output_file = output_name
            return output_name

        except Exception as e:
            job.status = "error"
            job.message = f"Extraction failed: {str(e)}"
            raise e
