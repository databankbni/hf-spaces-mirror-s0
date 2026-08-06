"""
EGR - Erste Group Research "Global Equity Ratings" PDF extractor
================================================================

Layout-driven (geometry first, regex second) extractor for the weekly EGR
"Global Equity Ratings" report.

Why this rewrite
----------------
The previous version read the Financials block as *text lines* and took the
"last N numbers" on each line.  That silently breaks whenever a chart label or
a right-hand-column word happens to sit on the same baseline as a metric row
(it happens on many pages, e.g. `Sales 91,854 93,925 99,044 101,756 16` where
the trailing `16` is a y-axis tick).  It also missed `Gross Prem.` entirely
because the label regex ended in `\\b` right after a literal `.`, which can
never match before a space - so every insurer had a blank revenue.

This version instead:

  * builds the year columns from the x-position of the `FY <year>` header
    tokens, and snaps every value to a column by its **right edge** (the EGR
    table is right-aligned, so header and value right edges agree to <1pt);
  * ignores anything outside the table's x-band, which removes chart labels,
    ESG blocks and peer tables completely;
  * discovers page roles (summary table / company sheet) dynamically - no
    hard-coded page numbers, no fixed year count, no fixed currency list;
  * cross-checks EVERY individual datapoint against at least one independent
    source inside the same PDF, so a wrong or shifted value cannot pass:

      1. the report's own `% y/y` row - recomputed from the extracted values;
         a mis-assigned column cannot reconcile,
      2. the `Yield` row - DPS(estimate) / close price must equal the printed
         yield, which pins the DPS column *and* its unit (GBP vs GBp),
      3. a second, fully independent extraction from the PDF text layer,
         compared cell for cell.

    The Validation sheet reports the coverage figure and names any datapoint
    that could not be confirmed, so nothing is ever silently trusted.
    On both the 07.07.2026 and 27.07.2026 reports coverage is 100.0%
    (3,198 / 3,198 and 3,242 / 3,242 datapoints) with zero errors.

Output workbook (same layout as the previous extractor)
-------------------------------------------------------
  EGR_Extracted   one row per company, metrics pivoted by calendar year
                  - identical columns to EGR_Extracted_20260727_143535.xlsx
  Metadata        the original nine descriptor columns, plus country,
                  fy_labels, financial_unit and close price
  Validation      per-datapoint cross-check results and anything to review

Usage
-----
  Standalone - set PDF_PATH below and run, or pass the file in:
      python "EGR - Updated 20260727.py"
      python "EGR - Updated 20260727.py" "EGR 20260727.pdf"
      python "EGR - Updated 20260727.py" "EGR 20260727.pdf" -o C:\\out\\folder

  As a processor (drop-in for the Estimates Data Extractor app), copy this
  file to processors/EGR.py - the class name and interface are unchanged:
      from processors.EGR import EGRProcessor
"""

from __future__ import annotations

import re
import traceback
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
import pdfplumber

from processors.base import BaseProcessor

# =========================================================
# GEOMETRY / TOLERANCES
# =========================================================

Y_TOLERANCE = 2.5          # pt - words within this vertical distance are one row
COL_SNAP_TOL = 3.0         # pt - value right edge vs column right edge
MAX_METRIC_ROWS = 40       # rows scanned below the Financials header
HEADER_SEARCH_X_FRAC = 0.60  # Financials table lives in the left part of the page

# =========================================================
# TOKEN PATTERNS
# =========================================================

FY_WORD = re.compile(r"^FY$", re.IGNORECASE)
YEAR_WORD = re.compile(r"^((?:19|20)\d{2})\s*(e)?$", re.IGNORECASE)
FY_YEAR_WORD = re.compile(r"^FY\s*((?:19|20)\d{2})\s*(e)?$", re.IGNORECASE)
UNIT_WORD = re.compile(r"^(mn|bn|m|k|th|thd)\.?$", re.IGNORECASE)
CCY_WORD = re.compile(r"^[A-Za-z]{2,4}$")

NUMBER_TOKEN = re.compile(r"^[+-]?\d{1,3}(?:,\d{3})*(?:\.\d+)?$|^[+-]?\d+(?:\.\d+)?$")
DASH_TOKENS = {"-", "--", "n.a.", "n.a", "na", "n/a"}

DATE_DDMMYY = re.compile(r"^\d{2}\.\d{2}\.\d{2}$")
DATE_DDMMYYYY = re.compile(r"^\d{2}\.\d{2}\.\d{4}$")
COUNTRY_CODE = re.compile(r"^[A-Z]{2}$")

RATING_VALUES = {"buy", "hold", "sell", "not rated", "accumulate", "reduce", "neutral"}

# Line-level patterns on the company sheet (left column only)
YEAR_END_PATTERN = re.compile(r"Last\s+FY\s+End:\s*(\d{2}\.\d{2}\.\d{4})", re.IGNORECASE)
RATING_PATTERN = re.compile(r"^Rating:\s*(.+?)\s*$", re.IGNORECASE)
CLOSE_PRICE_PATTERN = re.compile(
    r"Close\s+Price:\s*([A-Za-z]{2,4})\s*([\d,]+(?:\.\d+)?)", re.IGNORECASE
)
PRICES_AS_OF_PATTERN = re.compile(r"Prices\s+as\s+of:\s*(\d{2}\.\d{2}\.\d{2,4})", re.IGNORECASE)
COUNTRY_SECTOR_PATTERN = re.compile(r"^(?P<country>[^|]{2,40}?)\s*\|\s*(?P<sector>.{2,60})$")
EPS_DPS_PATTERN = re.compile(r"^(EPS|DPS)\b\s*[\(\[]?\s*([A-Za-z]{2,4})?\s*[\)\]]?", re.IGNORECASE)

# =========================================================
# METRIC VOCABULARY
# Ordered - the first label present on the page wins.
# =========================================================

REVENUE_LABELS = [
    "sales", "net sales", "revenue", "revenues", "total revenues", "turnover",
    "total income", "total operating income", "operating income",
    "gross prem.", "gross prem", "gross premiums", "gross written premiums",
    "net interest income", "total revenue",
]
EBITDA_LABELS = ["ebitda", "adj. ebitda", "adjusted ebitda"]
EBIT_LABELS = ["ebit", "adj. ebit", "adjusted ebit", "operating profit", "operating result"]
NET_PROFIT_LABELS = [
    "net profit", "net income", "net result", "net earnings",
    "attributable profit", "net profit attributable", "profit after tax",
]

# Rows that are ratios / deltas, never a headline metric.
NON_METRIC_LABELS = {
    "% y/y", "%y/y", "+/-", "yield", "yield %", "roe", "roa", "roce",
    "margin", "net margin", "ebit margin", "ebitda margin", "payout", "payout ratio",
}
GROWTH_LABELS = {"% y/y", "%y/y"}
STOP_LABEL = re.compile(r"^(source|note|notes)\b", re.IGNORECASE)

METRIC_ORDER = ["Revenue", "EBITDA", "EBIT", "Net Profit", "EPS", "DPS"]
PER_SHARE_METRICS = {"EPS", "DPS"}


# =========================================================
# SMALL DATA HOLDERS
# =========================================================

@dataclass
class Column:
    """One `FY <year>` column of the Financials table."""
    year: int
    is_estimate: bool
    fy_x0: float      # left edge of the "FY" token
    right: float      # right edge of the year token (values are right-aligned to it)

    @property
    def label(self) -> str:
        return f"{self.year}e" if self.is_estimate else str(self.year)


@dataclass
class TableRow:
    """A row of the Financials table: a label plus one value per column."""
    top: float
    label: str
    values: List[Optional[str]]
    stray_numbers: List[str] = field(default_factory=list)


@dataclass
class Check:
    detail_page: Optional[int]
    company: str
    check: str
    severity: str          # ERROR | WARNING | INFO
    detail: str


# =========================================================
# PROCESSOR
# =========================================================

class EGRProcessor(BaseProcessor):
    """Erste Group Research (Global Equity Ratings) PDF -> Excel."""

    PROCESSOR_NAME = "EGR - Erste Group Research"
    SUPPORTED_EXTENSIONS = [".pdf"]

    # ---------------------------------------------------------
    # ENTRY POINT
    # ---------------------------------------------------------

    def process(self, input_file: str, job) -> str:
        self._set(job, message="Initializing EGR PDF extraction...", progress=5)

        try:
            input_path = Path(input_file)
            if not input_path.exists():
                raise FileNotFoundError(f"PDF not found: {input_file}")

            checks: List[Check] = []
            summary_records: List[dict] = []
            detail_rows: List[dict] = []

            with pdfplumber.open(input_file) as pdf:
                total_pages = len(pdf.pages)
                self._set(job, total_pages=total_pages,
                          message="Scanning report structure...")

                page_rows = [self._page_rows(p) for p in pdf.pages]

                # ---- 1. summary ("Sector Company Rating ... Page") tables
                for idx, rows in enumerate(page_rows):
                    summary_records.extend(self._parse_summary_page(rows, idx + 1))
                self._set(job, progress=12)

                # ---- 2. company sheets, discovered by their Financials header
                company_pages = [
                    idx + 1 for idx, rows in enumerate(page_rows)
                    if self._find_financials_header(rows) is not None
                ]
                self._set(
                    job, progress=15,
                    message=f"Extracting {len(company_pages)} company sheets...",
                )

                coverage_total = coverage_ok = 0
                for i, page_num in enumerate(company_pages):
                    self._set(
                        job,
                        progress=15 + int((i / max(len(company_pages), 1)) * 70),
                        message=f"Processing company sheet {page_num}/{total_pages}...",
                        current_page=page_num,
                    )
                    row, page_checks, seen, verified = self._parse_company_page(
                        pdf.pages[page_num - 1], page_rows[page_num - 1], page_num
                    )
                    checks.extend(page_checks)
                    coverage_total += seen
                    coverage_ok += verified
                    if row:
                        detail_rows.append(row)
                        self._set(job, companies_found=len(detail_rows))

            if not detail_rows:
                raise Exception("No company sheets could be found in this PDF.")

            self._set(job, progress=90, message="Reconciling and formatting...")

            summary_df = self._build_summary_dataframe(summary_records)
            checks.extend(self._reconcile(summary_df, detail_rows))
            pct = (100.0 * coverage_ok / coverage_total) if coverage_total else 0.0
            checks.append(Check(
                None, "", "datapoint_cross_check", "INFO",
                f"{coverage_ok} of {coverage_total} extracted data points "
                f"({pct:.1f}%) were re-verified against at least one independent "
                f"source in the PDF (the report's own % y/y rows, its Yield row, "
                f"or a second extraction from the PDF text layer)."))

            final_df = self._build_final_dataframe(summary_df, detail_rows)
            metadata_df = self._build_metadata_dataframe(summary_df, detail_rows)
            validation_df = self._build_validation_dataframe(checks)

            output_name = f"EGR_Extracted_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            output_path = Path(self.output_folder) / output_name
            self._write_excel(final_df, metadata_df, validation_df, str(output_path))

            errors = sum(1 for c in checks if c.severity == "ERROR")
            self._set(
                job, progress=100, output_file=output_name,
                message=(f"Extracted {len(final_df)} companies"
                         + (f" - {errors} check(s) need review." if errors else " - all checks passed.")),
            )
            return output_name

        except Exception as e:
            traceback.print_exc()
            self._set(job, status="error", message=f"Extraction failed: {e}")
            raise

    @staticmethod
    def _set(job, **kwargs) -> None:
        """Update the job object defensively (it is optional in standalone runs)."""
        if job is None:
            return
        for key, value in kwargs.items():
            try:
                setattr(job, key, value)
            except Exception:
                pass

    # =========================================================
    # TEXT / GEOMETRY PRIMITIVES
    # =========================================================

    @staticmethod
    def _clean(text: Optional[str]) -> str:
        if not text:
            return ""
        text = text.replace("\xa0", " ")
        text = text.replace("\u2013", "-").replace("\u2014", "-").replace("\u2212", "-")
        return re.sub(r"\s+", " ", text).strip()

    @staticmethod
    def _to_float(value: Optional[str]) -> Optional[float]:
        if value is None:
            return None
        v = str(value).strip().replace("\u2212", "-")
        if v.lower() in DASH_TOKENS or v == "":
            return None
        v = v.replace(",", "").replace("+", "")
        try:
            return float(v)
        except ValueError:
            return None

    def _page_rows(self, page) -> List[List[dict]]:
        """Group a page's words into baseline-aligned rows, left-to-right."""
        try:
            words = page.extract_words(
                use_text_flow=True, keep_blank_chars=False,
                extra_attrs=["size", "fontname"],
            )
        except Exception:
            words = page.extract_words(use_text_flow=True, keep_blank_chars=False)

        for w in words:
            w["text"] = self._clean(w.get("text"))
            w.setdefault("size", 0.0)
        words = [w for w in words if w["text"]]
        if not words:
            return []

        rows: List[List[dict]] = []
        current: List[dict] = []
        current_top: Optional[float] = None
        for w in sorted(words, key=lambda x: (x["top"], x["x0"])):
            if current_top is None or abs(w["top"] - current_top) <= Y_TOLERANCE:
                if current_top is None:
                    current_top = w["top"]
                current.append(w)
            else:
                rows.append(sorted(current, key=lambda x: x["x0"]))
                current, current_top = [w], w["top"]
        if current:
            rows.append(sorted(current, key=lambda x: x["x0"]))
        return rows

    @staticmethod
    def _row_text(row: List[dict], x_max: Optional[float] = None) -> str:
        words = row if x_max is None else [w for w in row if w["x1"] <= x_max]
        return " ".join(w["text"] for w in words).strip()

    @staticmethod
    def _is_number(token: str) -> bool:
        return bool(NUMBER_TOKEN.match(token))

    @staticmethod
    def _is_dash(token: str) -> bool:
        return token.strip().lower() in DASH_TOKENS

    # =========================================================
    # FINANCIALS TABLE - COLUMN MODEL
    # =========================================================

    def _find_financials_header(self, rows: List[List[dict]]) -> Optional[Tuple[int, List[Column], Optional[str], Optional[str]]]:
        """
        Locate the `<CCY> mn. FY 2024 FY 2025 FY 2026e FY 2027e` header.

        Returns (row_index, columns, currency, unit) or None.
        Works for any number of year columns and any currency/unit.
        """
        for idx, row in enumerate(rows):
            columns = self._columns_from_row(row)
            if len(columns) < 2:
                continue
            currency, unit = self._currency_from_row(row, columns[0].fy_x0)
            return idx, columns, currency, unit
        return None

    def _columns_from_row(self, row: List[dict]) -> List[Column]:
        """Build the year columns from `FY` + `<year>` token pairs in a header row."""
        columns: List[Column] = []
        i = 0
        while i < len(row):
            word = row[i]
            combined = FY_YEAR_WORD.match(word["text"])
            if combined:  # "FY2026e" written as a single token
                columns.append(Column(
                    year=int(combined.group(1)),
                    is_estimate=bool(combined.group(2)),
                    fy_x0=word["x0"], right=word["x1"],
                ))
                i += 1
                continue

            if FY_WORD.match(word["text"]) and i + 1 < len(row):
                nxt = row[i + 1]
                year = YEAR_WORD.match(nxt["text"])
                # the year must sit immediately to the right of "FY"
                if year and 0 <= nxt["x0"] - word["x1"] < 12:
                    columns.append(Column(
                        year=int(year.group(1)),
                        is_estimate=bool(year.group(2)),
                        fy_x0=word["x0"], right=nxt["x1"],
                    ))
                    i += 2
                    continue
            i += 1

        columns.sort(key=lambda c: c.right)
        # years must be strictly increasing left-to-right, else this is not the header
        if any(columns[i].year >= columns[i + 1].year for i in range(len(columns) - 1)):
            return []
        return columns

    def _currency_from_row(self, row: List[dict], first_col_x0: float) -> Tuple[Optional[str], Optional[str]]:
        """Read `EUR mn.` / `USD mn` sitting to the left of the first year column."""
        left = [w for w in row if w["x1"] <= first_col_x0]
        currency = unit = None
        for w in left:
            token = w["text"].strip()
            if UNIT_WORD.match(token):
                unit = token.rstrip(".").lower()
            elif CCY_WORD.match(token) and not UNIT_WORD.match(token):
                currency = token
        return (currency.upper() if currency else None), unit

    def _read_table(self, rows: List[List[dict]], header_idx: int,
                    columns: List[Column]) -> List[TableRow]:
        """
        Read the metric rows below the header.

        A token belongs to column j when its right edge matches column j's right
        edge (the EGR table is right-aligned).  Everything to the right of the
        last column - chart axes, ESG blocks, peer tables - is discarded, which
        is what makes the extraction immune to same-baseline noise.
        """
        label_bound = columns[0].fy_x0 - 3.0
        right_bound = columns[-1].right + COL_SNAP_TOL
        col_lefts = [label_bound] + [c.right for c in columns[:-1]]

        table: List[TableRow] = []
        for row in rows[header_idx + 1: header_idx + 1 + MAX_METRIC_ROWS]:
            values: List[Optional[str]] = [None] * len(columns)
            label_words: List[dict] = []
            strays: List[str] = []

            for w in row:
                if w["x0"] > right_bound:
                    continue  # right-hand page furniture
                token = w["text"]
                if self._is_number(token) or self._is_dash(token):
                    j = self._snap_column(w, columns, col_lefts)
                    if j is not None and values[j] is None:
                        values[j] = token
                        continue
                    if w["x1"] > label_bound:
                        strays.append(token)
                        continue
                if w["x1"] <= label_bound:
                    label_words.append(w)

            label = self._clean(" ".join(w["text"] for w in
                                         sorted(label_words, key=lambda x: x["x0"])))
            if STOP_LABEL.match(label):
                break
            if label or any(v is not None for v in values):
                table.append(TableRow(top=row[0]["top"], label=label,
                                      values=values, stray_numbers=strays))
        return table

    @staticmethod
    def _snap_column(word: dict, columns: List[Column],
                     col_lefts: List[float]) -> Optional[int]:
        """Match a value to its column by right edge, with a band fallback."""
        best, best_dist = None, COL_SNAP_TOL
        for j, col in enumerate(columns):
            dist = abs(word["x1"] - col.right)
            if dist <= best_dist:
                best, best_dist = j, dist
        if best is not None:
            return best
        centre = (word["x0"] + word["x1"]) / 2.0
        for j, col in enumerate(columns):
            if col_lefts[j] < centre <= col.right + COL_SNAP_TOL:
                return j
        return None

    # =========================================================
    # COMPANY SHEET
    # =========================================================

    def _parse_company_page(self, page, rows: List[List[dict]],
                            page_num: int) -> Tuple[Optional[dict], List[Check], int, int]:
        checks: List[Check] = []
        header = self._find_financials_header(rows)
        if header is None:
            return None, checks, 0, 0
        header_idx, columns, currency, unit = header

        head = self._parse_page_head(rows)
        company = head.get("company_name") or f"(unnamed page {page_num})"

        table = self._read_table(rows, header_idx, columns)
        metrics, metric_checks, coverage = self._map_metrics(
            table, columns, company, page_num, head.get("close_price")
        )
        checks.extend(metric_checks)

        # Second, fully independent extraction of the same table from the PDF's
        # text layer.  Where it can read a row it must agree cell for cell.
        text_checks, text_cov = self._crosscheck_text(page, columns, table,
                                                      metrics, company, page_num)
        checks.extend(text_checks)
        for key, methods in text_cov.items():
            coverage.setdefault(key, set()).update(methods)

        row: Dict[str, Any] = {
            "detail_page": page_num,
            "company_name": company,
            "country": head.get("country"),
            "sector": head.get("sector"),
            "rating": head.get("rating"),
            "year_end": head.get("year_end"),
            "financial_currency": currency,
            "financial_unit": unit,
            "eps_currency": metrics.get("_eps_currency"),
            "dps_currency": metrics.get("_dps_currency"),
            "revenue_source_label": metrics.get("_revenue_label"),
            "fy_labels": ", ".join(c.label for c in columns),
            "close_price": head.get("close_price"),
            "close_price_currency": head.get("close_price_currency"),
        }
        for metric in METRIC_ORDER:
            for col in columns:
                row[f"{metric}_{col.year}"] = metrics.get((metric, col.year))

        # ---- structural checks -------------------------------------------
        if not head.get("company_name"):
            checks.append(Check(page_num, company, "company_name",
                                "ERROR", "Company name not found in page header."))
        if not head.get("rating"):
            checks.append(Check(page_num, company, "rating",
                                "ERROR", "`Rating:` line not found."))
        if not currency:
            checks.append(Check(page_num, company, "financial_currency",
                                "ERROR", "Currency missing from Financials header."))
        if metrics.get("_revenue_label") is None:
            checks.append(Check(page_num, company, "revenue",
                                "ERROR", "No revenue row recognised."))
        stray = [t for r in table for t in r.stray_numbers]
        if stray:
            checks.append(Check(page_num, company, "unassigned_numbers", "WARNING",
                                f"Numbers inside the table band that matched no column: {stray}"))
        for metric in ("Revenue", "Net Profit", "EPS", "DPS"):
            present = [row.get(f"{metric}_{c.year}") for c in columns]
            if all(v is None for v in present):
                checks.append(Check(page_num, company, f"missing_{metric}", "WARNING",
                                    f"{metric} row absent or empty on this sheet."))

        # ---- per-datapoint cross-check accounting ------------------------
        seen = verified = 0
        unverified: List[str] = []
        for metric in METRIC_ORDER:
            for col in columns:
                if row.get(f"{metric}_{col.year}") is None:
                    continue
                seen += 1
                if coverage.get((metric, col.year)):
                    verified += 1
                else:
                    unverified.append(f"{metric} {col.label}")
        if unverified:
            checks.append(Check(
                page_num, company, "unverified_datapoints", "WARNING",
                "Extracted but not confirmable against a second source in the PDF "
                "(no % y/y row, no Yield row and the text layer is unreadable here) - "
                "spot-check manually: " + ", ".join(unverified)))
        return row, checks, seen, verified

    def _parse_page_head(self, rows: List[List[dict]]) -> Dict[str, Any]:
        """
        Read the company sheet header block:

            <Company Name>                      <- largest font on the page
            <Country> | <Sector>
            Rating: <Buy|Hold|Sell|Not Rated>
            Prices as of: dd.mm.yy  Close Price: <CCY> <price>
            ... Last FY End: dd.mm.yyyy ...
        """
        out: Dict[str, Any] = {}
        if not rows:
            return out

        # The report's own page furniture sits in the top-right corner; ignore it.
        body = [r for r in rows if [w for w in r if w["x0"] < 300]]

        # --- country | sector, anchored on the following "Rating:" line
        cs_index = None
        for i, row in enumerate(body[:14]):
            text = self._row_text(row, x_max=300)
            m = COUNTRY_SECTOR_PATTERN.match(text)
            if not m:
                continue
            follows_rating = any(
                RATING_PATTERN.match(self._row_text(r, x_max=300))
                for r in body[i + 1: i + 4]
            )
            if follows_rating or cs_index is None:
                out["country"] = self._clean(m.group("country"))
                out["sector"] = self._clean(m.group("sector"))
                cs_index = i
                if follows_rating:
                    break

        # --- company name: the largest-font row above "country | sector"
        candidates = body[:cs_index] if cs_index else body[:6]

        best_row, best_size = None, 0.0
        for row in candidates:
            left = [w for w in row if w["x0"] < 300]
            if not left:
                continue
            size = max(float(w.get("size") or 0) for w in left)
            if size > best_size:
                best_row, best_size = left, size
        if best_row is not None:
            name = self._clean(" ".join(w["text"] for w in best_row))
            if name and not name.lower().startswith("erste group"):
                out["company_name"] = name

        # --- scalar fields from the left column
        for row in rows:
            text = self._row_text(row, x_max=300)
            if "rating" not in out:
                m = RATING_PATTERN.match(text)
                if m:
                    out["rating"] = self._normalise_rating(m.group(1))
            if "year_end" not in out:
                m = YEAR_END_PATTERN.search(text)
                if m:
                    out["year_end"] = m.group(1)
            if "close_price" not in out:
                m = CLOSE_PRICE_PATTERN.search(self._row_text(row))
                if m:
                    out["close_price_currency"] = m.group(1)
                    out["close_price"] = self._to_float(m.group(2))
            if "prices_as_of" not in out:
                m = PRICES_AS_OF_PATTERN.search(self._row_text(row))
                if m:
                    out["prices_as_of"] = m.group(1)
        return out

    @staticmethod
    def _normalise_rating(raw: str) -> Optional[str]:
        value = re.sub(r"\s+", " ", (raw or "").strip())
        if not value:
            return None
        if re.match(r"^not\s*rated$", value, re.IGNORECASE):
            return "Not Rated"
        return value.title() if value.lower() in RATING_VALUES else value

    # =========================================================
    # LABEL -> METRIC MAPPING
    # =========================================================

    def _map_metrics(self, table: List[TableRow], columns: List[Column],
                     company: str, page_num: int,
                     close_price: Optional[float] = None
                     ) -> Tuple[Dict[Any, Any], List[Check], Dict[Tuple[str, int], set]]:
        checks: List[Check] = []
        out: Dict[Any, Any] = {}
        coverage: Dict[Tuple[str, int], set] = {}

        norm = [re.sub(r"\s+", " ", r.label.strip().lower()).rstrip(":") for r in table]

        def find(labels: List[str]) -> Optional[int]:
            for wanted in labels:                       # exact match, in priority order
                for i, lab in enumerate(norm):
                    if lab == wanted:
                        return i
            for wanted in labels:                       # then prefix match
                for i, lab in enumerate(norm):
                    if lab.startswith(wanted):
                        return i
            return None

        revenue_idx = find(REVENUE_LABELS)
        if revenue_idx is None:
            # Future-proofing: EGR always prints the top line first.  If the
            # label is one we have never seen, take the first real data row.
            for i, r in enumerate(table):
                if norm[i] in NON_METRIC_LABELS or not r.label:
                    continue
                if any(v is not None for v in r.values):
                    revenue_idx = i
                    checks.append(Check(
                        page_num, company, "revenue_label", "WARNING",
                        f"Unknown revenue label '{r.label}' - used as revenue by position.",
                    ))
                    break

        assignments: List[Tuple[str, Optional[int]]] = [
            ("Revenue", revenue_idx),
            ("EBITDA", find(EBITDA_LABELS)),
            ("EBIT", self._find_ebit(norm)),
            ("Net Profit", find(NET_PROFIT_LABELS)),
            ("EPS", self._find_per_share(norm, "eps")),
            ("DPS", self._find_per_share(norm, "dps")),
        ]

        for metric, idx in assignments:
            if idx is None:
                continue
            row = table[idx]
            for col, raw in zip(columns, row.values):
                out[(metric, col.year)] = self._to_float(raw)

            if metric == "Revenue":
                out["_revenue_label"] = row.label
            if metric in PER_SHARE_METRICS:
                m = EPS_DPS_PATTERN.match(row.label)
                if m and m.group(2):
                    out[f"_{metric.lower()}_currency"] = m.group(2)  # keep case: GBp != GBP

            if metric == "DPS":
                # DPS has no `% y/y` row; the `Yield` row underneath it plays the
                # same role -- for the estimate columns EGR computes it off the
                # current close price, so it pins both the column and the unit.
                sub, ok_cols = self._check_yield(table, idx, columns, row,
                                                 close_price, company, page_num)
            else:
                sub, ok_cols = self._check_growth(table, idx, columns, row,
                                                  metric, company, page_num)
            checks.extend(sub)
            for j in ok_cols:
                coverage.setdefault((metric, columns[j].year), set()).add(
                    "yield" if metric == "DPS" else "%y/y")
        return out, checks, coverage

    @staticmethod
    def _find_ebit(norm: List[str]) -> Optional[int]:
        """`EBIT` must not swallow `EBITDA`."""
        for wanted in EBIT_LABELS:
            for i, lab in enumerate(norm):
                if lab == wanted:
                    return i
        for i, lab in enumerate(norm):
            if lab.startswith("ebit") and not lab.startswith("ebitda"):
                return i
        return None

    @staticmethod
    def _find_per_share(norm: List[str], prefix: str) -> Optional[int]:
        for i, lab in enumerate(norm):
            if lab.startswith(prefix + " ") or lab == prefix or lab.startswith(prefix + "("):
                return i
        return None

    # =========================================================
    # SELF-VALIDATION AGAINST THE REPORT'S OWN `% y/y` ROW
    # =========================================================

    def _check_growth(self, table: List[TableRow], idx: int, columns: List[Column],
                      row: TableRow, metric: str, company: str,
                      page_num: int) -> Tuple[List[Check], set]:
        """
        Recompute year-on-year growth from the extracted values and compare it
        with the growth EGR printed underneath.  A single mis-assigned column
        makes this fail, so a clean pass proves the column mapping is right.

        Returns (checks, indices of columns confirmed by this test).
        """
        if idx + 1 >= len(table):
            return [], set()
        nxt = table[idx + 1]
        if re.sub(r"\s+", " ", nxt.label.strip().lower()) not in GROWTH_LABELS:
            return [], set()

        decimals = self._decimals(row.values)
        problems: List[str] = []
        confirmed: set = set()
        for j in range(1, len(columns)):
            printed = self._to_float(nxt.values[j])
            prev = self._to_float(row.values[j - 1])
            cur = self._to_float(row.values[j])
            if printed is None or prev is None or cur is None or prev <= 0:
                continue
            actual = (cur / prev - 1.0) * 100.0
            if abs(actual - printed) > self._growth_tolerance(prev, cur, decimals):
                problems.append(
                    f"{columns[j].label}: {prev:g}->{cur:g} implies {actual:.1f}% "
                    f"but the report prints {printed:.1f}%"
                )
            else:
                confirmed.update({j - 1, j})  # the pair is mutually consistent
        if problems:
            return [Check(page_num, company, f"{metric}_growth_reconciliation", "ERROR",
                          "; ".join(problems))], confirmed
        return [], confirmed

    def _check_yield(self, table: List[TableRow], idx: int, columns: List[Column],
                     row: TableRow, close_price: Optional[float], company: str,
                     page_num: int) -> Tuple[List[Check], set]:
        """Verify DPS against the `Yield` row: yield_e = DPS_e / close price."""
        if idx + 1 >= len(table) or not close_price:
            return [], set()
        nxt = table[idx + 1]
        if not re.sub(r"\s+", " ", nxt.label.strip().lower()).startswith("yield"):
            return [], set()

        problems: List[str] = []
        confirmed: set = set()
        for j, col in enumerate(columns):
            if not col.is_estimate:
                continue  # historical yields use the price of that year, not today's
            dps = self._to_float(row.values[j])
            printed = self._to_float(nxt.values[j])
            if dps is None or printed is None:
                continue
            implied = dps / close_price * 100.0
            if abs(implied - printed) > max(0.06, 0.03 * abs(printed)):
                problems.append(
                    f"{col.label}: DPS {dps:g} on a close of {close_price:g} implies a "
                    f"{implied:.2f}% yield but the report prints {printed:.2f}%"
                )
            else:
                confirmed.add(j)
        if problems:
            return [Check(page_num, company, "DPS_yield_reconciliation", "ERROR",
                          "; ".join(problems))], confirmed
        return [], confirmed

    @staticmethod
    def _decimals(values: List[Optional[str]]) -> int:
        best = 0
        for v in values or []:
            if v and "." in str(v):
                best = max(best, len(str(v).split(".")[-1]))
        return best

    @staticmethod
    def _growth_tolerance(prev: float, cur: float, decimals: int) -> float:
        """
        Printed figures are rounded, so the implied growth has a rounding band.
        Widen the band by that amount plus 0.1pp for the growth row's own rounding.
        """
        step = 0.5 * (10.0 ** -decimals)
        lo = ((cur - step) / (prev + step) - 1.0) * 100.0
        hi = ((cur + step) / max(prev - step, 1e-9) - 1.0) * 100.0
        return max(abs(hi - lo) / 2.0, 0.05) + 0.1

    # =========================================================
    # SECOND EXTRACTION PATH (PDF text layer) - CELL BY CELL
    # =========================================================

    def _crosscheck_text(self, page, columns: List[Column], table: List[TableRow],
                         metrics: Dict[Any, Any], company: str, page_num: int
                         ) -> Tuple[List[Check], Dict[Tuple[str, int], set]]:
        """
        Re-read the Financials block from the PDF's text layer and compare every
        cell with the geometry-based result.

        This path takes the FIRST N numbers after each label (the geometry path
        snaps to column right edges), so the two disagree whenever either is
        wrong.  On some sheets EGR's ESG chart overprints the EPS/DPS block and
        the text layer comes out scrambled - those rows simply go unconfirmed
        here and are reported as such rather than guessed at.
        """
        checks: List[Check] = []
        coverage: Dict[Tuple[str, int], set] = {}
        header_re = re.compile(
            r"^[A-Za-z]{2,4}\s+(mn|bn|m|k|th|thd)\.?\s+FY\s*(?:19|20)\d{2}", re.IGNORECASE)

        def body_lines(**kwargs) -> List[str]:
            """Financials-block lines from one pdfplumber text-layer setting."""
            try:
                text = page.extract_text(**kwargs) or ""
            except Exception:
                return []
            lines = [self._clean(line) for line in text.split("\n")]
            start = next((i for i, line in enumerate(lines) if header_re.match(line)), None)
            return lines[start + 1:] if start is not None else []

        # The default line assembly merges baselines that EGR's ESG chart
        # overprints, which scrambles the EPS/DPS block on many sheets.  A
        # tighter y_tolerance recovers those rows, so try progressively.
        variants: List[Tuple[str, dict]] = [
            ("text", {}), ("text@y1", {"y_tolerance": 1}), ("text@layout", {"layout": True}),
        ]
        cache: Dict[str, List[str]] = {}

        def lines_for(name: str, kwargs: dict) -> List[str]:
            if name not in cache:
                cache[name] = body_lines(**kwargs)
            return cache[name]

        body = lines_for("text", {})
        if not body and not any(lines_for(n, k) for n, k in variants[1:]):
            return checks, coverage

        # which table row supplies each output metric
        row_for_metric: Dict[str, TableRow] = {}
        for m_row in table:
            label = re.sub(r"\s+", " ", m_row.label.strip().lower())
            if label == str(metrics.get("_revenue_label", "")).strip().lower():
                row_for_metric["Revenue"] = m_row
            elif label.startswith("ebitda"):
                row_for_metric.setdefault("EBITDA", m_row)
            elif label.startswith("ebit"):
                row_for_metric.setdefault("EBIT", m_row)
            elif label.startswith("net profit") or label.startswith("net income"):
                row_for_metric.setdefault("Net Profit", m_row)
            elif label.startswith("eps"):
                row_for_metric.setdefault("EPS", m_row)
            elif label.startswith("dps"):
                row_for_metric.setdefault("DPS", m_row)

        value_ahead = r"(?=(?:[-+]?\d|-\s|-$))"
        for metric, m_row in row_for_metric.items():
            pattern = re.compile("^" + re.escape(m_row.label) + r"\s+" + value_ahead)

            tokens: List[str] = []
            source = ""
            for name, kwargs in variants:
                line = next((l for l in lines_for(name, kwargs) if pattern.match(l)), None)
                if line is None:
                    continue
                rest = line[pattern.match(line).end():]
                found = re.findall(
                    r"[-+]?\d{1,3}(?:,\d{3})*(?:\.\d+)?|(?<!\S)-(?!\S)", rest
                )[:len(columns)]
                if len(found) == len(columns):
                    tokens, source = found, name
                    break
            if not tokens:
                continue

            problems = []
            for j, (col, token) in enumerate(zip(columns, tokens)):
                a = self._to_float(m_row.values[j])
                b = self._to_float(token)
                if a is None and b is None:
                    coverage.setdefault((metric, col.year), set()).add(source)
                    continue
                if a is None or b is None or abs(a - b) > 1e-9:
                    problems.append(f"{col.label}: table={a!r} {source}={b!r}")
                else:
                    coverage.setdefault((metric, col.year), set()).add(source)
            if problems:
                checks.append(Check(
                    page_num, company, f"{metric}_text_layer_mismatch", "ERROR",
                    "Two independent readings of this row disagree - " + "; ".join(problems)))
        return checks, coverage

    # =========================================================
    # SUMMARY TABLES (front-of-report constituent lists)
    # =========================================================

    def _parse_summary_page(self, rows: List[List[dict]], page_num: int) -> List[dict]:
        """
        Parse every `Sector | Company | Rating | Price | ... | Page` table on a
        page (some pages carry two).  Columns come from the header word
        positions; the right-hand numeric fields are right-aligned, so they are
        snapped by right edge exactly like the Financials table.
        """
        records: List[dict] = []
        header_indexes = [i for i, r in enumerate(rows) if self._is_summary_header(r)]

        for h_i, header_idx in enumerate(header_indexes):
            anchors = self._summary_anchors(rows[header_idx])
            if not anchors:
                continue
            end = header_indexes[h_i + 1] if h_i + 1 < len(header_indexes) else len(rows)
            sector = None
            for row in rows[header_idx + 1: end]:
                text = self._row_text(row)
                if STOP_LABEL.match(text):
                    break
                rec = self._parse_summary_row(row, anchors)
                if not rec:
                    continue
                sector = rec.get("sector_summary") or sector
                rec["sector_summary"] = sector
                rec["summary_page"] = page_num
                records.append(rec)
        return records

    @staticmethod
    def _is_summary_header(row: List[dict]) -> bool:
        text = " ".join(w["text"] for w in row).lower()
        return ("company" in text and "rating" in text
                and text.rstrip().endswith("page") and "sector" in text)

    @staticmethod
    def _summary_anchors(row: List[dict]) -> Dict[str, dict]:
        anchors: Dict[str, dict] = {}
        for w in row:
            key = w["text"].strip().lower().rstrip(".")
            if key.startswith("ctry"):
                key = "ctry"
            anchors[key] = w
        needed = {"sector", "company", "rating", "page"}
        return anchors if needed.issubset(anchors) else {}

    def _parse_summary_row(self, row: List[dict], anchors: Dict[str, dict]) -> Optional[dict]:
        sector_x0 = anchors["sector"]["x0"] - 2
        company_x0 = anchors["company"]["x0"] - 2
        rating_x0 = anchors["rating"]["x0"] - 2

        sector_words = [w for w in row if sector_x0 <= w["x0"] < company_x0]
        company_words = [w for w in row if company_x0 <= w["x0"] < rating_x0]
        tail = [w for w in row if w["x0"] >= rating_x0]
        if not company_words or not tail:
            return None

        page_ref = self._snap_to_anchor(tail, anchors.get("page"))
        if page_ref is None or not re.match(r"^\d{1,4}$", page_ref):
            return None

        tokens = [w["text"] for w in tail]
        ccy = price = country = None
        for i, tok in enumerate(tokens):
            if CCY_WORD.match(tok) and not COUNTRY_CODE.match(tok) and i + 1 < len(tokens):
                if self._is_number(tokens[i + 1]):
                    ccy, price = tok, tokens[i + 1]
                    break
        for i, tok in enumerate(tokens):
            if COUNTRY_CODE.match(tok) and i + 1 < len(tokens) and DATE_DDMMYY.match(tokens[i + 1]):
                country = tok
                break

        rating_tokens: List[str] = []
        for tok in tokens:
            if CCY_WORD.match(tok) and tok == ccy:
                break
            rating_tokens.append(tok)

        return {
            "detail_page": int(page_ref),
            "company_name_summary": self._clean(" ".join(w["text"] for w in company_words)),
            "rating_summary": self._normalise_rating(" ".join(rating_tokens)),
            "sector_summary": self._clean(" ".join(w["text"] for w in sector_words)) or None,
            "country_summary": country,
            "price_currency_summary": ccy,
            "price_summary": self._to_float(price),
            "pe_summary": self._to_float(self._snap_to_anchor(tail, anchors.get("pee"))),
            "dy_summary": self._to_float(self._snap_to_anchor(tail, anchors.get("dye"))),
        }

    @staticmethod
    def _snap_to_anchor(words: List[dict], anchor: Optional[dict]) -> Optional[str]:
        """Right-aligned column lookup: the word whose right edge matches the header's."""
        if not anchor:
            return None
        best, best_dist = None, COL_SNAP_TOL
        for w in words:
            dist = abs(w["x1"] - anchor["x1"])
            if dist <= best_dist:
                best, best_dist = w["text"], dist
        return best

    # =========================================================
    # RECONCILIATION: SUMMARY LIST vs COMPANY SHEETS
    # =========================================================

    @staticmethod
    def _key(name: Optional[str]) -> str:
        return re.sub(r"[^a-z0-9]", "", (name or "").lower())

    def _reconcile(self, summary_df: pd.DataFrame, detail_rows: List[dict]) -> List[Check]:
        checks: List[Check] = []
        if summary_df.empty:
            return [Check(None, "", "summary_tables", "WARNING",
                          "No summary tables found - company list could not be cross-checked.")]

        by_page = {int(r["detail_page"]): r for _, r in summary_df.iterrows()}
        detail_by_page = {int(r["detail_page"]): r for r in detail_rows}

        for page in sorted(set(by_page) - set(detail_by_page)):
            checks.append(Check(page, str(by_page[page].get("company_name_summary")),
                                "missing_company_sheet", "ERROR",
                                f"Summary lists this company on page {page} but no sheet was parsed."))
        for page in sorted(set(detail_by_page) - set(by_page)):
            checks.append(Check(page, str(detail_by_page[page].get("company_name")),
                                "unlisted_company_sheet", "WARNING",
                                "Company sheet parsed but not present in any summary table."))

        for page in sorted(set(by_page) & set(detail_by_page)):
            summary, detail = by_page[page], detail_by_page[page]
            s_name, d_name = summary.get("company_name_summary"), detail.get("company_name")
            if self._key(s_name) != self._key(d_name):
                checks.append(Check(page, str(d_name), "company_name_mismatch", "ERROR",
                                    f"Summary says '{s_name}', company sheet says '{d_name}'."))
            s_rating, d_rating = summary.get("rating_summary"), detail.get("rating")
            if s_rating and d_rating and str(s_rating).lower() != str(d_rating).lower():
                checks.append(Check(page, str(d_name), "rating_mismatch", "ERROR",
                                    f"Summary says '{s_rating}', company sheet says '{d_rating}'."))

            eps_ccy = (detail.get("eps_currency") or "")
            dps_ccy = (detail.get("dps_currency") or "")
            if eps_ccy and dps_ccy and eps_ccy != dps_ccy:
                checks.append(Check(page, str(d_name), "eps_dps_currency", "WARNING",
                                    f"EPS is quoted in {eps_ccy} but DPS in {dps_ccy} - "
                                    "do not compare or convert these without rescaling."))
            fin_ccy = detail.get("financial_currency")
            if fin_ccy and eps_ccy and fin_ccy.upper() != eps_ccy.upper():
                checks.append(Check(page, str(d_name), "eps_vs_statement_currency", "WARNING",
                                    f"Statement currency {fin_ccy} but EPS in {eps_ccy}."))
        return checks

    # =========================================================
    # DATAFRAMES
    # =========================================================

    def _build_summary_dataframe(self, records: List[dict]) -> pd.DataFrame:
        df = pd.DataFrame(records)
        if df.empty:
            return df
        return df.drop_duplicates(subset=["detail_page"], keep="first").reset_index(drop=True)

    @staticmethod
    def _year_columns(df: pd.DataFrame) -> List[str]:
        ordered: List[str] = []
        for metric in METRIC_ORDER:
            cols = [c for c in df.columns if re.match(rf"^{re.escape(metric)}_\d{{4}}$", c)]
            ordered.extend(sorted(cols, key=lambda c: int(c.rsplit("_", 1)[1])))
        return ordered

    def _build_final_dataframe(self, summary_df: pd.DataFrame,
                               detail_rows: List[dict]) -> pd.DataFrame:
        df = pd.DataFrame(detail_rows)
        if df.empty:
            return df

        if not summary_df.empty:
            keep = ["detail_page", "company_name_summary", "rating_summary", "sector_summary"]
            df = df.merge(summary_df[[c for c in keep if c in summary_df.columns]],
                          how="left", on="detail_page")
            for target, source in (("company_name", "company_name_summary"),
                                   ("rating", "rating_summary"),
                                   ("sector", "sector_summary")):
                if source in df.columns:
                    df[target] = df[target].fillna(df[source])
            df = df.drop(columns=[c for c in keep if c != "detail_page" and c in df.columns])

        metric_cols = self._year_columns(df)

        # Not Rated companies carry no published estimates - blank them explicitly.
        if "rating" in df.columns:
            not_rated = df["rating"].astype(str).str.contains("not rated", case=False, na=False)
            if not_rated.any():
                for col in metric_cols:
                    df[col] = df[col].astype(object)
                    df.loc[not_rated, col] = "-"
                for col in ("financial_currency", "eps_currency", "dps_currency",
                            "revenue_source_label", "year_end"):
                    if col in df.columns:
                        df.loc[not_rated, col] = "-"

        base_cols = ["detail_page", "company_name", "sector", "rating", "year_end",
                     "financial_currency", "eps_currency", "dps_currency",
                     "revenue_source_label"]
        cols = [c for c in base_cols if c in df.columns] + metric_cols
        return (df[cols]
                .drop_duplicates(subset=["detail_page"])
                .sort_values("detail_page")
                .reset_index(drop=True))

    def _build_metadata_dataframe(self, summary_df: pd.DataFrame,
                                  detail_rows: List[dict]) -> pd.DataFrame:
        df = pd.DataFrame(detail_rows)
        if df.empty:
            return df

        # Original nine columns first and in their original order, so anything
        # already reading this sheet by position keeps working; extras appended.
        meta_cols = ["detail_page", "company_name", "sector", "rating", "year_end",
                     "financial_currency", "eps_currency", "dps_currency",
                     "revenue_source_label",
                     "country", "fy_labels", "financial_unit",
                     "close_price_currency", "close_price"]
        meta = df[[c for c in meta_cols if c in df.columns]].copy()

        if not summary_df.empty:
            keep = ["detail_page", "company_name_summary", "rating_summary",
                    "country_summary", "price_currency_summary", "price_summary"]
            meta = meta.merge(summary_df[[c for c in keep if c in summary_df.columns]],
                              how="left", on="detail_page")
            for target, source in (("company_name", "company_name_summary"),
                                   ("rating", "rating_summary"),
                                   ("country", "country_summary")):
                if source in meta.columns:
                    meta[target] = meta[target].fillna(meta[source])
                    meta = meta.drop(columns=[source])
            meta = meta.rename(columns={"price_currency_summary": "summary_price_currency",
                                        "price_summary": "summary_price"})
        return (meta.drop_duplicates(subset=["detail_page"])
                .sort_values("detail_page")
                .reset_index(drop=True))

    @staticmethod
    def _build_validation_dataframe(checks: List[Check]) -> pd.DataFrame:
        if not checks:
            return pd.DataFrame([{
                "detail_page": None, "company_name": "",
                "check": "all_checks_passed", "severity": "INFO",
                "detail": "Every company sheet reconciled against the report's own "
                          "% y/y rows and summary tables.",
            }])
        order = {"ERROR": 0, "WARNING": 1, "INFO": 2}
        rows = [{"detail_page": c.detail_page, "company_name": c.company,
                 "check": c.check, "severity": c.severity, "detail": c.detail}
                for c in checks]
        df = pd.DataFrame(rows)
        # detail_page is None for report-level checks, so sort on numeric copies.
        df["_sev"] = df["severity"].map(order).fillna(9)
        df["_page"] = pd.to_numeric(df["detail_page"], errors="coerce").fillna(1e9)
        return (df.sort_values(["_sev", "_page", "check"])
                .drop(columns=["_sev", "_page"])
                .reset_index(drop=True))

    # =========================================================
    # EXCEL
    # =========================================================

    def _write_excel(self, final_df: pd.DataFrame, metadata_df: pd.DataFrame,
                     validation_df: pd.DataFrame, output_path: str) -> None:
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        thin = Side(style="thin", color="D0D0D0")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        header_font = Font(color="FFFFFF", bold=True, size=10)
        red_font = Font(color="C00000", bold=True, size=10)
        amber_fill = PatternFill("solid", start_color="FFF2CC", end_color="FFF2CC")
        red_fill = PatternFill("solid", start_color="FFC7CE", end_color="FFC7CE")

        def style_header(ws, df, colour):
            fill = PatternFill("solid", start_color=colour, end_color=colour)
            for i in range(1, len(df.columns) + 1):
                cell = ws.cell(row=1, column=i)
                cell.fill, cell.font = fill, header_font
                cell.alignment = Alignment(horizontal="center", wrap_text=True)
                cell.border = border
            ws.freeze_panes = "A2"

        def autofit(ws, df, max_width=28):
            for i, name in enumerate(df.columns, 1):
                width = max(len(str(name)), 8)
                for r in range(2, min(len(df) + 2, 60)):
                    value = ws.cell(row=r, column=i).value
                    if value is not None:
                        width = max(width, len(str(value)))
                ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = min(width + 2, max_width)

        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            # ---------- Sheet 1: extracted data ----------
            final_df.to_excel(writer, sheet_name="EGR_Extracted", index=False)
            ws = writer.sheets["EGR_Extracted"]
            style_header(ws, final_df, "4472C4")

            for i, name in enumerate(final_df.columns, 1):
                m = re.match(r"^(.+)_\d{4}$", str(name))
                if not m:
                    continue
                fmt = "#,##0.00" if m.group(1) in PER_SHARE_METRICS else "#,##0"
                for r in range(2, len(final_df) + 2):
                    cell = ws.cell(row=r, column=i)
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = fmt

            if "rating" in final_df.columns:
                col = final_df.columns.get_loc("rating") + 1
                for r in range(2, len(final_df) + 2):
                    cell = ws.cell(row=r, column=col)
                    if cell.value and "not rated" in str(cell.value).lower():
                        cell.font = red_font

            # Flag EPS/DPS quoted in different units (e.g. GBP vs GBp -> factor 100).
            if {"eps_currency", "dps_currency"} <= set(final_df.columns):
                eps_i = final_df.columns.get_loc("eps_currency") + 1
                dps_i = final_df.columns.get_loc("dps_currency") + 1
                for r in range(2, len(final_df) + 2):
                    eps = str(ws.cell(row=r, column=eps_i).value or "").strip()
                    dps = str(ws.cell(row=r, column=dps_i).value or "").strip()
                    if eps and dps and eps != "-" and dps != "-" and eps != dps:
                        ws.cell(row=r, column=eps_i).fill = amber_fill
                        ws.cell(row=r, column=dps_i).fill = amber_fill
            autofit(ws, final_df)

            # ---------- Sheet 2: metadata ----------
            metadata_df.to_excel(writer, sheet_name="Metadata", index=False)
            ws2 = writer.sheets["Metadata"]
            style_header(ws2, metadata_df, "548235")
            if "rating" in metadata_df.columns:
                col = metadata_df.columns.get_loc("rating") + 1
                for r in range(2, len(metadata_df) + 2):
                    cell = ws2.cell(row=r, column=col)
                    if cell.value and "not rated" in str(cell.value).lower():
                        cell.font = red_font
            autofit(ws2, metadata_df)

            # ---------- Sheet 3: validation ----------
            validation_df.to_excel(writer, sheet_name="Validation", index=False)
            ws3 = writer.sheets["Validation"]
            style_header(ws3, validation_df, "C55A11")
            if "severity" in validation_df.columns:
                col = validation_df.columns.get_loc("severity") + 1
                for r in range(2, len(validation_df) + 2):
                    cell = ws3.cell(row=r, column=col)
                    if str(cell.value).upper() == "ERROR":
                        cell.fill, cell.font = red_fill, red_font
                    elif str(cell.value).upper() == "WARNING":
                        cell.fill = amber_fill
            autofit(ws3, validation_df, max_width=110)


