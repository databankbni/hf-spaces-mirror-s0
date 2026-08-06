import pandas as pd
import openpyxl
import re
import time
from datetime import datetime
from processors.base import BaseProcessor

class HAYProcessor(BaseProcessor):
    PROCESSOR_NAME = "HAY Broker Report"
    SUPPORTED_EXTENSIONS = [".xlsx", ".xls"]

    def process(self, input_file: str, job) -> str:
        job.message = "Initializing HAY extraction (Excel)..."
        job.progress = 0
        
        try:
            input_path = str(input_file)
            use_pandas_fallback = False
            
            if input_path.lower().endswith('.xls'):
                use_pandas_fallback = True
            else:
                try:
                    # Load raw workbook
                    wb = openpyxl.load_workbook(input_file, data_only=True)
                    ws = wb.active
                    
                    job.message = "Reading sheet data..."
                    job.progress = 10
                    
                    # Extract headers and data rows
                    headers = [cell.value if cell.value else "" for cell in next(ws.iter_rows(min_row=1, max_row=1))]
                    data_rows = list(ws.iter_rows(min_row=2))
                    
                    # Extract cell values, including hyperlinks
                    data = []
                    for row in data_rows:
                        row_data = []
                        for cell in row:
                            if cell.hyperlink and cell.hyperlink.target:
                                row_data.append(cell.hyperlink.target)
                            else:
                                row_data.append(cell.value)
                        data.append(row_data)
                        
                    df = pd.DataFrame(data, columns=headers)
                except Exception as e:
                    # If openpyxl fails for any reason (e.g. it's an .xls file misnamed as .xlsx)
                    print(f"openpyxl failed to load {input_path}, falling back to pandas: {str(e)}")
                    use_pandas_fallback = True
            
            if use_pandas_fallback:
                job.message = "Reading sheet data (pandas fallback)..."
                job.progress = 10
                # Fallback to pandas for legacy formats since openpyxl does not support them
                try:
                    df = pd.read_excel(input_path, engine='xlrd')
                except ImportError:
                    df = pd.read_excel(input_path)
            
            job.message = "Processing metrics..."
            job.progress = 40
            
            # Convert FYE-YYMM to base fiscal year
            fye_str = pd.to_numeric(df['FYE-YYMM'], errors='coerce').fillna(0).astype(int).astype(str).str.zfill(4)
            fye_dates = pd.to_datetime(fye_str, format='%y%m', errors='coerce')
            fye_years = fye_dates.dt.year
            
            # Set up fixed columns
            out = pd.DataFrame({
                'Company Name': df['COMPANY NAME'],
                'Ticker': df['TICKER & EXCHANGE CODE'].astype(str).str.split().str[0],
                'Analyst Rating': df['Analyst Rec'],
                'Target Price (PTG)': df['PTG'],
                'Analyst Name': df['ANALYST (Full Name)'],
                'FYE': fye_dates.dt.strftime('%b-%Y'),
                'Currency': df['CURRENCY']
            })
            
            # Process metric columns
            for col in df.columns:
                col_clean = str(col).strip()
                match = re.match(r'^(EPS|SALES|CPS|FFOPS)[\s\w\-]*(FY0|FYO|FY(\d+))(?:\s*QTR(\d))?$', col_clean, flags=re.IGNORECASE)
                if not match:
                    continue
                
                metric = match.group(1).upper()
                fy_offset = 0 if match.group(2).upper() in ['FY0', 'FYO'] else int(match.group(3))
                quarter = match.group(4)
                
                values = pd.to_numeric(df[col], errors='coerce').round(3)
                
                for i in range(len(df)):
                    base_year = fye_years.iat[i]
                    if pd.isna(base_year):
                        continue
                    fiscal_year = base_year + fy_offset
                    if fiscal_year < 2024:
                        continue
                    label = f"{metric}_{fiscal_year}"
                    if quarter:
                        label += f"_Q{quarter}"
                    out.at[i, label] = values.iat[i]
            
            fixed_cols = ['Company Name', 'Ticker', 'Analyst Rating', 'Target Price (PTG)', 'Analyst Name', 'FYE', 'Currency']
            
            # Drop fully blank columns, but keep fixed columns
            cols_to_drop = [c for c in out.columns if out[c].isna().all() and c not in fixed_cols]
            out.drop(columns=cols_to_drop, inplace=True)
            
            # Sort metric columns
            def sort_key(col):
                if col in fixed_cols: return (0, col)
                m = re.match(r'(EPS|SALES|CPS|FFOPS)_(\d{4})(?:_Q(\d))?', col)
                if m:
                    metric_order = {'EPS': 0, 'SALES': 1, 'CPS': 2, 'FFOPS': 3}
                    return (10, metric_order.get(m.group(1), 999), int(m.group(2)), int(m.group(3)) if m.group(3) else 0)
                return (999, col)
            
            metric_cols = sorted([c for c in out.columns if c not in fixed_cols], key=sort_key)
            out = out[fixed_cols + metric_cols]
            
            # Fill remaining NaNs with "-" using column-wise object conversion to avoid strict dtype errors
            for col in out.columns:
                out[col] = out[col].astype(object).fillna("-")
            
            job.companies_found = len(out)
            output_name = f"HAY_Extracted_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            output_path = self.output_folder / output_name
            out.to_excel(output_path, index=False)
            
            job.progress = 100
            job.message = "Successfully generated Excel report."
            job.output_file = output_name
            return output_name
            
        except Exception as e:
            job.status = "error"
            job.message = f"Extraction failed: {str(e)}"
            raise e
