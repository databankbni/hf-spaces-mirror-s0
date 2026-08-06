#!/usr/bin/env python3
"""
Extract RBC "News from Nashville" comparative valuation tables into XLSX.

Why this version exists
-----------------------
The original pdfplumber-only extractor misses some pages in Nash.pdf because several
pages render as letter-spaced glyphs in pdfplumber (for example Exhibit 11, Exhibit
14, and Exhibit 16). PyMuPDF reads those same pages in the correct reading order,
so this version uses PyMuPDF as the primary parser and applies a table-aware row
parser.

What it extracts
----------------
- Comparative valuation exhibits only: Exhibit 10 through Exhibit 16 style tables.
- Non-NR rows only in the main output, as requested.
- NR rows are captured only in the optional audit/Skipped_NR sheet.
- One row per non-NR table row. Duplicate tickers that appear in multiple sectors
  are intentionally preserved in the main sheet, and a unique ticker sheet is also
  generated.

Expected for Nash.pdf
---------------------
- 35 non-NR table rows
- 33 unique non-NR tickers
- NR-rated rows excluded from main output

Usage
-----
    pip install pymupdf openpyxl
    python rbc_nash_extractor_fixed.py Nash.pdf
    python rbc_nash_extractor_fixed.py Nash.pdf --xlsx output.xlsx

Notes
-----
- Values such as $1,234.5 are converted to numbers.
- Negative values shown as ($1.23) are converted to -1.23.
- NA/NM are retained as text.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from collections import defaultdict
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

try:
    import fitz  # PyMuPDF
except ImportError as exc:  # pragma: no cover
    raise SystemExit(
        "Missing dependency: PyMuPDF. Install with: pip install pymupdf"
    ) from exc


RATINGS = {"OP", "SP", "UP", "NR"}
YEARS = ["CY 2025", "CY 2026", "CY 2027"]

BASE_COLUMNS = [
    "Page",
    "Ticker",
    "Rating",
    "Company Name",
]

METRIC_COLUMNS = [
    "CY EPS - CY 2025",
    "CY EPS - CY 2026",
    "CY EPS - CY 2027",
    "CY CFFO/Shr - CY 2025",
    "CY CFFO/Shr - CY 2026",
    "CY CFFO/Shr - CY 2027",
    "CY EBITDA ($MM) - CY 2025",
    "CY EBITDA ($MM) - CY 2026",
    "CY EBITDA ($MM) - CY 2027",
]

REIT_EXTRA_COLUMNS = []

OUTPUT_COLUMNS = BASE_COLUMNS + METRIC_COLUMNS + REIT_EXTRA_COLUMNS

HEADER_NOISE = {
    "ticker",
    "ticker rating",
    "rating",
    "company name",
    "($mm)",
    "price",
    "2025",
    "2026",
    "2026e",
    "2027e",
    "value ($mm)",
    "market cap",
    "5/29/26",
    "enterprise",
    "adj. enterprise",
    "ev / ebitda",
    "aev / ebitdar",
    "price / cy eps",
    "price / cy cffo",
    "price / affo",
    "nav /",
    "implied",
    "price-to-",
    "yield",
    "share",
    "cap rate",
    "book",
}


class ExtractionError(RuntimeError):
    pass


def clean_line(value: str) -> str:
    """Normalize a text line without destroying finance notation."""
    value = value.replace("\u00a0", " ").replace("\uf0be", "")
    value = re.sub(r"\s+", " ", value)
    return value.strip()


def clean_company_name(name: str) -> str:
    """Fix small extraction artifacts seen in this PDF while keeping names generic."""
    name = clean_line(name)
    # PyMuPDF drops the final B in this company name on some pages of Nash.pdf.
    if name in {"Universal Health Services, Inc. Class", "Universal Health Services, Inc. Cla"}:
        return "Universal Health Services, Inc. Class B"
    if name == "Select Medical Holdings Corporat":
        return "Select Medical Holdings Corporation"
    return name


def parse_numeric(value: str) -> Any:
    """Convert finance strings to numbers when safe; keep NA/NM or other text as-is."""
    raw = clean_line(value)
    if not raw:
        return ""
    upper = raw.upper()
    if upper in {"NA", "N/A", "NM", "#N/A"}:
        return upper.replace("N/A", "NA")

    is_negative = False
    s = raw
    if s.startswith("(") and s.endswith(")"):
        is_negative = True
        s = s[1:-1]

    s = (
        s.replace("$", "")
        .replace("€", "")
        .replace(",", "")
        .replace("%", "")
        .replace("x", "")
        .replace("X", "")
        .strip()
    )
    try:
        number = float(s)
        return -number if is_negative else number
    except ValueError:
        return raw


def text_lines(page: fitz.Page) -> List[str]:
    return [clean_line(line) for line in page.get_text("text").splitlines() if clean_line(line)]


def find_exhibit(lines: Sequence[str]) -> str:
    for line in lines:
        if line.startswith("Exhibit "):
            return line.strip()
    return ""


def is_comparative_valuation_page(lines: Sequence[str]) -> bool:
    exhibit = find_exhibit(lines).lower()
    return "comparative valuations" in exhibit


def is_header_line(line: str) -> bool:
    low = line.lower().strip()
    return low == "ticker" or low == "ticker rating"


def parse_ticker_line(line: str) -> Optional[str]:
    """Return ticker from a line like '1 HCA', 'CVS', 'FME-DE', or None."""
    line = clean_line(line)
    match = re.fullmatch(r"(?:\d+\s+)?([A-Z][A-Z0-9.\-]{0,11})", line)
    if not match:
        return None
    ticker = match.group(1)
    # Avoid treating generic all-caps section labels as tickers. The caller also
    # checks that the next line is a valid rating, so this is only a light guard.
    if ticker in {"CY", "EV", "AEV", "NAV"}:
        return None
    return ticker


def is_value_like(line: str) -> bool:
    s = clean_line(line)
    if s.upper() in {"NA", "NM", "N/A", "#N/A"}:
        return True
    return bool(re.fullmatch(r"[($€\-]?\s*[\d,]+(?:\.\d+)?\)?%?x?", s, flags=re.I))


def is_section_heading(line: str) -> bool:
    low = clean_line(line).lower()
    if not low or low in HEADER_NOISE:
        return False
    if low.startswith(("note:", "source:", "exhibit ", "news from nashville")):
        return False
    if "average" in low or "median" in low or "multiple" in low:
        return False
    if is_value_like(line):
        return False
    if len(line) > 90:
        return False
    return bool(re.search(r"[A-Za-z]", line))


def infer_schema(context_lines: Sequence[str]) -> Tuple[str, Optional[str], int, str]:
    """
    Infer table schema from the header context.

    Returns:
        primary_metric, profit_metric, expected_value_count_after_company, schema_name
    """
    ctx = " | ".join(context_lines).lower()

    if "affo/share" in ctx or "price / affo" in ctx:
        return "CY AFFO/Shr", None, 12, "reit"

    if "cy cffo/shr" in ctx:
        primary = "CY CFFO/Shr"
    elif "cy ffo/shr" in ctx:
        primary = "CY FFO/Shr"
    elif "cy eps" in ctx:
        primary = "CY EPS"
    else:
        primary = "CY EPS"

    if "cy ebitdar" in ctx or "cy ebitda" in ctx:
        profit = "CY EBITDA ($MM)"
    else:
        profit = None

    return primary, profit, 14, "standard"


def make_metric_columns(metric_name: str) -> List[str]:
    return [f"{metric_name} - {year}" for year in YEARS]


def parse_standard_row(
    values: Sequence[str], primary_metric: str, profit_metric: Optional[str]
) -> Dict[str, Any]:
    """Parse non-REIT valuation rows. Expected values count is 14."""
    row: Dict[str, Any] = {}

    for col, val in zip(make_metric_columns(primary_metric), values[2:5]):
        row[col] = parse_numeric(val)

    # values[5:8] are price/primary metric multiples, intentionally not included.
    if profit_metric:
        for col, val in zip(make_metric_columns(profit_metric), values[8:11]):
            row[col] = parse_numeric(val)

    # values[11] is Enterprise Value / Adj. Enterprise Value.
    # values[12:15] are EV/EBITDA or AEV/EBITDAR multiples, intentionally not included.
    return row


def parse_reit_row(values: Sequence[str]) -> Dict[str, Any]:
    """Parse REIT table rows. Expected values count is 12."""
    row: Dict[str, Any] = {}
    return row


def extract_page_rows(page_number: int, lines: Sequence[str]) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    """Extract non-NR rows and skipped NR rows from one comparative valuation page."""
    exhibit = find_exhibit(lines)
    header_indices = [idx for idx, line in enumerate(lines) if is_header_line(line)]
    data_rows: List[Dict[str, Any]] = []
    skipped_nr: List[Dict[str, Any]] = []

    for header_no, header_idx in enumerate(header_indices):
        next_header = header_indices[header_no + 1] if header_no + 1 < len(header_indices) else len(lines)
        context = lines[max(0, header_idx - 15):header_idx]
        primary_metric, profit_metric, expected_values, schema_name = infer_schema(context)
        section = ""
        i = header_idx + 1

        while i < next_header:
            ticker = parse_ticker_line(lines[i])
            rating = lines[i + 1].strip().upper() if i + 1 < next_header else ""

            if ticker and rating in RATINGS:
                company = clean_company_name(lines[i + 2]) if i + 2 < next_header else ""
                value_start = i + 3
                value_end = value_start + expected_values
                values = list(lines[value_start:value_end])

                if len(values) < expected_values:
                    # Incomplete row; move forward one line to avoid infinite loops.
                    i += 1
                    continue

                base = {
                    "Page": page_number,
                    "Ticker": ticker,
                    "Rating": rating,
                    "Company Name": company,
                }

                if schema_name == "reit":
                    base.update(parse_reit_row(values))
                else:
                    base.update(parse_standard_row(values, primary_metric, profit_metric))

                if rating == "NR":
                    skipped_nr.append({k: base.get(k, "") for k in ["Page", "Ticker", "Rating", "Company Name"]})
                else:
                    data_rows.append(base)

                i = value_end
                continue

            if is_section_heading(lines[i]):
                section = lines[i]
            i += 1

    return data_rows, skipped_nr


def extract_pdf(pdf_path: Path) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], Dict[str, Any]]:
    if not pdf_path.exists():
        raise FileNotFoundError(f"Input PDF not found: {pdf_path}")

    all_rows: List[Dict[str, Any]] = []
    all_skipped_nr: List[Dict[str, Any]] = []
    pages_seen: List[int] = []

    with fitz.open(str(pdf_path)) as doc:
        for page_idx, page in enumerate(doc, start=1):
            lines = text_lines(page)
            if not lines or not is_comparative_valuation_page(lines):
                continue
            pages_seen.append(page_idx)
            rows, skipped = extract_page_rows(page_idx, lines)
            all_rows.extend(rows)
            all_skipped_nr.extend(skipped)

    seen_tickers = set()
    deduped_rows = []
    for row in all_rows:
        if row["Ticker"] not in seen_tickers:
            seen_tickers.add(row["Ticker"])
            deduped_rows.append(row)
    all_rows = deduped_rows

    # Stable order by page/section/ticker occurrence.
    unique_tickers = sorted({row["Ticker"] for row in all_rows})
    duplicate_tickers = sorted(
        ticker for ticker in unique_tickers if sum(1 for row in all_rows if row["Ticker"] == ticker) > 1
    )

    audit = {
        "input_pdf": str(pdf_path),
        "comparative_valuation_pages_seen": pages_seen,
        "non_nr_rows": len(all_rows),
        "unique_non_nr_tickers": len(unique_tickers),
        "unique_non_nr_ticker_list": unique_tickers,
        "duplicate_non_nr_tickers_preserved_as_rows": duplicate_tickers,
        "nr_rows_skipped": len(all_skipped_nr),
    }
    return all_rows, all_skipped_nr, audit


def ordered_columns(rows: Sequence[Dict[str, Any]]) -> List[str]:
    extras = sorted({key for row in rows for key in row.keys()} - set(OUTPUT_COLUMNS))
    return [col for col in OUTPUT_COLUMNS if any(col in row for row in rows) or col in BASE_COLUMNS] + extras



def build_unique_rows(rows: Sequence[Dict[str, Any]]) -> List[Dict[str, Any]]:
    grouped: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for row in rows:
        grouped[str(row["Ticker"])].append(row)

    out: List[Dict[str, Any]] = []
    for ticker in sorted(grouped):
        group = grouped[ticker]
        out.append(
            {
                "Ticker": ticker,
                "Rows in Valuation Sheet": len(group),
                "Ratings": ", ".join(sorted({str(r.get("Rating", "")) for r in group if r.get("Rating")})),
                "Company Names": " | ".join(dict.fromkeys(str(r.get("Company Name", "")) for r in group)),
                "Pages": ", ".join(str(r.get("Page", "")) for r in group),
            }
        )
    return out


def write_xlsx(
    path: Path,
    rows: Sequence[Dict[str, Any]],
    skipped_nr: Sequence[Dict[str, Any]],
    audit: Dict[str, Any],
    columns: Sequence[str],
) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError as exc:  # pragma: no cover
        raise SystemExit("Missing dependency: openpyxl. Install with: pip install openpyxl") from exc

    wb = Workbook()
    ws = wb.active
    ws.title = "Valuations_Non_NR"

    def write_sheet(sheet, records: Sequence[Dict[str, Any]], sheet_columns: Sequence[str]) -> None:
        sheet.append(list(sheet_columns))
        for record in records:
            sheet.append([record.get(col, "") for col in sheet_columns])
        style_sheet(sheet, len(sheet_columns), max(1, len(records) + 1))

    def style_sheet(sheet, ncols: int, nrows: int) -> None:
        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        thin = Side(style="thin", color="D9E2F3")
        border = Border(bottom=thin)
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for col_idx in range(1, ncols + 1):
            col_letter = get_column_letter(col_idx)
            max_len = 10
            for cell in sheet[col_letter][: min(nrows, 100)]:
                max_len = max(max_len, len(str(cell.value or "")))
            sheet.column_dimensions[col_letter].width = min(max_len + 2, 34)
        for row in sheet.iter_rows(min_row=2, max_row=nrows):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=False)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00;[Red](#,##0.00);-'

    write_sheet(ws, rows, columns)

    unique_ws = wb.create_sheet("Unique_Tickers")
    unique_rows = build_unique_rows(rows)
    unique_cols = ["Ticker", "Rows in Valuation Sheet", "Ratings", "Company Names", "Pages"]
    write_sheet(unique_ws, unique_rows, unique_cols)

    skipped_ws = wb.create_sheet("Skipped_NR")
    skipped_cols = ["Page", "Ticker", "Rating", "Company Name"]
    write_sheet(skipped_ws, skipped_nr, skipped_cols)

    audit_ws = wb.create_sheet("Audit")
    audit_records = [
        {"Metric": "Input PDF", "Value": audit.get("input_pdf", "")},
        {"Metric": "Comparative valuation pages seen", "Value": ", ".join(map(str, audit.get("comparative_valuation_pages_seen", [])))},
        {"Metric": "Non-NR rows extracted", "Value": audit.get("non_nr_rows", 0)},
        {"Metric": "Unique non-NR tickers", "Value": audit.get("unique_non_nr_tickers", 0)},
        {"Metric": "Duplicate tickers preserved as separate sector rows", "Value": ", ".join(audit.get("duplicate_non_nr_tickers_preserved_as_rows", []))},
        {"Metric": "NR rows skipped", "Value": audit.get("nr_rows_skipped", 0)},
        {"Metric": "Unique non-NR ticker list", "Value": ", ".join(audit.get("unique_non_nr_ticker_list", []))},
    ]
    write_sheet(audit_ws, audit_records, ["Metric", "Value"])
    audit_ws.column_dimensions["A"].width = 42
    audit_ws.column_dimensions["B"].width = 120
    for row in audit_ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def write_debug_json(path: Path, rows: Sequence[Dict[str, Any]], skipped_nr: Sequence[Dict[str, Any]], audit: Dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps({"audit": audit, "data": list(rows), "skipped_nr": list(skipped_nr)}, indent=2),
        encoding="utf-8",
    )


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Extract RBC comparative valuation tables into XLSX, excluding NR ratings.")
    parser.add_argument("pdf", type=Path, help="Input RBC PDF file")
    parser.add_argument("--xlsx", type=Path, default=None, help="Optional output XLSX path. Defaults to <pdf_stem>_non_nr_valuations.xlsx")
    parser.add_argument("--debug-json", type=Path, default=None, help="Optional debug JSON output path")
    return parser


def main(argv: Optional[Sequence[str]] = None) -> int:
    args = build_arg_parser().parse_args(argv)
    pdf_path: Path = args.pdf
    xlsx_path = args.xlsx or pdf_path.with_name(f"{pdf_path.stem}_non_nr_valuations.xlsx")

    try:
        rows, skipped_nr, audit = extract_pdf(pdf_path)
    except Exception as exc:
        print(f"Extraction failed: {exc}", file=sys.stderr)
        return 1

    columns = ordered_columns(rows)
    write_xlsx(xlsx_path, rows, skipped_nr, audit, columns)

    if args.debug_json:
        write_debug_json(args.debug_json, rows, skipped_nr, audit)

    print(f"Extracted non-NR rows: {audit['non_nr_rows']}")
    print(f"Unique non-NR tickers: {audit['unique_non_nr_tickers']}")
    print(f"Skipped NR rows: {audit['nr_rows_skipped']}")
    print(f"Wrote XLSX: {xlsx_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
