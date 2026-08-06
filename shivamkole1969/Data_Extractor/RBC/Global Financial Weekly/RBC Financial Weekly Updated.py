"""
RBC Global Financial Weekly - PDF Data Extractor (Updated)
Dynamically detects column positions from header rows instead of
hardcoding indices, so it works reliably across different PDF dates.
"""

import re
import sys
from pathlib import Path
from datetime import datetime
import camelot
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter


RATING_TOKENS = {'O', 'SP', 'U', 'UP', 'NR', 'R',
                 'O*', 'SP*', 'U*', 'UP*', 'NR*', 'R*'}
CURRENCIES = {'USD', 'CAD', 'EUR', 'GBP', 'GBp', 'AUD', 'CHF',
              'CHF1', 'USD2', 'EUR2', 'GBP2', 'GBX', 'USD1'}
EXCLUDED_TICKERS = {'BN', 'IAG', 'Ticker'}

TARGET_COLUMNS = [
    'Name', 'Ticker', 'Inv. Rat.', 'Price Target', 'CCY',
    'EPS 2025', 'EPS 2026', 'EPS 2027', 'EPS 2028',
    'BVPS 2025', 'BVPS 2026', 'BVPS 2027', 'BVPS 2028',
    'EBITDA 2025', 'EBITDA 2026', 'EBITDA 2027', 'EBITDA 2028'
]

SECTION_EXCLUDE_KEYWORDS = [
    'Average', 'Source:', 'Exhibit', 'Canadian Banks',
    'Insurance', 'Asset Managers'
]
ALLOWED_FINANCIAL_NAMES = {
    'SLM Corporation', 'Synchrony Financial', 'Ally Financial Inc.',
    'The PNC Financial Services Group, Inc.',
    'Citizens Financial Group, Inc.', 'iA Financial Corporation Inc.',
    'Manulife Financial Corporation', 'Sun Life Financial Inc.'
}

YEAR_PAT = re.compile(r'^(20\d{2})[AE]?$')


def clean_text(x):
    if x is None:
        return ''
    x = str(x).replace('\n', ' ').strip()
    return re.sub(r'\s+', ' ', x)


def clean_num(x):
    x = clean_text(x)
    if x in {'', 'NA', 'N/A', 'nm', 'NM', 'na', 'n/a', '-'}:
        return ''
    x = re.sub(r'[\$£€]|GBp|GBX', '', x)
    x = x.replace('%', '').replace('x', '').strip()
    return x


def find_header_row(df):
    """Find the header row index that contains year labels (2025, 2026, 2027)."""
    for i, row in df.iterrows():
        vals = [clean_text(v) for v in row.tolist()]
        year_count = sum(1 for v in vals if YEAR_PAT.match(v))
        if year_count >= 6:
            return i
    return None


def find_year_groups(header_vals):
    """Find indices of consecutive year-column groups in the header row.
    Year columns come in repeating groups (e.g. 2025,2026,2027, 2025,2026,2027...).
    We split on non-year values AND on year-sequence restarts (when year decreases)."""
    groups = []
    current_group = []
    prev_year = 0
    for i, v in enumerate(header_vals):
        m = YEAR_PAT.match(v)
        if m:
            year = int(m.group(1))
            # Start new group if year sequence restarts (e.g. 2027 -> 2025)
            if current_group and year <= prev_year:
                groups.append(current_group)
                current_group = []
            current_group.append((i, v))
            prev_year = year
        else:
            if current_group:
                groups.append(current_group)
                current_group = []
            prev_year = 0
    if current_group:
        groups.append(current_group)
    return groups


def find_label_mapping(df, header_row_idx):
    """Look at the label row (1-2 rows above header) to identify which
    year-column groups correspond to EPS, BVPS, EBITDA."""
    mapping = {'eps': None, 'bvps': None, 'ebitda': None}

    for offset in [1, 2]:
        label_idx = header_row_idx - offset
        if label_idx < 0:
            continue
        label_vals = [clean_text(v) for v in df.iloc[label_idx].tolist()]
        label_joined = ' '.join(label_vals).lower()

        if 'earning' in label_joined or 'eps' in label_joined:
            for i, v in enumerate(label_vals):
                vl = v.lower()
                if 'earning' in vl or 'eps' in vl:
                    mapping['eps_label_col'] = i
                    break

        if 'actual bvps' in label_joined:
            for i, v in enumerate(label_vals):
                vl = v.lower()
                # Only match "RBC / Actual BVPS", not generic "BV/PS"
                if 'actual bvps' in vl:
                    mapping['bvps_label_col'] = i
                    break

        if 'ebitda' in label_joined:
            for i, v in enumerate(label_vals):
                vl = v.lower()
                # Match "EBITDA ($)" but not "EV/EBITDA"
                if 'ebitda' in vl and 'ev/' not in vl:
                    mapping['ebitda_label_col'] = i
                    break

    return mapping


def find_ticker_col(header_vals):
    """Find the column index that contains 'Ticker'."""
    for i, v in enumerate(header_vals):
        if 'Ticker' in v:
            return i
    return None


def detect_columns(df):
    """Dynamically detect column positions from the header rows.
    Returns a dict with: header_row, ticker_col, eps_cols, bvps_cols, ebitda_cols, pt_col."""
    header_idx = find_header_row(df)
    if header_idx is None:
        return None

    header_vals = [clean_text(v) for v in df.iloc[header_idx].tolist()]
    year_groups = find_year_groups(header_vals)

    if not year_groups:
        return None

    # EPS is always the first group of year columns
    eps_cols = [g[0] for g in year_groups[0]]  # list of col indices

    # Find Ticker column
    ticker_col = find_ticker_col(header_vals)

    # Price Target: typically 1 column after Ticker (Price), then next is Target
    # But the exact position varies. We find it by looking at the columns
    # between Ticker and the first year column.
    # Pattern: Ticker, Price, PriceTarget, YTD%, [then EPS years]
    # So PT is at ticker_col + 2 (normally)
    first_year_col = year_groups[0][0][0]
    if ticker_col is not None:
        # The columns between ticker and first year col are: Price, PriceTarget, YTD%
        gap = first_year_col - ticker_col
        if gap >= 4:
            pt_col = ticker_col + 2  # standard: Ticker, Price, PT, YTD, EPS...
        elif gap == 3:
            pt_col = ticker_col + 2  # Ticker, Price, PT, EPS... (no YTD shown separately)
        else:
            pt_col = ticker_col + 1
    else:
        pt_col = first_year_col - 2  # fallback

    # Use label row to identify BVPS and EBITDA groups
    label_info = find_label_mapping(df, header_idx)

    bvps_cols = []
    ebitda_cols = []

    bvps_label_col = label_info.get('bvps_label_col')
    ebitda_label_col = label_info.get('ebitda_label_col')

    # Match year groups to BVPS/EBITDA based on closest proximity to label column
    def find_closest_group(label_col, groups):
        best_group = None
        best_dist = float('inf')
        for group in groups:
            group_start = group[0][0]
            dist = abs(group_start - label_col)
            if dist < best_dist and dist <= 4:
                best_dist = dist
                best_group = group
        return [g[0] for g in best_group] if best_group else []

    if bvps_label_col is not None:
        bvps_cols = find_closest_group(bvps_label_col, year_groups[1:])
    if ebitda_label_col is not None:
        ebitda_cols = find_closest_group(ebitda_label_col, year_groups[1:])

    return {
        'header_row': header_idx,
        'ticker_col': ticker_col,
        'pt_col': pt_col,
        'eps_cols': eps_cols,
        'bvps_cols': bvps_cols,
        'ebitda_cols': ebitda_cols,
        'first_year_col': first_year_col,
        'header_vals': header_vals,
    }


def parse_name_ticker_rating(vals, ticker_col, first_year_col):
    """Extract name, ticker, rating, CCY, and price target from the info columns."""
    rating = ''
    ccy = ''
    name = ''
    ticker = ''

    if ticker_col is None:
        return None

    # Get ticker from the expected column
    ticker = clean_text(vals[ticker_col]) if ticker_col < len(vals) else ''

    # Info columns are everything before the ticker column
    info_parts = [clean_text(vals[i]) for i in range(ticker_col)]

    # Try to extract rating and CCY from info_parts
    # Patterns seen:
    #   ['SP', 'USD The Bank of NY...']  (2 cols, merged CCY+Name)
    #   ['O', 'EUR', 'Intesa Sanpaolo']   (3 cols, separate CCY and Name)
    #   ['SP', 'CAD', 'Bank of Montreal'] (3 cols)
    #   ['O', 'USD Arch Capital...']      (2 cols, merged)
    #   ['Rat. CCY Name', 'content']      (merged header)

    if not info_parts:
        return None

    # Flatten all info text and find rating
    all_tokens = []
    for part in info_parts:
        all_tokens.extend(part.split())

    if not all_tokens:
        return None

    # First token should be rating
    if all_tokens[0] in RATING_TOKENS:
        rating = all_tokens[0]
    else:
        return None  # Not a data row

    # Find CCY - could be second token or embedded in a longer string
    remaining_tokens = all_tokens[1:]
    if remaining_tokens and remaining_tokens[0] in CURRENCIES:
        ccy = remaining_tokens[0]
        name_tokens = remaining_tokens[1:]
    else:
        name_tokens = remaining_tokens

    # Name: remaining tokens from info columns after rating and CCY
    name = ' '.join(name_tokens)

    # If name is empty, try getting it from the column right before ticker
    if not name and ticker_col >= 2:
        candidate = clean_text(vals[ticker_col - 1])
        if candidate and candidate not in CURRENCIES and candidate not in RATING_TOKENS:
            name = candidate

    # Shift detection: if ticker looks numeric, it might actually be a price
    if ticker and re.match(r'^-?[\d.,\$]+$', ticker):
        # Look for the real ticker in the column before
        for col_idx in range(ticker_col - 1, -1, -1):
            candidate = clean_text(vals[col_idx])
            # A ticker is typically short, uppercase
            if candidate and len(candidate) <= 8 and re.match(r'^[A-Z][A-Z0-9./]*$', candidate):
                # Reconstruct: real ticker found, adjust
                old_ticker_val = ticker
                ticker = candidate
                # Name is everything between rating/CCY cols and the real ticker col
                name_parts = []
                for ni in range(col_idx):
                    part = clean_text(vals[ni])
                    tokens = part.split()
                    for t in tokens:
                        if t not in RATING_TOKENS and t not in CURRENCIES:
                            name_parts.append(t)
                if name_parts:
                    name = ' '.join(name_parts)
                break

    return rating, ccy, name, ticker


def parse_row(vals, col_config):
    """Parse a single data row using dynamically detected column positions."""
    ticker_col = col_config['ticker_col']
    pt_col = col_config['pt_col']
    eps_cols = col_config['eps_cols']
    bvps_cols = col_config['bvps_cols']
    ebitda_cols = col_config['ebitda_cols']
    first_year_col = col_config['first_year_col']

    result = parse_name_ticker_rating(vals, ticker_col, first_year_col)
    if result is None:
        return None

    rating, ccy, name, ticker = result

    if not rating or not ticker or not name:
        return None

    if ticker in EXCLUDED_TICKERS:
        return None

    # Exclude section headers/labels
    if 'Financials' in name and name not in ALLOWED_FINANCIAL_NAMES:
        if not ticker:
            return None
    for kw in SECTION_EXCLUDE_KEYWORDS:
        if kw in name:
            if name not in ALLOWED_FINANCIAL_NAMES and ticker:
                pass  # keep it if it has a valid ticker
            elif not ticker:
                return None

    out = {col: '' for col in TARGET_COLUMNS}
    out['Name'] = name
    out['Ticker'] = ticker
    out['Inv. Rat.'] = rating
    out['CCY'] = ccy

    # Price Target
    if pt_col < len(vals):
        out['Price Target'] = clean_num(vals[pt_col])

    # EPS
    for i, col_idx in enumerate(eps_cols[:3]):
        year_key = f'EPS 202{5 + i}'
        if year_key in out and col_idx < len(vals):
            out[year_key] = clean_num(vals[col_idx])

    # BVPS
    for i, col_idx in enumerate(bvps_cols[:3]):
        year_key = f'BVPS 202{5 + i}'
        if year_key in out and col_idx < len(vals):
            out[year_key] = clean_num(vals[col_idx])

    # EBITDA
    for i, col_idx in enumerate(ebitda_cols[:3]):
        year_key = f'EBITDA 202{5 + i}'
        if year_key in out and col_idx < len(vals):
            out[year_key] = clean_num(vals[col_idx])

    # Replace blanks with "-"
    for col in TARGET_COLUMNS:
        if not out[col]:
            out[col] = '-'

    return out


def dedupe_rows(rows):
    seen = set()
    out = []
    for r in rows:
        key = tuple(r[c] for c in TARGET_COLUMNS)
        if key not in seen:
            seen.add(key)
            out.append(r)
    return out


def get_page_count(pdf_path):
    """Get total number of pages in the PDF."""
    from pypdf import PdfReader
    reader = PdfReader(pdf_path)
    return len(reader.pages)


def extract(pdf_path):
    """Extract data from all data pages of the PDF."""
    all_rows = []
    total_pages = get_page_count(pdf_path)

    for page in range(2, total_pages + 1):
        try:
            tables = camelot.read_pdf(str(pdf_path), pages=str(page), flavor='stream')
        except Exception:
            continue

        if tables.n == 0:
            continue

        df = tables[0].df

        # Detect columns dynamically
        col_config = detect_columns(df)
        if col_config is None:
            continue  # Not a data table page (e.g., disclosures)

        header_row = col_config['header_row']

        # Process data rows (after header + any sub-header/section rows)
        for row_idx in range(header_row + 1, len(df)):
            vals = [clean_text(v) for v in df.iloc[row_idx].tolist()]

            # Skip empty rows
            if all(v == '' for v in vals):
                continue

            parsed = parse_row(vals, col_config)
            if parsed:
                all_rows.append(parsed)

    return dedupe_rows(all_rows)


def write_excel(rows, output_path):
    """Write extracted data to a formatted Excel file."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Wide Data'
    ws.append(TARGET_COLUMNS)
    for r in rows:
        ws.append([r[c] for c in TARGET_COLUMNS])

    header_fill = PatternFill('solid', fgColor='1F4E78')
    header_font = Font(color='FFFFFF', bold=True)
    thin = Side(style='thin', color='D9E2F3')
    border = Border(bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    widths = {
        'A': 34, 'B': 12, 'C': 12, 'D': 14, 'E': 10,
        'F': 12, 'G': 12, 'H': 12, 'I': 12,
        'J': 12, 'K': 12, 'L': 12, 'M': 12,
        'N': 14, 'O': 14, 'P': 14, 'Q': 14,
    }
    for col, width in widths.items():
        if col in ws.column_dimensions:
            ws.column_dimensions[col].width = width

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    ws.sheet_view.showGridLines = False

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical='center')

    wb.save(output_path)


def main():
    if len(sys.argv) < 2:
        # Default: look for PDF files in current directory
        pdf_files = list(Path('.').glob('PDF*.pdf'))
        if not pdf_files:
            print("Usage: python 'RBC Financial Weekly Updated.py' <pdf_path>")
            print("No PDF files found in current directory.")
            sys.exit(1)
        pdf_path = pdf_files[0]
        print(f"Auto-detected PDF: {pdf_path}")
    else:
        pdf_path = Path(sys.argv[1])

    if not pdf_path.exists():
        print(f"Error: File not found: {pdf_path}")
        sys.exit(1)

    print(f"Processing: {pdf_path}")
    rows = extract(str(pdf_path))
    print(f"Extracted {len(rows)} companies")

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_name = f"RBC_Global_Weekly_{timestamp}.xlsx"
    output_path = Path('.') / output_name

    write_excel(rows, str(output_path))
    print(f"Output: {output_path}")


if __name__ == '__main__':
    main()
